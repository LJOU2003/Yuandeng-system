import os
import time
from datetime import datetime, date, timedelta, timezone  # FIX: 加 timezone 供台灣時區轉換

import streamlit as st
from dotenv import load_dotenv
from notion_client import Client
import calendar
import textwrap
import math
import streamlit.components.v1 as components
from streamlit_js_eval import get_geolocation
import bcrypt
import re


# =========================
# 0) 讀取環境變數 (Notion Token / DB ID)
# =========================
load_dotenv()

def _get_cfg(key: str, default=None):
    """優先讀取 Streamlit Cloud 的 st.secrets，其次讀取環境變數；都沒有則回傳 default。
    ✅ 兼容：大小寫不同的 key（例如 secrets 用 notion_token / NOTION_TOKEN）
    """
    keys_to_try = [key, str(key).upper(), str(key).lower()]
    # 1) Streamlit Secrets
    try:
        if hasattr(st, "secrets"):
            for k in keys_to_try:
                if k in st.secrets:
                    return st.secrets[k]
                # 有些人會放在 [general] 或其他 section 內（st.secrets 會是 dict-like）
                try:
                    v = st.secrets.get(k, None)  # type: ignore[attr-defined]
                    if v is not None:
                        return v
                except Exception:
                    pass
            # 掃描一層巢狀（避免 secrets.toml 分段）
            try:
                for _, section in dict(st.secrets).items():
                    if isinstance(section, dict):
                        for k in keys_to_try:
                            if k in section:
                                return section[k]
            except Exception:
                pass
    except Exception:
        pass

    # 2) Environment Variables
    for k in keys_to_try:
        v = os.getenv(k)
        if v is not None:
            return v
    return default

NOTION_TOKEN = _get_cfg("NOTION_TOKEN")
ACCOUNT_DB_ID = _get_cfg("ACCOUNT_DB_ID")
LEAVE_DB_ID = _get_cfg("LEAVE_DB_ID")
VACATION_DB_ID = _get_cfg("VACATION_DB_ID")
SALARY_DB_ID = _get_cfg("SALARY_DB_ID")  # ✅ 薪資計算表
OPLOG_DB_ID = _get_cfg("OPLOG_DB_ID") or _get_cfg("OP_LOG_DB_ID") or _get_cfg("OPERATION_LOG_DB_ID")  # ✅ 操作記錄表
CASHOUT_RULE_DB_ID = _get_cfg("CASHOUT_RULE_DB_ID")
ANNOUNCE_DB_ID = _get_cfg("ANNOUNCE_DB_ID")  # ✅ 公告紀錄表
PUNCH_DB_ID = _get_cfg("PUNCH_DB_ID")
COMPANY_LAT = float(_get_cfg("COMPANY_LAT", "0") or 0)
COMPANY_LON = float(_get_cfg("COMPANY_LON", "0") or 0)
COMPANY_RADIUS_M = float(_get_cfg("COMPANY_RADIUS_M", "100") or 100)
DUTY_DB_ID = _get_cfg("DUTY_DB_ID")                 # ✅ 值班排班表（直式記錄）
OVERTIME_RULE_DB_ID = _get_cfg("OVERTIME_RULE_DB_ID") # ✅ 加班設定表
OVERTIME_COUNT_DB_ID = _get_cfg("OVERTIME_COUNT_DB_ID") # ✅ 加班次數表



# 🍱 LUNCH：午餐訂餐表
LUNCH_DB_ID = _get_cfg("LUNCH_DB_ID")        # ✅ 午餐訂餐表

# ✅ 出勤記錄表
ATTEND_DB_ID = _get_cfg("ATTEND_DB_ID")      # ✅ 出勤記錄表

# （保留變數：目前午餐不再依它計算，但不移除）
HOLIDAY_DB_ID = _get_cfg("HOLIDAY_DB_ID")    # ✅ 放假/行事曆表（可選）

if not NOTION_TOKEN:
    raise RuntimeError("❌ 請先在 .env 設定 NOTION_TOKEN")
if not ACCOUNT_DB_ID:
    raise RuntimeError("❌ 請先在 .env 設定 ACCOUNT_DB_ID（帳號管理表 Database ID）")
if not LEAVE_DB_ID:
    raise RuntimeError("❌ 請先在 .env 設定 LEAVE_DB_ID（請假紀錄表 Database ID）")
if not VACATION_DB_ID:
    raise RuntimeError("❌ 請先在 .env 設定 VACATION_DB_ID（年度特休表 Database ID）")
if not SALARY_DB_ID:
    raise RuntimeError("❌ 請先在 .env 設定 SALARY_DB_ID（薪資計算表 Database ID）")

notion = Client(auth=NOTION_TOKEN)

# =========================
# ✅ 表格欄位清理（員工視角不顯示建立/更新時間）
# =========================
META_COLUMNS = {"建立時間", "最後更新時間"}

def strip_meta_columns(rows: list[dict] | None) -> list[dict]:
    """移除員工不應看到的系統欄位（建立時間 / 最後更新時間）。"""
    if not rows:
        return []
    cleaned: list[dict] = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        cleaned.append({k: v for k, v in r.items() if k not in META_COLUMNS})
    return cleaned


# =========================
# ✅ 折算規則（預設值）
# =========================
DEFAULT_HOURS_PER_DAY = 8.0
DEFAULT_CASHOUT_CAP_DAYS = 5.0
DEFAULT_CASHOUT_AMOUNT_PER_DAY = 1000.0
DEFAULT_CASHOUT_WHOLE_DAYS_ONLY = True

# 🍱 LUNCH：午餐規則
LUNCH_ALLOWANCE_PER_DAY = 90  # 一天 90 元
# 工作日：週一(0)~週六(5)，週日(6)不算
WORKDAY_WEEKDAYS = {0, 1, 2, 3, 4, 5}

# ✅ 出勤狀態（你指定：出席/請假/遲到）
ATTEND_PRESENT_STATUS = "出席"
ATTEND_LEAVE_STATUS = "請假"
ATTEND_LATE_STATUS = "遲到"

# ✅ 工餐規則：出席、遲到 都算 90；請假不算
ATTEND_LUNCH_ELIGIBLE_STATUSES = {ATTEND_PRESENT_STATUS, ATTEND_LATE_STATUS}


# =========================
# 2) 工具：讀資料庫欄位 / Select 選項
# =========================
def _rt_get_first_plain_text(prop: dict) -> str:
    """Notion rich_text 取第一段 plain_text"""
    rt = (prop or {}).get("rich_text", []) or []
    return (rt[0].get("plain_text") or "").strip() if rt else ""



def _get_prop_plain_text(prop: dict) -> str:
    """更通用的 Notion 文字讀取：支援 title / rich_text / select / multi_select / number / checkbox.
    ✅ 重點：title/rich_text 可能被切成多段（例如 bcrypt hash），必須把所有片段串起來。
    """
    if not prop:
        return ""

    # title / rich_text（把所有片段串起來，避免長字串被截斷）
    if "title" in prop:
        arr = prop.get("title") or []
        return "".join([(x.get("plain_text") or "") for x in arr]).strip() if arr else ""
    if "rich_text" in prop:
        arr = prop.get("rich_text") or []
        return "".join([(x.get("plain_text") or "") for x in arr]).strip() if arr else ""

    # select / status / multi_select
    if "select" in prop and prop.get("select"):
        return (prop["select"].get("name") or "").strip()
    if "status" in prop and prop.get("status"):
        return (prop["status"].get("name") or "").strip()
    if "multi_select" in prop and prop.get("multi_select"):
        ms = prop.get("multi_select") or []
        return ", ".join([(x.get("name") or "").strip() for x in ms if x.get("name")])

    # number / checkbox
    if "number" in prop and prop.get("number") is not None:
        return str(prop.get("number"))
    if "checkbox" in prop and prop.get("checkbox") is not None:
        return "True" if prop.get("checkbox") else "False"

    return ""
def _build_notion_prop_value(db_id: str, props_meta: dict, prop_name: str, value):
    """依據資料庫欄位型態，自動組出 Notion API properties payload；不匹配就回傳 None（略過該欄位）。"""
    meta = (props_meta or {}).get(prop_name, {}) or {}
    ptype = meta.get("type")
    if value is None:
        value = ""
    if isinstance(value, str):
        value = value.strip()
    # 文字類
    if ptype == "title":
        return {"title": [{"text": {"content": value or "—"}}]}
    if ptype == "rich_text":
        return {"rich_text": [{"text": {"content": value}}]} if value else {"rich_text": []}
    if ptype in ("email", "url", "phone_number"):
        return {ptype: value} if value else {ptype: None}
    # 選單類
    if ptype == "select":
        if not value:
            return None
        options = get_select_options(db_id, prop_name) or []
        if value in options:
            return {"select": {"name": value}}
        # 若選項不存在：改用第一個選項（避免整筆寫入失敗）
        if options:
            return {"select": {"name": options[0]}}
        return None
    if ptype == "multi_select":
        if not value:
            return None
        # 支援以逗號分隔
        if isinstance(value, str):
            vals = [v.strip() for v in value.split(",") if v.strip()]
        else:
            vals = list(value) if isinstance(value, (list, tuple, set)) else []
        options = set(get_select_options(db_id, prop_name) or [])
        payload = [{"name": v} for v in vals if (not options) or (v in options)]
        return {"multi_select": payload} if payload else None
    # 日期
    if ptype == "date":
        # value 可傳 datetime / ISO string
        if isinstance(value, datetime):
            start = value.isoformat()
        else:
            start = str(value).strip()
        return {"date": {"start": start}} if start else None
    # 數值 / 勾選
    if ptype == "number":
        try:
            return {"number": float(value)} if str(value).strip() != "" else None
        except Exception:
            return None
    if ptype == "checkbox":
        return {"checkbox": bool(value)}
    return None


def _title_get_first_plain_text(prop: dict) -> str:
    """Notion title 取第一段 plain_text"""
    t = (prop or {}).get("title", []) or []
    return (t[0].get("plain_text") or "").strip() if t else ""


def hash_password_bcrypt(plain: str) -> str:
    plain = (plain or "").encode("utf-8")
    salt = bcrypt.gensalt(rounds=12)
    return bcrypt.hashpw(plain, salt).decode("utf-8")


def verify_password_bcrypt(plain: str, hashed: str) -> bool:
    if not plain or not hashed:
        return False
    try:
        # ✅ Notion 的 rich_text / title 有時會把長字串切段或夾雜換行、空白
        #    雲端部署時最常見的就是 login_hash 讀出來含有 \n / 空白，導致 bcrypt 驗證永遠失敗
        cleaned = re.sub(r"\s+", "", str(hashed))
        # 只保留 bcrypt hash 允許的字元（避免 zero-width/奇怪符號造成雲端驗證失敗）
        cleaned = re.sub(r"[^0-9A-Za-z./$]", "", cleaned)
        return bcrypt.checkpw(plain.encode("utf-8"), cleaned.encode("utf-8"))
    except Exception:
        return False


def get_account_page_by_username(username: str) -> dict | None:
    """用員工姓名找帳號管理表那一筆 page（不依賴 schema；依序嘗試 title / rich_text）"""
    username = (username or "").strip()
    if not username:
        return None

    # ✅ 雲端偶爾會因為 schema 讀取失敗而導致查不到帳號（進而「帳號或密碼錯誤」）
    #   這裡改成「不依賴 notion.databases.retrieve」，直接嘗試兩種常見型態的 filter。
    try:
        res = notion.databases.query(
            database_id=ACCOUNT_DB_ID,
            filter={"property": "員工姓名", "title": {"equals": username}},
            page_size=1,
        )
        results = res.get("results", [])
        if results:
            return results[0]
    except Exception:
        pass

    try:
        res = notion.databases.query(
            database_id=ACCOUNT_DB_ID,
            filter={"property": "員工姓名", "rich_text": {"equals": username}},
            page_size=1,
        )
        results = res.get("results", [])
        return results[0] if results else None
    except Exception:
        return None

@st.cache_data(ttl=60)
def get_db_properties(database_id: str) -> dict:
    try:
        db = notion.databases.retrieve(database_id=database_id)
        return db.get("properties", {}) or {}
    except Exception as e:
        # ✅ 佈署到 Streamlit Cloud 時，如果 secrets/token/權限或 DB_ID 有問題，這裡會失敗
        #    開啟 DEBUG_NOTION=1 才顯示錯誤，避免一般使用者看到內部訊息
        if os.getenv("DEBUG_NOTION", "").strip() == "1":
            st.error(f"❌ Notion 讀取資料庫欄位失敗（{database_id}）：{e}")
        return {}



@st.cache_data(ttl=60)
def get_select_options(database_id: str, property_name: str) -> list[str]:
    try:
        props = get_db_properties(database_id)
        prop = props.get(property_name, {})
        if prop.get("type") != "select":
            return []
        options = prop["select"].get("options", [])
        return [o.get("name") for o in options if o.get("name")]
    except Exception as e:
        st.error(f"讀取 Notion 選項失敗（{property_name}）：{e}")
        return []


def _first_title_prop_name(props_meta: dict) -> str | None:
    """回傳資料庫中第一個 title 欄位名稱（Notion 每個 DB 一定會有一個 title）。"""
    for name, meta in (props_meta or {}).items():
        if (meta or {}).get("type") == "title":
            return name
    return None


def _build_text_property_by_type(prop_type: str, value: str):
    """依 Notion property type 產生正確 payload（只處理文字相關）。"""
    v = (value or "").strip()
    if prop_type == "title":
        return {"title": [{"text": {"content": v}}]} if v else {"title": []}
    if prop_type == "rich_text":
        return {"rich_text": [{"text": {"content": v}}]} if v else {"rich_text": []}
    # 其他型態不支援 → 回 None
    return None


def _best_set_text(props: dict, props_meta: dict, prop_name: str, value: str) -> None:
    """如果欄位存在且是 title/rich_text，盡力寫入；否則忽略。"""
    meta = (props_meta or {}).get(prop_name)
    if not meta:
        return
    payload = _build_text_property_by_type((meta or {}).get("type"), value)
    if payload is not None:
        props[prop_name] = payload


def _best_set_select(props: dict, props_meta: dict, db_id: str, prop_name: str, value: str) -> None:
    meta = (props_meta or {}).get(prop_name)
    if not meta or (meta.get("type") != "select"):
        return
    v = (value or "").strip()
    if not v:
        return
    options = get_select_options(db_id, prop_name) or []
    if (not options) or (v in options):
        props[prop_name] = {"select": {"name": v}}


def _equals_filter_by_type(props_meta: dict, prop_name: str, value: str) -> dict | None:
    """依欄位型態產生 Notion filter（title/rich_text）。"""
    meta = (props_meta or {}).get(prop_name) or {}
    t = meta.get("type")
    v = (value or "").strip()
    if not v:
        return None
    if t == "title":
        return {"property": prop_name, "title": {"equals": v}}
    if t == "rich_text":
        return {"property": prop_name, "rich_text": {"equals": v}}
    return None


def haversine_m(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """回傳兩點距離（公尺）"""
    R = 6371000.0
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c


def gps_widget_queryparam():
    """用瀏覽器 Geolocation 拿到座標後寫到 query params，再刷新"""
    html = """
    <script>
    navigator.geolocation.getCurrentPosition(
      (pos) => {
        const lat = pos.coords.latitude;
        const lon = pos.coords.longitude;
        const url = new URL(window.location.href);
        url.searchParams.set("lat", lat);
        url.searchParams.set("lon", lon);
        url.searchParams.delete("gps_err");
        window.location.href = url.toString();
      },
      (err) => {
        const url = new URL(window.location.href);
        url.searchParams.set("gps_err", err.message);
        window.location.href = url.toString();
      },
      {enableHighAccuracy:true, timeout:10000, maximumAge:0}
    );
    </script>
    """
    components.html(html, height=0)


def _qp_get_first(qp, key: str) -> str | None:
    """
    FIX: Streamlit query_params 在不同版本/情境下可能回傳 list / tuple / str
    這裡統一取第一個字串值。
    """
    try:
        v = qp.get(key)
        if v is None:
            return None
        if isinstance(v, (list, tuple)):
            return str(v[0]) if v else None
        return str(v)
    except Exception:
        return None


def _sync_gps_to_session_state():
    """把 query params 的 lat/lon 同步進 session_state，並清掉 query params 避免重跑循環"""
    qp = st.query_params

    gps_err = _qp_get_first(qp, "gps_err")
    if gps_err:
        st.session_state["gps_err"] = str(gps_err)
        try:
            st.query_params.clear()
        except Exception:
            pass
        return

    lat_s = _qp_get_first(qp, "lat")
    lon_s = _qp_get_first(qp, "lon")

    if lat_s is not None and lon_s is not None:
        try:
            st.session_state["gps_lat"] = float(lat_s)
            st.session_state["gps_lon"] = float(lon_s)
            st.session_state["gps_err"] = ""
        except Exception:
            st.session_state["gps_err"] = "GPS 座標解析失敗"

        try:
            st.query_params.clear()
        except Exception:
            pass


def _day_range(d: date) -> tuple[datetime, datetime]:
    start = datetime.combine(d, datetime.min.time())
    end = start + timedelta(days=1)
    return start, end


@st.cache_data(ttl=30)
def has_punch(employee_name: str, d: date, punch_type: str) -> bool:
    """
    回傳：該員工在 d 當天是否已存在 punch_type（上班/下班）打卡記錄
    Notion 欄位建議：
      - 員工姓名 (title)
      - 打卡類型 (select) -> 上班/下班
      - 打卡時間 (date)
    """
    if not PUNCH_DB_ID:
        return False

    start_dt, end_dt = _day_range(d)

    try:
        res = notion.databases.query(
            database_id=PUNCH_DB_ID,
            filter={
                "and": [
                    {"property": "員工姓名", "title": {"equals": employee_name}},
                    {"property": "打卡類型", "select": {"equals": punch_type}},
                    {"property": "打卡時間", "date": {"on_or_after": start_dt.isoformat()}},
                    {"property": "打卡時間", "date": {"before": end_dt.isoformat()}},
                ]
            },
            page_size=1,
        )
        return bool(res.get("results"))
    except Exception:
        return False


def create_punch_record(
    employee_name: str,
    punch_type: str,             # "上班" / "下班"
    lat: float,
    lon: float,
    dist_m: float,
    passed: bool,
    note: str = "",
    actor: str = "",
) -> bool:

    if not PUNCH_DB_ID:
        st.error("❌ 尚未設定 PUNCH_DB_ID（打卡記錄表 Database ID）")
        return False

    employee_name = (employee_name or "").strip()
    punch_type = (punch_type or "").strip()
    note = (note or "").strip()

    if not employee_name:
        st.error("❌ 員工姓名不可為空")
        return False

    try:
        props_meta = get_db_properties(PUNCH_DB_ID) or {}

        def has_prop(n: str) -> bool:
            return n in props_meta

        props = {}

        # 必填
        if has_prop("員工姓名"):
            props["員工姓名"] = {"title": [{"text": {"content": employee_name}}]}
        if has_prop("打卡類型"):
            props["打卡類型"] = {"select": {"name": punch_type}}
        if has_prop("打卡時間"):
            props["打卡時間"] = {"date": {"start": datetime.now().isoformat()}}

        # GPS資訊（可選）
        if has_prop("緯度"):
            props["緯度"] = {"number": float(lat)}
        if has_prop("經度"):
            props["經度"] = {"number": float(lon)}
        if has_prop("距離"):
            props["距離"] = {"number": float(dist_m)}
        if has_prop("GPS通過"):
            props["GPS通過"] = {"checkbox": bool(passed)}
        if has_prop("備註"):
            props["備註"] = {"rich_text": [{"text": {"content": note}}]} if note else {"rich_text": []}

        notion.pages.create(parent={"database_id": PUNCH_DB_ID}, properties=props)
        log_action(actor or employee_name, "打卡", f"{employee_name}｜{punch_type}｜距離{dist_m:.1f}m", "成功")
        return True

    except Exception as e:
        st.error(f"打卡寫入失敗：{e}")
        log_action(actor or employee_name, "打卡", f"寫入失敗：{e}", "系統錯誤")
        return False


@st.cache_data(ttl=60)
def list_punch_records(employee_name: str, y: int, m: int, limit: int = 500) -> list[dict]:
    """
    查詢某員工某月打卡
    回傳欄位：打卡時間、類型、距離、GPS通過
    """
    if not PUNCH_DB_ID:
        return []

    employee_name = (employee_name or "").strip()
    if not employee_name:
        return []

    start_d = date(int(y), int(m), 1)
    if int(m) == 12:
        end_d = date(int(y) + 1, 1, 1)
    else:
        end_d = date(int(y), int(m) + 1, 1)

    try:
        # FIX: Notion query 有分頁，原本只抓前 100 筆會漏資料
        rows = []
        next_cursor = None

        props_meta = get_db_properties(PUNCH_DB_ID) or {}
        name_filter = _equals_filter_by_type(props_meta, "員工姓名", employee_name)
        if not name_filter:
            return []

        base_query = {
            "database_id": PUNCH_DB_ID,
            "filter": {
                "and": [
                    name_filter,
                    {"property": "打卡時間", "date": {"on_or_after": datetime.combine(start_d, datetime.min.time()).isoformat()}},
                    {"property": "打卡時間", "date": {"before": datetime.combine(end_d, datetime.min.time()).isoformat()}},
                ]
            },
            "sorts": [{"property": "打卡時間", "direction": "descending"}],
            "page_size": 100,
        }

        while True:
            q = dict(base_query)
            if next_cursor:
                q["start_cursor"] = next_cursor

            res = notion.databases.query(**q)

            for page in res.get("results", []):
                props = page.get("properties", {}) or {}

                def get_date_start(name: str) -> str:
                    d = (props.get(name, {}) or {}).get("date")
                    return d.get("start", "") if d else ""

                def get_select(name: str) -> str:
                    s = (props.get(name, {}) or {}).get("select")
                    return s.get("name", "") if s else ""

                def get_number(name: str) -> float:
                    return float((props.get(name, {}) or {}).get("number") or 0.0)

                def get_checkbox(name: str) -> bool:
                    v = (props.get(name, {}) or {}).get("checkbox")
                    return bool(v) if v is not None else False

                rows.append({
                    "打卡時間": get_date_start("打卡時間"),
                    "打卡類型": get_select("打卡類型"),
                    "距離": get_number("距離"),
                    "GPS通過": get_checkbox("GPS通過"),
                })

                if len(rows) >= int(limit):
                    return rows

            if not res.get("has_more"):
                break

            next_cursor = res.get("next_cursor")

        return rows

    except Exception:
        return []


def sanitize_announce_text(s: str) -> str:
    """
    防止公告內容被 Streamlit/Markdown 當成程式碼區塊或 HTML 注入
    - 轉義 &,<,>
    - 把 ``` 破壞掉，避免變 code block
    - 保留換行為 <br>
    """
    if not s:
        return ""
    s = str(s)

    # 先處理最關鍵：三個反引號（Markdown code fence）
    s = s.replace("```", "``\u200b`")  # 插入零寬字元打斷

    # HTML escape
    s = (s.replace("&", "&amp;")
           .replace("<", "&lt;")
           .replace(">", "&gt;"))

    # 換行轉 <br>
    s = s.replace("\n", "<br>")
    return s


def list_employee_names(limit: int = 200) -> list[str]:
    """從帳號管理表抓出所有員工姓名（自動適配 title / rich_text）。"""
    try:
        if not ACCOUNT_DB_ID:
            return []

        props_meta = get_db_properties(ACCOUNT_DB_ID) or {}
        if "員工姓名" not in props_meta:
            return []

        ptype = (props_meta.get("員工姓名", {}) or {}).get("type")

        res = notion.databases.query(
            database_id=ACCOUNT_DB_ID,
            page_size=min(limit, 100),
        )

        names: list[str] = []
        for page in res.get("results", []):
            p = page.get("properties", {}) or {}
            cell = p.get("員工姓名", {}) or {}

            if ptype == "title":
                t = cell.get("title", []) or []
                name = (t[0].get("plain_text") or "").strip() if t else ""
            elif ptype == "rich_text":
                name = _rt_get_first_plain_text(cell)
            else:
                name = ""

            if name:
                names.append(name)

        # 去重 + 排序
        names = sorted(list(dict.fromkeys(names)))
        return names
    except Exception:
        return []
    except Exception:
        return []

DUTY_SHIFT_COLUMNS = [
    "檢驗線(中)",
    "檢驗線(晚)",
    "收費員(中)",
    "收費員(晚)",
    "打掃工作",
]

WEEKDAY_MAP = ["一", "二", "三", "四", "五", "六", "日"]

def create_duty_record(duty_date, shift_name, employee_name, note="", weekday_text=""):
    try:
        if not DUTY_DB_ID:
            return False

        y = duty_date.year
        m = duty_date.month

        if not weekday_text:
            weekday_text = WEEKDAY_MAP[duty_date.weekday()]

        title = f"{y}-{m:02d}-{duty_date.day:02d} {shift_name} {employee_name}"

        props = {
            "員工姓名": {"title": [{"text": {"content": str(employee_name)}}]},
            "年份": {"number": int(y)},
            "月份": {"number": int(m)},
            "日期": {"date": {"start": duty_date.isoformat()}},
            "星期": {"rich_text": [{"text": {"content": str(weekday_text)}}]},
            "班別": {"select": {"name": str(shift_name)}},
        }

        if note:
            props["備註"] = {"rich_text": [{"text": {"content": str(note)}}]}

        notion.pages.create(parent={"database_id": DUTY_DB_ID}, properties=props)
        return True
    except Exception:
        return False



def count_employee_duty_times(employee_name: str, y: int, m: int, shift_filter: list[str] | None = None) -> int:
    """統計某員工在某年某月出現幾次值班（以值班記錄表直式資料為準）"""
    try:
        if not DUTY_DB_ID:
            return 0

        duty_props = get_db_properties(DUTY_DB_ID) or {}
        if "年份" not in duty_props or "月份" not in duty_props:
            return 0

        # 先找員工 page_id（同上：若你員工表不是 ACCOUNT_DB_ID，要改）
        emp_page = get_account_page_by_username(employee_name)
        emp_id = emp_page["id"] if emp_page else None
        if not emp_id:
            return 0

        filters = [
            {"property": "年份", "number": {"equals": int(y)}},
            {"property": "月份", "number": {"equals": int(m)}},
            {"property": "值班人員", "relation": {"contains": emp_id}},
        ]

        if shift_filter:
            # Notion 的 select filter：只能 equals / does_not_equal / is_empty...
            # 多選用 or 包起來
            or_filters = [{"property": "班別", "select": {"equals": s}} for s in shift_filter]
            filters.append({"or": or_filters})

        q = {
            "database_id": DUTY_DB_ID,
            "page_size": 100,
            "filter": {"and": filters},
        }
        res = notion.databases.query(**q)
        return len(res.get("results", []))
    except Exception:
        return 0


# ============================================================
# 📢 公告（Notion 公告紀錄表）功能：管理員可新增/完成；員工只可看
# ============================================================

@st.cache_data(ttl=60)
def resolve_title_prop_name(database_id: str) -> str | None:
    """
    Notion DB 一定有一個 title 欄位，但名稱可能是 Name / 標題 / 任何你改過的名字
    這裡自動找第一個 type=title 的欄位名。
    """
    props = get_db_properties(database_id) or {}
    for k, meta in props.items():
        if (meta or {}).get("type") == "title":
            return k
    return None


def _safe_iso(dt: datetime) -> str:
    return dt.isoformat()


def _now_iso() -> str:
    return datetime.now().isoformat()


def _make_announce_title(content: str, pub_date: date) -> str:
    c = (content or "").strip().replace("\n", " ")
    c = c[:20] + ("…" if len(c) > 20 else "")
    return f"{pub_date.isoformat()}｜{c or '公告'}"


def create_announcement(publish_date: date, content: str, end_date: date | None, actor: str = "") -> bool:
    if not ANNOUNCE_DB_ID:
        st.error("❌ 尚未設定 ANNOUNCE_DB_ID（公告紀錄表 Database ID）")
        return False

    content = (content or "").strip()
    if not content:
        st.error("❌ 公告內容不可為空")
        return False

    try:
        props_meta = get_db_properties(ANNOUNCE_DB_ID) or {}
        title_prop = resolve_title_prop_name(ANNOUNCE_DB_ID)  # 自動找 title 欄位

        def has_prop(n: str) -> bool:
            return n in props_meta

        props = {}

        # ✅ Title（Notion 必填）
        if title_prop:
            props[title_prop] = {"title": [{"text": {"content": _make_announce_title(content, publish_date)}}]}

        # ✅ 完成情況（預設 False）
        if has_prop("完成情況"):
            props["完成情況"] = {"checkbox": False}

        # ✅ 發布日期
        if has_prop("發布日期"):
            props["發布日期"] = {"date": {"start": datetime.combine(publish_date, datetime.min.time()).isoformat()}}

        # ✅ 公告內容
        if has_prop("公告內容"):
            # rich_text
            if (props_meta.get("公告內容", {}) or {}).get("type") == "rich_text":
                props["公告內容"] = {"rich_text": [{"text": {"content": content}}]}
            # 也有人把公告內容做成 title（就當備援）
            elif (props_meta.get("公告內容", {}) or {}).get("type") == "title":
                props["公告內容"] = {"title": [{"text": {"content": content}}]}
            else:
                # 保底：仍用 rich_text 方式寫
                props["公告內容"] = {"rich_text": [{"text": {"content": content}}]}

        # ✅ 結束時間（可空）
        if end_date and has_prop("結束時間"):
            props["結束時間"] = {"date": {"start": datetime.combine(end_date, datetime.min.time()).isoformat()}}

        notion.pages.create(parent={"database_id": ANNOUNCE_DB_ID}, properties=props)
        log_action(actor or "—", "公告管理", f"新增公告：{publish_date.isoformat()}｜{content[:30]}", "成功")
        return True

    except Exception as e:
        st.error(f"新增公告失敗：{e}")
        log_action(actor or "—", "公告管理", f"新增公告失敗：{e}", "系統錯誤")
        return False


def mark_announcement_done(page_id: str, done: bool, actor: str = "") -> bool:
    if not ANNOUNCE_DB_ID:
        return False
    try:
        props_meta = get_db_properties(ANNOUNCE_DB_ID) or {}
        if "完成情況" not in props_meta:
            st.warning("⚠️ 公告表沒有『完成情況』欄位（checkbox），無法勾選完成。")
            return False

        notion.pages.update(page_id=page_id, properties={"完成情況": {"checkbox": bool(done)}})
        log_action(actor or "—", "公告管理", f"勾選完成：{page_id} -> {done}", "成功")
        return True
    except Exception as e:
        st.error(f"更新完成情況失敗：{e}")
        log_action(actor or "—", "公告管理", f"更新完成情況失敗：{e}", "系統錯誤")
        return False


def archive_announcement(page_id: str, actor: str = "") -> bool:
    if not ANNOUNCE_DB_ID:
        return False
    try:
        notion.pages.update(page_id=page_id, archived=True)
        log_action(actor or "—", "公告管理", f"封存公告：{page_id}", "成功")
        return True
    except Exception as e:
        st.error(f"封存公告失敗：{e}")
        log_action(actor or "—", "公告管理", f"封存公告失敗：{e}", "系統錯誤")
        return False


def _extract_announce_row(page: dict) -> dict:
    props = page.get("properties", {}) or {}

    def get_checkbox(name: str) -> bool:
        v = (props.get(name, {}) or {}).get("checkbox")
        return bool(v) if v is not None else False

    def get_date_start(name: str) -> str:
        d = (props.get(name, {}) or {}).get("date")
        if d and d.get("start"):
            return d["start"]
        return ""

    def get_rich(name: str) -> str:
        rt = (props.get(name, {}) or {}).get("rich_text", []) or []
        return rt[0].get("plain_text", "") if rt else ""

    def get_title(name: str) -> str:
        t = (props.get(name, {}) or {}).get("title", []) or []
        return t[0].get("plain_text", "") if t else ""

    content = ""
    # 公告內容 可能是 rich_text 或 title
    if "公告內容" in props:
        ptype = (props.get("公告內容", {}) or {}).get("type")
        if ptype == "rich_text":
            content = get_rich("公告內容")
        elif ptype == "title":
            content = get_title("公告內容")

    return {
        "_page_id": page.get("id"),
        "完成情況": get_checkbox("完成情況") if "完成情況" in props else False,
        "發布日期": get_date_start("發布日期"),
        "公告內容": content,
        "結束時間": get_date_start("結束時間"),
        "建立時間": page.get("created_time", ""),
        "最後更新時間": page.get("last_edited_time", ""),
    }


@st.cache_data(ttl=60)
def list_announcements(include_hidden: bool, limit: int = 200) -> list[dict]:
    """
    include_hidden=True  -> 管理員看全部（含已完成/過期）
    include_hidden=False -> 只回傳未隱藏（給首頁/員工）
    """
    if not ANNOUNCE_DB_ID:
        return []

    props_meta = get_db_properties(ANNOUNCE_DB_ID) or {}
    has_done = "完成情況" in props_meta
    has_end = "結束時間" in props_meta

    filters = []

    if (not include_hidden) and (has_done or has_end):
        and_list = []
        if has_done:
            and_list.append({"property": "完成情況", "checkbox": {"equals": False}})
        if has_end:
            and_list.append({
                "or": [
                    {"property": "結束時間", "date": {"is_empty": True}},
                    {"property": "結束時間", "date": {"after": _now_iso()}},
                ]
            })
        if and_list:
            filters = [{"and": and_list}]

    query = {
        "database_id": ANNOUNCE_DB_ID,
        "page_size": 100,
        "sorts": [{"property": "發布日期", "direction": "descending"}] if "發布日期" in props_meta else [{"timestamp": "created_time", "direction": "descending"}],
    }
    if filters:
        query["filter"] = filters[0]

    try:
        rows: list[dict] = []
        next_cursor = None
        while True:
            if next_cursor:
                query["start_cursor"] = next_cursor
            res = notion.databases.query(**query)
            for page in res.get("results", []):
                rows.append(_extract_announce_row(page))
                if len(rows) >= int(limit):
                    return rows
            if not res.get("has_more"):
                break
            next_cursor = res.get("next_cursor")
        return rows
    except Exception as e:
        st.error(f"讀取公告失敗：{e}")
        return []


# ============================================================
# ✅ 值班排班表（橫向填寫 -> Notion直式 -> Excel橫向輸出）
# ============================================================

WEEKDAY_MAP = {0: "一", 1: "二", 2: "三", 3: "四", 4: "五", 5: "六", 6: "日"}

DUTY_COLUMNS = ["日期", "星期", "檢驗線(中)", "檢驗線(晚)", "收費員(中)", "收費員(晚)", "打掃工作", "備註"]

def _month_date_range(y: int, m: int) -> tuple[date, date]:
    last_day = calendar.monthrange(y, m)[1]
    start = date(y, m, 1)
    end_exclusive = date(y, m, last_day) + timedelta(days=1)
    return start, end_exclusive

def build_month_template(y: int, m: int) -> list[dict]:
    last_day = calendar.monthrange(y, m)[1]
    rows = []
    for d in range(1, last_day + 1):
        wd = datetime(y, m, d).weekday()
        rows.append({
            "日期": d,
            "星期": WEEKDAY_MAP.get(wd, ""),
            "檢驗線(中)": [],
            "檢驗線(晚)": [],
            "收費員(中)": [],
            "收費員(晚)": [],
            "打掃工作": "",
            "備註": "",
        })
    return rows

def query_duty_rows_from_notion(y: int, m: int) -> list[dict]:
    if not DUTY_DB_ID:
        st.error("❌ 尚未設定 DUTY_DB_ID（值班排班表 Database ID）")
        return []

    props_meta = get_db_properties(DUTY_DB_ID) or {}

    # 取得真正欄名（容錯）
    k_year = resolve_prop_key(props_meta, "年份")
    k_month = resolve_prop_key(props_meta, "月份")
    k_day = resolve_prop_key(props_meta, "日期")
    k_week = resolve_prop_key(props_meta, "星期")

    # 你的表格截圖：檢驗線/收費員 都是文字欄位（rich_text）
    k_mid_chk = resolve_prop_key(props_meta, "檢驗線(中)")
    k_night_chk = resolve_prop_key(props_meta, "檢驗線(晚)")
    k_mid_cash = resolve_prop_key(props_meta, "收費員(中)")
    k_night_cash = resolve_prop_key(props_meta, "收費員(晚)")
    k_clean = resolve_prop_key(props_meta, "打掃工作")
    k_note = resolve_prop_key(props_meta, "備註")

    # 以年/月過濾（你 Notion 有 年份/月 兩個 number 欄）
    filters = []
    if k_year:
        filters.append({"property": k_year, "number": {"equals": int(y)}})
    if k_month:
        filters.append({"property": k_month, "number": {"equals": int(m)}})

    try:
        res = notion.databases.query(
            database_id=DUTY_DB_ID,
            page_size=200,
            filter={"and": filters} if filters else None,
        )
    except Exception as e:
        st.error(f"查詢 Notion 值班排班失敗：{e}")
        return []

    def get_rich_text(props: dict, key: str) -> str:
        if not key:
            return ""
        p = props.get(key, {}) or {}
        rt = (p.get("rich_text") or [])
        if rt:
            return "".join([x.get("plain_text", "") for x in rt]).strip()
        t = (p.get("title") or [])
        if t:
            return "".join([x.get("plain_text", "") for x in t]).strip()
        return ""

    rows = []
    for pg in res.get("results", []):
        props = pg.get("properties", {}) or {}

        day_txt = get_rich_text(props, k_day)
        # day_txt 可能是 "1" / "01" / "1日" -> 抓數字
        d = ""
        if day_txt:
            digits = "".join(ch for ch in day_txt if ch.isdigit())
            d = int(digits) if digits else ""

        row = {
            "日期": d,
            "星期": get_rich_text(props, k_week),
            "檢驗線(中)": get_rich_text(props, k_mid_chk),
            "檢驗線(晚)": get_rich_text(props, k_night_chk),
            "收費員(中)": get_rich_text(props, k_mid_cash),
            "收費員(晚)": get_rich_text(props, k_night_cash),
            "打掃工作": get_rich_text(props, k_clean),
            "備註": get_rich_text(props, k_note),
            "_page_id": pg.get("id"),
        }
        rows.append(row)

    # 用日期排序（有些可能空）
    rows.sort(key=lambda r: (999 if r.get("日期") in ("", None) else int(r["日期"])))
    return rows



def export_duty_excel_bytes(y: int, m: int, rows: list[dict]) -> bytes:
    """輸出成你參考圖那種橫向班表（簡化版：可再加顏色/合併儲存格）。"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"{m:02d}月值班表"

    title = f"{y}年{m}月份值班表"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(DUTY_COLUMNS))
    ws.cell(row=1, column=1, value=title).font = Font(size=16, bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # header
    for c, name in enumerate(DUTY_COLUMNS, start=1):
        cell = ws.cell(row=2, column=c, value=name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # body
    for i, r in enumerate(rows, start=3):
        for c, name in enumerate(DUTY_COLUMNS, start=1):
            v = r.get(name, "")
            if isinstance(v, list):
                v = "、".join(v)
            cell = ws.cell(row=i, column=c, value=v)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    # column widths（你可再微調成更像參考圖）
    widths = {
        "日期": 6,
        "星期": 6,
        "檢驗線(中)": 16,
        "檢驗線(晚)": 16,
        "收費員(中)": 16,
        "收費員(晚)": 16,
        "打掃工作": 14,
        "備註": 18,
    }
    for i, name in enumerate(DUTY_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 14)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

def _norm_prop_name(s: str) -> str:
    """把欄位名稱正規化：去空白、全形括號轉半形，避免 Notion 欄名些微差異造成找不到。"""
    if s is None:
        return ""
    s = str(s)
    trans = str.maketrans({"（": "(", "）": ")", "　": " ", "\u00A0": " "})
    s = s.translate(trans)
    s = s.replace(" ", "")
    return s.strip().lower()

def resolve_prop_key(props_meta: dict, want: str) -> str | None:
    """用 want 去 Notion DB properties 裡找真正的 key（容錯：全形括號/空白/大小寫）。"""
    if not props_meta:
        return None
    want_n = _norm_prop_name(want)
    # 先精準
    if want in props_meta:
        return want
    # 再容錯比對
    for k in props_meta.keys():
        if _norm_prop_name(k) == want_n:
            return k
    return None

def resolve_title_prop(database_id: str) -> str | None:
    """找 Notion DB 的 title 欄位名稱（title 是 Notion 必填）。"""
    props = get_db_properties(database_id) or {}
    for k, meta in props.items():
        if (meta or {}).get("type") == "title":
            return k
    return None




def render_duty_schedule_page():
    st.header("值班排班表（管理員）")

    if not DUTY_DB_ID:
        st.error("❌ 尚未設定 DUTY_DB_ID（值班排班表 DB ID）")
        st.stop()

    # 模式：list(查詢) / create(新增)
    if "duty_mode" not in st.session_state:
        st.session_state["duty_mode"] = "list"

    # ✅ 右上角按鈕（新增值班 / 新增加班設定）
    colL, colR = st.columns([0.60, 0.40])
    with colR:
        if st.session_state["duty_mode"] == "list":
            b1, b2 = st.columns(2)
            with b1:
                if st.button("➕ 新增值班排班", use_container_width=True):
                    st.session_state["duty_mode"] = "create"
                    st.rerun()
            with b2:
                if st.button("🕒 新增加班設定", use_container_width=True):
                    st.session_state["open_overtime_rule_dialog"] = True
                    st.rerun()
        else:
            if st.button("← 回到查詢", use_container_width=True):
                st.session_state["duty_mode"] = "list"
                st.rerun()

    # ✅ 彈窗：新增加班設定
    if st.session_state.get("open_overtime_rule_dialog"):
        # Streamlit 版本支援 st.dialog 才能真正「懸浮式」
        if hasattr(st, "dialog"):
            @st.dialog("🕒 新增加班設定")
            def _overtime_rule_dialog():
                if not OVERTIME_RULE_DB_ID:
                    st.error("❌ 尚未設定 OVERTIME_RULE_DB_ID（加班設定表 DB ID）")
                    if st.button("關閉"):
                        st.session_state["open_overtime_rule_dialog"] = False
                        st.rerun()
                    return

                yy = st.number_input("年份", min_value=2020, max_value=2100, value=int(st.session_state.get("duty_y", datetime.now().year)), step=1, key="ot_rule_y")
                mm = st.number_input("月份", min_value=1, max_value=12, value=int(st.session_state.get("duty_m", datetime.now().month)), step=1, key="ot_rule_m")

                st.caption("名稱會自動產生（YYYY-MM）")
                st.text_input("名稱", value=f"{int(yy)}-{int(mm):02d}", disabled=True, key="ot_rule_name")

                c1, c2 = st.columns(2)
                with c1:
                    shift_hours = st.number_input("班次換算時數（1 次 = 幾小時）", min_value=0.0, value=1.0, step=0.5, key="ot_rule_shift_hours")
                with c2:
                    hourly_rate = st.number_input("加班時薪", min_value=0.0, value=0.0, step=10.0, key="ot_rule_hourly_rate")
                note = st.text_area("備註", placeholder="可留空", key="ot_rule_note")

                cc1, cc2 = st.columns(2)
                with cc1:
                    if st.button("✅ 匯入 Notion（同年同月覆蓋/新增）", use_container_width=True):
                        try:
                            page_id = upsert_overtime_rule_to_notion(
                                int(yy),
                                int(mm),
                                float(shift_hours),
                                float(hourly_rate),
                                note or "",
                            )
                            st.success(f"✅ 已寫入 Notion（page_id: {page_id[:8]}...）")
                            st.session_state["open_overtime_rule_dialog"] = False
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ 寫入失敗：{e}")
                with cc2:
                    if st.button("取消", use_container_width=True):
                        st.session_state["open_overtime_rule_dialog"] = False
                        st.rerun()

            _overtime_rule_dialog()
        else:
            # 如果 Streamlit 版本太舊：退化成 expander（不是真正懸浮）
            with st.expander("🕒 新增加班設定（你的 Streamlit 版本不支援 dialog，這裡用展開區代替）", expanded=True):
                st.session_state["open_overtime_rule_dialog"] = False
                st.warning("你的 Streamlit 版本不支援 st.dialog，請升級 Streamlit 才能使用懸浮式表單。")
    # 共用：年/月
    y = st.number_input("排班年份", min_value=2020, max_value=2100, value=datetime.now().year, step=1, key="duty_y")
    m = st.number_input("排班月份", min_value=1, max_value=12, value=datetime.now().month, step=1, key="duty_m")

    employees = list_employee_names()
    if not employees:
        st.warning("⚠️ 目前抓不到員工名單，請確認員工資料表欄位『員工姓名』")
        st.stop()

    # ✅ 員工選項順序一定要穩定，不然 data_editor 會被視為「結構改變」而刷新
    employees = sorted([str(e).strip() for e in employees if str(e).strip()])

    import pandas as pd

    # ✅ 每個月份一份草稿（暫存在瀏覽器 session_state）
    df_key = f"duty_df_{int(y)}_{int(m)}"
    editor_key = f"duty_editor_{int(y)}_{int(m)}"  # 只用來固定 widget，不去寫 st.session_state[editor_key]
    emp_key = f"duty_employees_{int(y)}_{int(m)}"
    if emp_key not in st.session_state:
        emps = list_employee_names()
        st.session_state[emp_key] = sorted([str(e).strip() for e in emps if str(e).strip()])
    employees = st.session_state[emp_key]

    def coerce_duty_df_list_columns(df):
        """
        ✅ 防呆版：確保值班欄位永遠是 list
        - 若拿到 dict（某些情況會是 widget state），就直接回傳原樣避免爆炸
        """
        if df is None:
            return df
        if isinstance(df, dict):
            return df
        for c in DUTY_SHIFT_COLUMNS:
            if c in df.columns:
                df[c] = df[c].apply(normalize_multi_people_cell)
        return df

    def _build_month_df(_y: int, _m: int) -> pd.DataFrame:
        days = calendar.monthrange(int(_y), int(_m))[1]
        rows = []
        for d in range(1, days + 1):
            dt = date(int(_y), int(_m), int(d))
            rows.append({
                "日期": d,
                "星期": WEEKDAY_MAP[dt.weekday()],  # 文字型態
                "檢驗線(中)": [],
                "檢驗線(晚)": [],
                "收費員(中)": [],
                "收費員(晚)": [],
                "打掃工作": [],
                "備註": "",
            })
        df0 = pd.DataFrame(rows)
        return coerce_duty_df_list_columns(df0)



    # ==========================
    # A) 查詢模式（主頁）
    # ==========================
    if st.session_state["duty_mode"] == "list":
        if st.button("🔎 查詢", use_container_width=True):
            duty_df = query_duty_month_to_horizontal_df(int(y), int(m), employees)
            st.session_state["duty_query_df"] = duty_df

        duty_df = st.session_state.get("duty_query_df")
        if duty_df is None or duty_df.empty:
            st.info("請選擇年份月份並按『查詢』，查到才會顯示表格。")
            return

        st.data_editor(duty_df, use_container_width=True, hide_index=True, disabled=True)
        return

    # ==========================
    # B) 新增模式（建立本月）
    # ==========================
    st.caption("每格可多選員工（可 2~3 人或更多）；填寫內容會先暫存在瀏覽器，按下『一鍵匯入 Notion 並下載 Excel』才會寫入 Notion。")

    # ✅ 建立 / 重新建立本月表格
    if st.button("產生本月表格", use_container_width=True, key=f"gen_duty_{int(y)}_{int(m)}"):
        st.session_state[df_key] = _build_month_df(int(y), int(m))
        st.rerun()

    # ✅ 取草稿
    if df_key not in st.session_state or st.session_state[df_key] is None or st.session_state[df_key].empty:
        st.info("請先按『產生本月表格』。")
        return


    def _apply_duty_editor_delta():
        """把 data_editor 尚未完整回傳的變更（edited_rows）套回 df，避免點下一格就消失。"""
        state = st.session_state.get(editor_key)
        if not isinstance(state, dict):
            return

        df = st.session_state.get(df_key)
        if df is None or df.empty:
            return

        # 1) 套用 edited_rows
        edited_rows = state.get("edited_rows", {}) or {}
        for r_idx, changes in edited_rows.items():
            # r_idx 是 row index（通常是 0..n-1）
            for col, val in (changes or {}).items():
                if col in df.columns and r_idx in df.index:
                    df.at[r_idx, col] = val

        # 2) 需要的話也可處理新增/刪除（你目前 num_rows 固定，所以通常不會用到）
        # added_rows = state.get("added_rows", []) or []
        # deleted_rows = state.get("deleted_rows", []) or []

        df = coerce_duty_df_list_columns(df).reset_index(drop=True)
        st.session_state[df_key] = df


    edited = st.data_editor(
        st.session_state[df_key],
        key=editor_key,
        use_container_width=True,
        hide_index=True,
        disabled=["日期", "星期"],
        column_config={
            "檢驗線(中)": st.column_config.MultiselectColumn("檢驗線(中)", options=employees),
            "檢驗線(晚)": st.column_config.MultiselectColumn("檢驗線(晚)", options=employees),
            "收費員(中)": st.column_config.MultiselectColumn("收費員(中)", options=employees),
            "收費員(晚)": st.column_config.MultiselectColumn("收費員(晚)", options=employees),
            "打掃工作": st.column_config.MultiselectColumn("打掃工作", options=employees),
            "備註": st.column_config.TextColumn("備註"),
        },
        on_change=_apply_duty_editor_delta,  # ✅ 關鍵：變更立刻寫回 df_key
    )

    # ✅ 保底：有些情況回傳 edited 已經含最新值，仍然同步一次
    edited = coerce_duty_df_list_columns(edited).reset_index(drop=True)
    st.session_state[df_key] = edited


    # -------------------------
    # ✅ 一鍵：下載 Excel + 同時匯入 Notion
    # -------------------------
    def _do_import_duty():
        df_now = st.session_state.get(df_key)
        if df_now is None or df_now.empty:
            st.session_state["duty_import_result"] = ("⚠️ 沒有可匯入的資料", 0, 0)
            return

        try:
            # ✅ 以「一天一列」upsert：Notion 內為直式（一天一筆）
            upsert_duty_rows_to_notion(int(y), int(m), df_now.to_dict("records"))
            # ✅ 同步更新：加班次數表（平日出現次數 -> 時數）
            ot_ok, ot_fail = sync_overtime_count_from_duty_rows(int(y), int(m), df_now.to_dict("records"), actor=str(st.session_state.get("user", "")))
            st.session_state["duty_import_result"] = (f"✅ 匯入完成（加班次數表：成功 {ot_ok}，失敗 {ot_fail}）", 1, 0)
        except Exception as e:
            st.session_state["duty_import_result"] = (f"❌ 匯入失敗：{e}", 0, 1)

    with st.expander("➕ 一鍵匯入 Notion 並下載 Excel", expanded=True):
        excel_bytes = export_duty_excel_bytes(int(y), int(m), st.session_state[df_key].to_dict("records"))

        st.download_button(
            "✅ 一鍵匯入 Notion 並下載 Excel",
            data=excel_bytes,
            file_name=f"{int(y)}-{int(m):02d}_值班排班表.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=f"duty_oneclick_{int(y)}_{int(m)}",
            on_click=_do_import_duty,
        )

        msg = st.session_state.get("duty_import_result")
        if msg:
            title, ok, fail = msg
            st.success(f"{title}：成功 {ok}，失敗 {fail}（Excel 已下載）")

def get_overtime_rule(y: int, m: int) -> dict:
    """
    從 Notion【加班設定表】取得某年某月的規則
    回傳：
      {
        "shift_hours": float,
        "hourly_rate": float,
              }
    若查不到就回傳預設（0）
    """
    if not OVERTIME_RULE_DB_ID:
        return {"shift_hours": 0.0, "hourly_rate": 0.0}

    res = notion.databases.query(
        database_id=OVERTIME_RULE_DB_ID,
        page_size=5,
        filter={
            "and": [
                {"property": "年份", "number": {"equals": int(y)}},
                {"property": "月份", "number": {"equals": int(m)}},
            ]
        },
    )
    results = (res or {}).get("results", []) or []
    if not results:
        return {"shift_hours": 0.0, "hourly_rate": 0.0}

    p = (results[0] or {}).get("properties", {}) or {}

    def _num(name: str) -> float:
        try:
            return float(((p.get(name) or {}).get("number")) or 0.0)
        except Exception:
            return 0.0

    return {
        "shift_hours": _num("班次換算時數"),
        "hourly_rate": _num("加班時薪"),
    }


def get_overtime_count_hours(employee: str, y: int, m: int) -> float:
    """
    從【加班次數表】讀取該員工該年月的「時數」(number)。
    欄位假設：
      - 員工姓名：title
      - 年份：number
      - 月份：number
      - 時數：number
    """
    employee = (employee or "").strip()
    if (not OVERTIME_COUNT_DB_ID) or (not employee):
        return 0.0

    try:
        props_meta = get_db_properties(OVERTIME_COUNT_DB_ID) or {}
        k_emp = resolve_title_prop(OVERTIME_COUNT_DB_ID) or resolve_prop_key(props_meta, "員工姓名") or "員工姓名"
        k_year = resolve_prop_key(props_meta, "年份") or "年份"
        k_month = resolve_prop_key(props_meta, "月份") or "月份"
        k_hours = resolve_prop_key(props_meta, "時數") or "時數"

        # title filter
        flt = {
            "and": [
                {"property": k_year, "number": {"equals": int(y)}},
                {"property": k_month, "number": {"equals": int(m)}},
                {"property": k_emp, "title": {"equals": employee}},
            ]
        }
        res = notion.databases.query(database_id=OVERTIME_COUNT_DB_ID, page_size=5, filter=flt)
        results = (res or {}).get("results", []) or []
        if not results:
            return 0.0
        p = (results[0] or {}).get("properties", {}) or {}
        try:
            return float(((p.get(k_hours) or {}).get("number")) or 0.0)
        except Exception:
            return 0.0
    except Exception:
        return 0.0


def upsert_overtime_count_to_notion(employee: str, y: int, m: int, hours: float, actor: str = "") -> str | None:
    """
    同年同月同人：有就更新，沒有就新增（加班次數表）。回傳 page_id。
    """
    employee = (employee or "").strip()
    if (not OVERTIME_COUNT_DB_ID) or (not employee):
        return None

    try:
        props_meta = get_db_properties(OVERTIME_COUNT_DB_ID) or {}
        k_emp = resolve_title_prop(OVERTIME_COUNT_DB_ID) or resolve_prop_key(props_meta, "員工姓名") or "員工姓名"
        k_year = resolve_prop_key(props_meta, "年份") or "年份"
        k_month = resolve_prop_key(props_meta, "月份") or "月份"
        k_hours = resolve_prop_key(props_meta, "時數") or "時數"

        flt = {
            "and": [
                {"property": k_year, "number": {"equals": int(y)}},
                {"property": k_month, "number": {"equals": int(m)}},
                {"property": k_emp, "title": {"equals": employee}},
            ]
        }
        res = notion.databases.query(database_id=OVERTIME_COUNT_DB_ID, page_size=5, filter=flt)
        results = (res or {}).get("results", []) or []
        page_id = results[0]["id"] if results else None

        props = {
            k_emp: {"title": [{"text": {"content": employee}}]},
            k_year: {"number": int(y)},
            k_month: {"number": int(m)},
            k_hours: {"number": float(hours or 0.0)},
        }

        if page_id:
            notion.pages.update(page_id=page_id, properties=props)
            log_action(actor or "—", "加班次數表", f"覆蓋：{employee} {y}-{m:02d} 時數={float(hours or 0.0)}", "成功")
            return page_id
        else:
            created = notion.pages.create(parent={"database_id": OVERTIME_COUNT_DB_ID}, properties=props)
            pid = (created or {}).get("id")
            log_action(actor or "—", "加班次數表", f"新增：{employee} {y}-{m:02d} 時數={float(hours or 0.0)}", "成功")
            return pid
    except Exception as e:
        log_action(actor or "—", "加班次數表", f"寫入失敗：{employee} {y}-{m:02d}｜{e}", "系統錯誤")
        return None


def _parse_names_cell(v) -> list[str]:
    """把 data_editor 的 cell 值轉成 ['A','B']。支援 list / str(頓號、逗號、換行分隔)。"""
    if v is None:
        return []
    if isinstance(v, list):
        return [str(x).strip() for x in v if str(x).strip()]
    s = str(v).strip()
    if not s:
        return []
    for sep in ["、", ",", "，", ";", "；", "\n", "\t"]:
        s = s.replace(sep, " ")
    return [x.strip() for x in s.split(" ") if x.strip()]


def calc_overtime_hours_from_duty_rows(y: int, m: int, rows: list[dict]) -> dict[str, float]:
    """
    從「值班排班表（橫向）」計算每位員工在該月【平日(週一~週五)】出現的次數。

    ✅ 依你 2026-02-13 最新需求：
    - 「時數」= 員工在該月平日於排班表中出現的【總次數】
    - 同一天若在不同欄位出現多次（例如：檢驗線(中) + 收費員(晚)），要【累加】成 2 次
      （不再做同日去重）

    回傳：{員工姓名: 次數(float)}
    """
    counts: dict[str, float] = {}
    if not rows:
        return counts

    # 欄位：除了 日期/星期/備註 之外，都視為「會放員工名單」的欄位
    skip_cols = {"日期", "星期", "備註"}
    name_cols = [c for c in (rows[0].keys() if isinstance(rows[0], dict) else []) if c not in skip_cols]

    for r in rows:
        # 日期
        try:
            d = int(r.get("日期") or 0)
        except Exception:
            d = 0
        if d <= 0:
            continue

        # 只算平日（週一~週五）
        try:
            wd = datetime(int(y), int(m), int(d)).weekday()  # 0=Mon
        except Exception:
            continue
        if wd >= 5:  # 5=Sat,6=Sun
            continue

        # ✅ 不去重：每個欄位出現一次就 +1
        for col in name_cols:
            for emp in _parse_names_cell(r.get(col)):
                emp = (emp or "").strip()
                if not emp:
                    continue
                counts[emp] = float(counts.get(emp, 0.0) + 1.0)

    return counts


def sync_overtime_count_from_duty_rows(y: int, m: int, rows: list[dict], actor: str = "") -> tuple[int, int]:
    """
    把值班排班表的「平日出現次數」同步到【加班次數表】：
    - 有就覆蓋，沒有就新增
    回傳：(成功筆數, 失敗筆數)
    """
    if not OVERTIME_COUNT_DB_ID:
        return (0, 0)

    counts = calc_overtime_hours_from_duty_rows(int(y), int(m), rows or [])
    ok = 0
    fail = 0
    for emp, hours in counts.items():
        try:
            pid = upsert_overtime_count_to_notion(emp, int(y), int(m), float(hours or 0.0), actor=actor)
            if pid:
                ok += 1
            else:
                fail += 1
        except Exception:
            fail += 1
    return (ok, fail)


def calc_weekday_ot_from_duty(employee: str, y: int, m: int) -> dict:
    """
    ✅ 新版【平日(中晚)加班費】計算方式：

    1) 從【加班次數表】抓：員工姓名 + 年份 + 月份 → 時數
    2) 從【加班設定表】抓：年份 + 月份 → 加班時薪
    3) 金額 = 時數 * 加班時薪

    回傳：
      {"hours": float, "amount": float, "rule": {...}}
    """
    employee = (employee or "").strip()
    rule = get_overtime_rule(int(y), int(m))
    hourly_rate = float(rule.get("hourly_rate") or 0.0)

    if (not employee) or hourly_rate <= 0 or (not OVERTIME_COUNT_DB_ID):
        return {"hours": 0.0, "amount": 0.0, "rule": rule}

    hours = float(get_overtime_count_hours(employee, int(y), int(m)) or 0.0)
    amount = float(hours * hourly_rate)
    return {"hours": hours, "amount": amount, "rule": rule}
def upsert_duty_rows_to_notion(y: int, m: int, rows: list[dict]) -> None:
    """同月同日：有就更新，沒有就新增（適用你目前的『值班排班表』橫式 Notion DB）"""
    if not DUTY_DB_ID:
        raise RuntimeError("❌ 尚未設定 DUTY_DB_ID（值班排班表 Database ID）")

    # 讀 Notion DB 欄位 schema
    duty_props = get_db_properties(DUTY_DB_ID) or {}
    if not duty_props:
        raise RuntimeError("❌ 讀不到 DUTY_DB_ID 的 properties，請確認權限與 DB ID 是否正確")

    def _prop_type(name: str) -> str | None:
        p = duty_props.get(name)
        return (p or {}).get("type")

    def _rt(val: str):
        return {"rich_text": [{"text": {"content": str(val)}}]}

    def _title(val: str):
        return {"title": [{"text": {"content": str(val)}}]}

    # 你 Notion DB 截圖的欄位（以你實際 DB 為準）
    # 注意：如果你的 DB 名稱跟這裡不一樣，會在下面噴錯並列出實際欄位清單
    REQUIRED = ["員工姓名", "年份", "月份", "日期", "星期", "檢驗線(中)", "檢驗線(晚)", "收費員(中)", "收費員(晚)"]
    missing = [k for k in REQUIRED if k not in duty_props]
    if missing:
        all_keys = "、".join(duty_props.keys())
        raise RuntimeError(
            "❌ 值班排班表 DB 欄位名稱對不上，缺少："
            + "、".join(missing)
            + "\n\n✅ 你這個 DB 目前實際欄位有：\n"
            + all_keys
            + "\n\n👉 請把程式裡使用的欄位名稱改成跟 Notion 完全一致（含括號/空白/全形半形）"
        )

    # 判斷「日期」欄位型態（你的截圖是 rich_text，但我做成自動判斷）
    date_type = _prop_type("日期")  # "rich_text" or "date" ...
    weekday_type = _prop_type("星期")

    # shift 欄位（你 UI 裡是 multiselect/list，但 Notion 這邊多半是 rich_text）
    SHIFT_FIELDS = ["檢驗線(中)", "檢驗線(晚)", "收費員(中)", "收費員(晚)", "打掃工作"]
    note_exists = ("備註" in duty_props)

    ok, fail = 0, 0
    errors: list[str] = []

    for i, r in enumerate(rows, start=1):
        try:
            # 允許你的 row 來源欄位是「日期」或「日期(字串)」
            # 這裡以你 dataframe 的 "日期" 欄位為主
            date_str = str(r.get("日期", "")).strip()  # e.g. "2026-02-13" 或 "2/13"
            weekday_str = str(r.get("星期", "")).strip()

            if not date_str:
                raise RuntimeError("row 缺少『日期』")

            # Notion Title：你目前 DB 第一欄叫「員工姓名」(title)，但其實你放日期更直覺
            # 如果你想 Title 顯示別的，改這行即可
            title_text = date_str

            props_payload = {
                "員工姓名": _title(title_text),
                "年份": {"number": int(y)},
                "月份": {"number": int(m)},
            }

            # 日期：依 DB 型態寫入
            if date_type == "date":
                props_payload["日期"] = {"date": {"start": date_str}}
            else:
                props_payload["日期"] = _rt(date_str)

            # 星期：依 DB 型態寫入（你截圖是 rich_text）
            if weekday_type == "select":
                props_payload["星期"] = {"select": {"name": weekday_str}}
            else:
                props_payload["星期"] = _rt(weekday_str)

            # shift 欄位：list -> "A, B, C"
            for f in SHIFT_FIELDS:
                if f not in duty_props:
                    continue
                v = r.get(f, "")
                if isinstance(v, list):
                    v = ", ".join([str(x).strip() for x in v if str(x).strip()])
                else:
                    v = str(v).strip()

                # 依欄位型態寫入（大多是 rich_text）
                t = _prop_type(f)
                if t == "multi_select":
                    # 若你 DB 真的是 multi_select，就用 multi_select 寫
                    names = [s.strip() for s in v.split(",") if s.strip()]
                    props_payload[f] = {"multi_select": [{"name": n} for n in names]}
                elif t == "select":
                    props_payload[f] = {"select": {"name": v}} if v else {"select": None}
                else:
                    props_payload[f] = _rt(v)

            if note_exists:
                note = str(r.get("備註", "")).strip()
                if note:
                    props_payload["備註"] = _rt(note)

            # ---- 查同一天是否已存在（年份+月份+日期）----
            # 日期如果是 rich_text，用 rich_text equals；如果是 date，用 date equals
            date_filter = (
                {"property": "日期", "date": {"equals": date_str}}
                if date_type == "date"
                else {"property": "日期", "rich_text": {"equals": date_str}}
            )

            res = notion.databases.query(
                database_id=DUTY_DB_ID,
                page_size=5,
                filter={
                    "and": [
                        {"property": "年份", "number": {"equals": int(y)}},
                        {"property": "月份", "number": {"equals": int(m)}},
                        date_filter,
                    ]
                },
            )

            results = res.get("results", []) or []
            if results:
                # update
                page_id = results[0]["id"]
                notion.pages.update(page_id=page_id, properties=props_payload)
            else:
                # create
                notion.pages.create(parent={"database_id": DUTY_DB_ID}, properties=props_payload)

            ok += 1

        except Exception as e:
            fail += 1
            errors.append(f"第 {i} 天寫入失敗：{e}")

    if errors:
        # 直接把錯誤集中丟出去，讓你前端一次看到
        raise RuntimeError("\n".join(errors) + f"\n\n✅ 成功 {ok} 筆，❌ 失敗 {fail} 筆")



# =========================
# 🕒 OVERTIME：加班設定表（管理員用）
# =========================
def _get_default_shift_options() -> list[str]:
    # ✅ 與值班排班「班別」一致（你目前固定是這四個）
    return ["收費員(中)", "收費員(晚)", "檢驗線(中)", "檢驗線(晚)"]


def upsert_overtime_rule_to_notion(
    y: int,
    m: int,
    shift_hours: float,
    hourly_rate: float,
    note: str = "",
) -> str:
    """同年同月：有就更新，沒有就新增（加班設定表）。回傳 page_id。"""
    if not OVERTIME_RULE_DB_ID:
        raise RuntimeError("尚未設定 OVERTIME_RULE_DB_ID（加班設定表 DB ID）")

    # 讀 DB schema（避免欄位型別不一致）
    db = notion.databases.retrieve(database_id=OVERTIME_RULE_DB_ID)
    props = (db or {}).get("properties", {}) or {}

    def _ptype(name: str) -> str | None:
        return (props.get(name) or {}).get("type")

    def _rt(val: str):
        return {"rich_text": [{"text": {"content": str(val)}}]}

    def _title(val: str):
        return {"title": [{"text": {"content": str(val)}}]}

    def _num(val):
        try:
            return {"number": float(val)}
        except Exception:
            return {"number": None}

    def _ms(vals: list[str]):
        return {"multi_select": [{"name": str(v)} for v in vals if str(v).strip()]}

    def _sel(val: str):
        return {"select": {"name": str(val)}} if str(val).strip() else {"select": None}

    def _set(name: str, value):
        t = _ptype(name)
        if t == "title":
            return _title(value)
        if t == "number":
            return _num(value)
        if t == "rich_text":
            return _rt(value)
        if t == "select":
            return _sel(value)
        if t == "multi_select":
            # value 可能是 list[str] 或字串
            if isinstance(value, (list, tuple)):
                return _ms(list(value))
            return _ms([str(value)])
        # fallback：當作 rich_text
        return _rt(value)

    name = f"{int(y)}-{int(m):02d}"

    payload = {
        "名稱": _set("名稱", name),
        "年份": _set("年份", int(y)),
        "月份": _set("月份", int(m)),
        "班次換算時數": _set("班次換算時數", shift_hours),
        "加班時薪": _set("加班時薪", hourly_rate),
        "備註": _set("備註", note or ""),
    }

    # 只送 DB 真的存在的欄位（避免 Notion 噴錯）
    payload = {k: v for k, v in payload.items() if k in props}

    # 查同年同月是否已存在
    res = notion.databases.query(
        database_id=OVERTIME_RULE_DB_ID,
        page_size=5,
        filter={
            "and": [
                {"property": "年份", "number": {"equals": int(y)}},
                {"property": "月份", "number": {"equals": int(m)}},
            ]
        },
    )
    results = (res or {}).get("results", []) or []
    if results:
        page_id = results[0]["id"]
        notion.pages.update(page_id=page_id, properties=payload)
        return page_id

    created = notion.pages.create(
        parent={"database_id": OVERTIME_RULE_DB_ID},
        properties=payload,
    )
    return (created or {}).get("id", "")


def query_duty_month_to_horizontal_df(y: int, m: int, employees: list[str]):
    """
    ✅ 依你目前 Notion【值班排班表】的欄位結構查詢（一天一筆）：
    - 年份(number)、月份(number)、日期(文字或數字) 、星期(文字)
    - 班別欄位本身就是：檢驗線(中)、檢驗線(晚)、收費員(中)、收費員(晚)、打掃工作、備註
    - 每格可能是「多個人名」，用 、 / , / 空白 / 換行 分隔

    回傳 DataFrame（日期 1..月底），班別欄位為 list[str]（方便後續統計）
    """
    import pandas as pd

    # 先建空表（日期 1..月底）
    days = calendar.monthrange(int(y), int(m))[1]
    base_rows = []
    for d in range(1, days + 1):
        dt = date(int(y), int(m), d)
        base_rows.append({
            "日期": d,
            "星期": WEEKDAY_MAP[dt.weekday()],
            "檢驗線(中)": [],
            "檢驗線(晚)": [],
            "收費員(中)": [],
            "收費員(晚)": [],
            "打掃工作": [],
            "備註": "",
        })
    idx = {r["日期"]: i for i, r in enumerate(base_rows)}

    # 從 Notion 拉本月所有 rows（字串欄位）
    notion_rows = query_duty_rows_from_notion(int(y), int(m))

    def _split_names(s: str) -> list[str]:
        s = (s or "").strip()
        if not s:
            return []
        for sep in ["、", ",", "，", ";", "；", "\n", "\t"]:
            s = s.replace(sep, " ")
        parts = [p.strip() for p in s.split(" ") if p.strip()]
        seen = set()
        out = []
        for p in parts:
            if p not in seen:
                seen.add(p)
                out.append(p)
        return out

    for r in notion_rows:
        d = r.get("日期")
        try:
            d = int(d)
        except Exception:
            d = None
        if not d or d not in idx:
            continue
        i = idx[d]

        wk = (r.get("星期") or "").strip()
        if wk:
            base_rows[i]["星期"] = wk

        for col in ["檢驗線(中)", "檢驗線(晚)", "收費員(中)", "收費員(晚)", "打掃工作"]:
            base_rows[i][col] = _split_names(r.get(col, ""))

        note = (r.get("備註") or "").strip()
        if note:
            base_rows[i]["備註"] = note

    return pd.DataFrame(base_rows)

def normalize_multi_people_cell(v):
    """把 data_editor/Notion 回來的值，統一轉成 list[str]，並處理 NaN/NA"""
    try:
        import pandas as pd
        if v is None or (isinstance(v, float) and math.isnan(v)) or (hasattr(pd, "isna") and pd.isna(v)):
            return []
    except Exception:
        if v is None:
            return []

    if isinstance(v, list):
        return [str(x).strip() for x in v if str(x).strip()]

    if isinstance(v, str):
        s = v.strip()
        if not s:
            return []
        parts = re.split(r"[、,，/]+|\s+", s)
        return [p.strip() for p in parts if p.strip()]

    s = str(v).strip()
    return [s] if s else []


def coerce_duty_df_list_columns(df):
    """確保值班欄位永遠是 list，避免被 Streamlit 當成文字欄位"""
    for c in DUTY_SHIFT_COLUMNS:
        if c in df.columns:
            df[c] = df[c].apply(normalize_multi_people_cell)
    return df


# =========================
# ✅（新增）薪資表：自動偵測「餐食」欄位真正名稱（避免欄位名/型態不符寫不進去）
# =========================
@st.cache_data(ttl=60)
def resolve_salary_food_prop_name() -> str | None:
    """
    自動找出 Notion 薪資表中「餐食」欄位真正的名稱（避免你 Notion 叫 餐費/午餐/工餐 等導致寫不進去）
    ✅ 只接受 type=number 的欄位。
    """
    props = get_db_properties(SALARY_DB_ID) or {}

    candidates = ["餐食", "餐費", "午餐", "工餐", "午餐差額", "餐食費", "餐食金額", "餐費差額"]

    for name in candidates:
        meta = props.get(name)
        if not meta:
            continue
        if meta.get("type") == "number":
            return name

    return None


# =========================
# ✅ 操作記錄表：寫入 / 讀取
# =========================
def log_action(employee_name: str, action_type: str, action_content: str, result: str):
    """寫入「操作記錄表」：不強制欄位型態，盡力填入可用欄位。
    ✅ 重點：
    - 盡力寫入 title 欄位（Notion DB 必有），避免出現「空白列」
    - 如果抓不到 schema，也會用常見欄位名稱做 fallback 寫入（至少要留下一筆可追蹤紀錄）
    """
    if not OPLOG_DB_ID:
        return

    emp = (employee_name or "").strip() or "—"
    act = (action_type or "").strip() or "—"
    content = (action_content or "").strip() or "—"
    res_txt = (result or "").strip() or "—"

    try:
        props_meta = get_db_properties(OPLOG_DB_ID) or {}
        props: dict = {}

        # 1) title 欄位（schema 有→找出 title 名稱；沒有→預設用「員工姓名」當 title）
        title_prop = _first_title_prop_name(props_meta) or "員工姓名" or "員工姓名"
        title_value = emp or act or "—"
        props[title_prop] = {"title": [{"text": {"content": title_value}}]}

        now_iso = datetime.now().isoformat()

        if props_meta:
            # 2) schema 存在：用既有 helper 盡力寫入
            _best_set_text(props, props_meta, "員工姓名", emp)
            _best_set_text(props, props_meta, "操作類型", act)
            _best_set_text(props, props_meta, "操作內容", content)

            # 操作結果（常見：select）
            meta_r = (props_meta.get("操作結果") or {})
            if meta_r.get("type") == "select" and res_txt:
                props["操作結果"] = {"select": {"name": res_txt}}
            else:
                _best_set_text(props, props_meta, "操作結果", res_txt)

            # 操作時間（常見：date）
            meta_t = (props_meta.get("操作時間") or {})
            if meta_t.get("type") == "date":
                props["操作時間"] = {"date": {"start": now_iso}}
        else:
            # 3) schema 取不到：用「常見欄位名稱」直接寫入（盡量不要再產生空白列）
            #    這些欄位若不存在或型態不同，Notion 會拒絕；因此這裡用 try/catch 包住
            try:
                props.setdefault("操作類型", {"rich_text": [{"text": {"content": act}}]})
                props.setdefault("操作內容", {"rich_text": [{"text": {"content": content}}]})
                # 操作結果常見是 select；若 DB 不是 select 會報錯，但至少 title 仍在
                props.setdefault("操作結果", {"select": {"name": res_txt}})
                props.setdefault("操作時間", {"date": {"start": now_iso}})
            except Exception:
                pass

        notion.pages.create(database_id=OPLOG_DB_ID, properties=props)

    except Exception as e:
        if os.getenv("DEBUG_NOTION", "").strip() == "1":
            st.error(f"❌ 寫入操作記錄失敗：{e}")
        return



def list_operation_logs(limit: int = 200):
    if not OPLOG_DB_ID:
        return []

    try:
        props_meta = get_db_properties(OPLOG_DB_ID) or {}
        op_time_meta = props_meta.get("操作時間", {}) or {}
        op_time_type = op_time_meta.get("type")

        if op_time_type == "created_time":
            sorts = [{"timestamp": "created_time", "direction": "descending"}]
        elif op_time_type == "last_edited_time":
            sorts = [{"timestamp": "last_edited_time", "direction": "descending"}]
        elif "操作時間" in props_meta:
            sorts = [{"property": "操作時間", "direction": "descending"}]
        else:
            sorts = [{"timestamp": "created_time", "direction": "descending"}]

        query = {
            "database_id": OPLOG_DB_ID,
            "page_size": min(int(limit), 100),
            "sorts": sorts,
        }
        res = notion.databases.query(**query)

        def fmt_time(s: str) -> str:
            if not s:
                return ""
            try:
                dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
                if dt.tzinfo is not None:
                    dt = dt.astimezone(timezone(timedelta(hours=8)))
                return dt.strftime("%Y-%m-%d %H:%M")
            except Exception:
                return s

        rows = []
        for page in res.get("results", []):
            props = page.get("properties", {}) or {}

            def get_op_time() -> str:
                p = props.get("操作時間", {}) or {}
                d = p.get("date")
                if d and d.get("start"):
                    return fmt_time(d.get("start", ""))
                ct = p.get("created_time")
                if ct:
                    return fmt_time(ct)
                lt = p.get("last_edited_time")
                if lt:
                    return fmt_time(lt)
                return fmt_time(page.get("created_time", ""))

            rows.append({
                "員工姓名": _get_prop_plain_text(props.get("員工姓名", {})),
                "操作類型": _get_prop_plain_text(props.get("操作類型", {})),
                "操作內容": _get_prop_plain_text(props.get("操作內容", {})),
                "操作結果": _get_prop_plain_text(props.get("操作結果", {})),
                "操作時間": get_op_time(),
            })

        return rows

    except Exception as e:
        st.error(f"讀取操作記錄失敗：{e}")
        return []


def update_password_and_logout(username: str, old_pwd: str, new_pwd: str, force: bool = False) -> bool:
    username = (username or "").strip()
    old_pwd = (old_pwd or "").strip()
    new_pwd = (new_pwd or "").strip()

    if not username or not new_pwd:
        st.error("❌ 參數不足")
        return False

    if len(new_pwd) < 6:
        st.error("❌ 新密碼至少 6 碼（你可以自行調整規則）")
        return False

    page = get_account_page_by_username(username)
    if not page:
        st.error("❌ 找不到帳號資料")
        return False

    page_id = page["id"]
    props = page.get("properties", {}) or {}

    login_hash = _get_prop_plain_text(props.get("login_hash", {}))
    legacy_pwd = _get_prop_plain_text(props.get("密碼", {}))

    if not force:
        ok_old = False
        if login_hash:
            ok_old = verify_password_bcrypt(old_pwd, login_hash)
        else:
            ok_old = (old_pwd == legacy_pwd)

        if not ok_old:
            st.error("❌ 舊密碼不正確")
            return False

    if (not force) and old_pwd and (old_pwd == new_pwd):
        st.error("❌ 新密碼不可與舊密碼相同")
        return False

    new_hash = hash_password_bcrypt(new_pwd)

    props_to_update = {
        "login_hash": {"rich_text": [{"text": {"content": new_hash}}]},
        "must_change_password": {"checkbox": False},
    }

    if "密碼" in (get_db_properties(ACCOUNT_DB_ID) or {}):
        props_to_update["密碼"] = {"rich_text": []}

    if "last_password_change" in (get_db_properties(ACCOUNT_DB_ID) or {}):
        props_to_update["last_password_change"] = {"date": {"start": datetime.now().isoformat()}}

    try:
        notion.pages.update(page_id=page_id, properties=props_to_update)
        log_action(username, "更改密碼", "更改密碼成功（已寫入 login_hash）", "成功")
        return True
    except Exception as e:
        st.error(f"❌ 寫回 Notion 失敗：{e}")
        log_action(username, "更改密碼", f"寫回 Notion 失敗：{e}", "系統錯誤")
        return False


@st.dialog("🔒 更改密碼")
def change_password_dialog(force: bool = False):
    user = st.session_state.get("user", "")

    st.caption("改完密碼會立刻登出，請用新密碼重新登入。")

    if not force:
        old_pwd = st.text_input("舊密碼", type="password")
    else:
        old_pwd = ""

    new_pwd = st.text_input("新密碼", type="password")
    new_pwd2 = st.text_input("確認新密碼", type="password")

    c1, c2 = st.columns(2)
    if c1.button("✅ 儲存新密碼", use_container_width=True):
        if new_pwd != new_pwd2:
            st.error("❌ 兩次輸入的新密碼不一致")
            return

        ok = update_password_and_logout(
            username=user,
            old_pwd=old_pwd,
            new_pwd=new_pwd,
            force=force,
        )
        if ok:
            st.session_state["logged_in"] = False
            st.session_state["user"] = ""
            st.session_state["is_admin"] = False
            st.session_state["force_change_pwd"] = False
            st.success("✅ 密碼已更新，請重新登入")
            time.sleep(0.6)
            st.rerun()

    if c2.button("取消", use_container_width=True, disabled=force):
        st.rerun()


# =========================
# 1) Notion 登入驗證
# =========================
def login(username: str, password: str):
    username = (username or "").strip()
    password = (password or "").strip()

# ---- deploy debug ----
deploy_debug = bool(st.session_state.get("deploy_debug", False))
debug_info = {
    "ts": datetime.now().isoformat(),
    "username": username,
    "has_password": bool(password),
    "ACCOUNT_DB_ID_set": bool(ACCOUNT_DB_ID),
    "OPLOG_DB_ID_set": bool(OPLOG_DB_ID),
}
if deploy_debug:
    st.session_state["login_debug"] = debug_info

    if not username or not password:
        log_action(username or "—", "登入", "帳號或密碼為空", "失敗")
        debug_info.update({"stage":"empty_credentials"})
        if deploy_debug: st.session_state["login_debug"] = debug_info
        return False, False, False

    try:
        page = get_account_page_by_username(username)
        if not page:
            log_action(username, "登入", "找不到帳號", "失敗")
            debug_info.update({"stage":"no_account_page"})
            if deploy_debug: st.session_state["login_debug"] = debug_info
            return False, False, False

        page_id = page["id"]
        props = page.get("properties", {}) or {}

        sel = (props.get("權限", {}) or {}).get("select")
        role = sel.get("name") if sel else None
        is_admin = (role == "管理員")

        login_hash = _get_prop_plain_text(props.get("login_hash", {}))
        legacy_pwd = _get_prop_plain_text(props.get("密碼", {}))
        must_change_flag = bool((props.get("must_change_password", {}) or {}).get("checkbox") or False)

        debug_info.update({
            "stage":"loaded_account_page",
            "page_id": page_id,
            "role": role,
            "is_admin": is_admin,
            "login_hash_len": len(login_hash) if isinstance(login_hash, str) else None,
            "login_hash_preview": (login_hash[:12] + "..." + login_hash[-6:]) if isinstance(login_hash, str) and len(login_hash) > 20 else login_hash,
            "legacy_pwd_len": len(legacy_pwd) if isinstance(legacy_pwd, str) else None,
            "must_change_flag": bool(must_change_flag),
        })
        if deploy_debug: st.session_state["login_debug"] = debug_info

        used_legacy = False

        if login_hash:
            ok = verify_password_bcrypt(password, login_hash)
        else:
            ok = (password == legacy_pwd)
            used_legacy = bool(ok)

        if not ok:
            log_action(username, "登入", "帳號或密碼錯誤", "失敗")
            debug_info.update({"stage":"password_verify_failed", "used_legacy": used_legacy})
            if deploy_debug: st.session_state["login_debug"] = debug_info
            return False, False, False

        try:
            notion.pages.update(
                page_id=page_id,
                properties={"最後登入時間": {"date": {"start": datetime.now().isoformat()}}},
            )
        except Exception:
            pass

        must_change = bool(must_change_flag or used_legacy)

        log_action(username, "登入", "登入成功", "成功")
        debug_info.update({"stage":"login_success", "must_change": must_change, "used_legacy": used_legacy})
        if deploy_debug: st.session_state["login_debug"] = debug_info
        return True, is_admin, must_change

    except Exception as e:
        st.error(f"Notion 登入驗證失敗：{e}")
        log_action(username, "登入", f"Notion 驗證例外：{e}", "系統錯誤")
        debug_info.update({"stage":"exception", "error": str(e)})
        if deploy_debug: st.session_state["login_debug"] = debug_info
        return False, False, False


# =========================
# 2b) 請假：更新狀態 / 刪除
# =========================
def update_leave_status(page_id: str, new_status: str, actor: str = "") -> bool:
    try:
        notion.pages.update(page_id=page_id, properties={"狀態": {"select": {"name": new_status}}})
        log_action(actor or "—", "請假審核", f"更新請假狀態為：{new_status}", "成功")
        return True
    except Exception as e:
        st.error(f"更新狀態失敗：{e}")
        log_action(actor or "—", "請假審核", f"更新狀態失敗：{e}", "系統錯誤")
        return False


def delete_leave_request(page_id: str, actor: str = "") -> bool:
    try:
        notion.pages.update(page_id=page_id, archived=True)
        log_action(actor or "—", "請假管理", "刪除（封存）請假紀錄", "成功")
        return True
    except Exception as e:
        st.error(f"刪除（封存）失敗：{e}")
        log_action(actor or "—", "請假管理", f"刪除（封存）失敗：{e}", "系統錯誤")
        return False


def make_leave_label(row: dict) -> str:
    return f"{row.get('員工姓名','')}｜{row.get('假別','')}｜{row.get('請假期間','')}｜{row.get('狀態','')}"


# =========================
# 3) Notion Date 解析/格式化
# =========================
def parse_notion_date(props: dict, prop_name: str) -> tuple[datetime | None, datetime | None, str]:
    d = (props.get(prop_name, {}) or {}).get("date")
    if not d:
        return None, None, ""

    start_s = d.get("start")
    end_s = d.get("end")

    def to_dt(s: str | None) -> datetime | None:
        if not s:
            return None
        try:
            dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
            # FIX: 轉台灣時區顯示，避免日期偏移
            if dt.tzinfo is not None:
                dt = dt.astimezone(timezone(timedelta(hours=8)))
            return dt
        except Exception:
            return None

    start_dt = to_dt(start_s)
    end_dt = to_dt(end_s)

    def fmt(dt: datetime | None) -> str:
        if not dt:
            return ""
        # FIX: 原本強制 ":00" 容易誤導，改正常顯示到分鐘
        return dt.strftime("%Y-%m-%d %H:%M")

    if start_dt and end_dt:
        display = f"{fmt(start_dt)} ~ {fmt(end_dt)}"
    elif start_dt:
        display = fmt(start_dt)
    else:
        display = ""

    return start_dt, end_dt, display


# =========================
# ✅ 特休折算表（Notion）：讀取/新增/覆蓋（Upsert）
# =========================
@st.cache_data(ttl=60)
def get_cashout_rule_by_year(year: int) -> dict | None:
    if not CASHOUT_RULE_DB_ID:
        return None

    try:
        res = notion.databases.query(
            database_id=CASHOUT_RULE_DB_ID,
            filter={"property": "年份", "number": {"equals": int(year)}},
            page_size=1,
        )
        results = res.get("results", [])
        if not results:
            return None

        page = results[0]
        props = page.get("properties", {}) or {}

        def n(name: str) -> float:
            return float((props.get(name, {}) or {}).get("number") or 0.0)

        return {
            "_page_id": page["id"],
            "年份": int(year),
            "可折算天數": n("可折算天數"),
            "一天時數": n("一天時數"),
            "一小時折算金額": n("一小時折算金額"),
        }

    except Exception as e:
        st.error(f"讀取特休折算規則失敗：{e}")
        return None


def upsert_cashout_rule(year: int, cap_days: float, hours_per_day: float, amount_per_hour: float, actor: str = "") -> bool:
    if not CASHOUT_RULE_DB_ID:
        st.error("❌ 尚未設定 CASHOUT_RULE_DB_ID（特休折算表 Database ID）")
        return False

    year = int(year)
    cap_days = float(cap_days or 0.0)
    hours_per_day = float(hours_per_day or 0.0)
    amount_per_hour = float(amount_per_hour or 0.0)

    if hours_per_day <= 0:
        st.error("❌ 一天時數必須 > 0")
        return False

    try:
        existing = get_cashout_rule_by_year(year)

        props = {
            "年份": {"number": int(year)},
            "可折算天數": {"number": float(cap_days)},
            "一天時數": {"number": float(hours_per_day)},
            "一小時折算金額": {"number": float(amount_per_hour)},
        }

        if existing and existing.get("_page_id"):
            notion.pages.update(page_id=existing["_page_id"], properties=props)
            log_action(actor or "—", "特休折算規則", f"覆蓋年度規則：{year}", "成功")
        else:
            notion.pages.create(parent={"database_id": CASHOUT_RULE_DB_ID}, properties=props)
            log_action(actor or "—", "特休折算規則", f"新增年度規則：{year}", "成功")

        try:
            get_cashout_rule_by_year.clear()
        except Exception:
            pass

        return True

    except Exception as e:
        st.error(f"寫入特休折算規則失敗：{e}")
        log_action(actor or "—", "特休折算規則", f"寫入失敗：{year}｜{e}", "系統錯誤")
        return False


def calc_cashout(remaining_hours: float, hours_per_day: float, cap_days: float, amount_per_day: float, whole_days_only: bool):
    remaining_hours = float(remaining_hours or 0.0)
    hours_per_day = float(hours_per_day or DEFAULT_HOURS_PER_DAY)
    cap_days = float(cap_days if cap_days is not None else 0.0)
    amount_per_day = float(amount_per_day or 0.0)

    if hours_per_day <= 0:
        hours_per_day = DEFAULT_HOURS_PER_DAY

    remaining_days = remaining_hours / hours_per_day

    if whole_days_only:
        raw_days = float(int(remaining_days))
    else:
        raw_days = float(round(remaining_days, 2))

    cap_days = max(0.0, cap_days)
    cashout_days = min(raw_days, cap_days)
    cashout_amount = cashout_days * amount_per_day

    return {
        "remaining_days": float(remaining_days),
        "cashout_days": float(cashout_days),
        "cashout_amount": float(cashout_amount),
    }


# =========================
# 4) 寫入 Notion【請假紀錄表】
# =========================
def create_leave_request(
    employee_name: str,
    leave_type: str,
    hours: int,
    start_dt: datetime,
    end_dt: datetime,
    reason: str,
    target_employee_name: str | None = None,
    created_by: str | None = None,
) -> bool:
    employee_name = (employee_name or "").strip()
    leave_type = (leave_type or "").strip()
    reason = (reason or "").strip()
    target_name = (target_employee_name or employee_name or "").strip()
    actor = (created_by or employee_name or "").strip()

    if not employee_name:
        st.error("❌ 找不到登入者姓名，請重新登入")
        log_action(actor or "—", "請假申請", "找不到登入者姓名", "失敗")
        return False
    if not target_name:
        st.error("❌ 請指定請假人（員工姓名）")
        log_action(actor or "—", "請假申請", "未指定請假人", "失敗")
        return False
    if hours is None or int(hours) <= 0:
        st.error("❌ 請假時數必須大於 0")
        log_action(actor or "—", "請假申請", "請假時數 <= 0", "失敗")
        return False
    if end_dt <= start_dt:
        st.error("❌ 結束時間必須晚於開始時間")
        log_action(actor or "—", "請假申請", "結束時間 <= 開始時間", "失敗")
        return False

    try:
        status_options = get_select_options(LEAVE_DB_ID, "狀態")
        default_status = "待審核" if "待審核" in status_options else (status_options[0] if status_options else "待審核")

        props = {
            "員工姓名": {"title": [{"text": {"content": target_name}}]},
            "假別": {"select": {"name": leave_type}},
            "請假時數": {"number": int(hours)},
            "請假期間": {"date": {"start": start_dt.isoformat(), "end": end_dt.isoformat()}},
            "請假事由": {"rich_text": [{"text": {"content": reason}}]},
            "狀態": {"select": {"name": default_status}},
        }

        notion.pages.create(parent={"database_id": LEAVE_DB_ID}, properties=props)
        log_action(actor or target_name, "請假申請", f"{target_name} 申請 {leave_type} {int(hours)} 小時", "成功")
        return True

    except Exception as e:
        st.error(f"請假申請寫入失敗：{e}")
        log_action(actor or target_name, "請假申請", f"寫入 Notion 失敗：{e}", "系統錯誤")
        return False


# =========================
# 5) 讀取請假清單
# =========================
def list_leave_requests(is_admin: bool, employee_name: str, limit: int = 50):
    try:
        props_meta = get_db_properties(LEAVE_DB_ID) or {}
        build_meta = props_meta.get("建立時間", {}) or {}
        build_type = build_meta.get("type")

        # FIX: created_time/last_edited_time 必須用 timestamp 排序，不是 property
        if build_type == "created_time":
            sorts = [{"timestamp": "created_time", "direction": "descending"}]
        elif build_type == "last_edited_time":
            sorts = [{"timestamp": "last_edited_time", "direction": "descending"}]
        else:
            # 若你「建立時間」是 date 才能用 property 排序
            sorts = [{"property": "建立時間", "direction": "descending"}]

        query = {
            "database_id": LEAVE_DB_ID,
            "page_size": min(limit, 100),
            "sorts": sorts,
        }

        if not is_admin:
            query["filter"] = {"property": "員工姓名", "title": {"equals": employee_name}}

        res = notion.databases.query(**query)
        rows = []

        for page in res.get("results", []):
            props = page["properties"]

            def get_title(name):
                v = props.get(name, {}).get("title", [])
                return v[0]["plain_text"] if v else ""

            def get_select(name):
                v = props.get(name, {}).get("select")
                return v.get("name") if v else ""

            def get_number(name):
                return props.get(name, {}).get("number")

            def get_rich(name):
                v = props.get(name, {}).get("rich_text", [])
                return v[0]["plain_text"] if v else ""

            _sdt, _edt, period_display = parse_notion_date(props, "請假期間")

            rows.append({
                "_page_id": page["id"],
                "員工姓名": get_title("員工姓名"),
                "假別": get_select("假別"),
                "請假時數": get_number("請假時數"),
                "請假期間": period_display,
                "請假事由": get_rich("請假事由"),
                "狀態": get_select("狀態"),
                "建立時間": props.get("建立時間", {}).get("created_time", page.get("created_time", "")),
                "最後更新時間": props.get("最後更新時間", {}).get("last_edited_time", page.get("last_edited_time", "")),
            })

        return rows

    except Exception as e:
        st.error(f"讀取請假紀錄失敗：{e}")
        return []


# =========================
# 6) 年度特休：計算已用
# =========================
def calc_used_vacation_hours(employee_name: str, year: int) -> float:
    employee_name = (employee_name or "").strip()
    if not employee_name:
        return 0.0

    status_options = get_select_options(LEAVE_DB_ID, "狀態") or []
    approved_candidates = ["通過", "已通過", "核准", "已核准", "同意", "Approved"]
    approved_status = next((c for c in approved_candidates if c in status_options), None) or "通過"

    try:
        res = notion.databases.query(
            database_id=LEAVE_DB_ID,
            filter={
                "and": [
                    {"property": "員工姓名", "title": {"equals": employee_name}},
                    {"property": "假別", "select": {"equals": "特休"}},
                    {"property": "狀態", "select": {"equals": approved_status}},
                ]
            },
            page_size=100,
        )

        total = 0.0
        for page in res.get("results", []):
            props = page["properties"]
            start_dt, _end_dt, _display = parse_notion_date(props, "請假期間")
            if not start_dt:
                continue
            if int(start_dt.year) != int(year):
                continue
            hours = props.get("請假時數", {}).get("number") or 0
            total += float(hours)

        return float(total)

    except Exception as e:
        st.error(f"計算已用特休失敗：{e}")
        return 0.0


# =========================
# 7) 員工清單
# =========================
def list_employee_names(limit: int = 200):
    try:
        res = notion.databases.query(database_id=ACCOUNT_DB_ID, page_size=min(limit, 100))
        names = []
        for page in res.get("results", []):
            props = page["properties"]
            t = props.get("員工姓名", {}).get("title", [])
            name = t[0]["plain_text"].strip() if t else ""
            if name:
                names.append(name)
        return sorted(list(set(names)))
    except Exception as e:
        st.error(f"讀取員工清單失敗：{e}")
        return []


# =========================
# 8) 年度特休：讀取/初始化/快照/不足阻擋
# =========================
def list_vacation_summary(is_admin: bool, employee_name: str, year: int | None = None, limit: int = 200):
    try:
        query = {
            "database_id": VACATION_DB_ID,
            "page_size": min(limit, 100),
            "sorts": [{"property": "年度", "direction": "descending"}],
        }

        filters = []
        if not is_admin:
            filters.append({"property": "員工姓名", "title": {"equals": employee_name}})
        if year is not None:
            filters.append({"property": "年度", "number": {"equals": int(year)}})

        if filters:
            query["filter"] = {"and": filters} if len(filters) > 1 else filters[0]

        res = notion.databases.query(**query)
        rows = []

        for page in res.get("results", []):
            props = page["properties"]

            def get_title(name):
                v = props.get(name, {}).get("title", [])
                return v[0]["plain_text"] if v else ""

            def get_number(name):
                return props.get(name, {}).get("number")

            name = get_title("員工姓名")
            y = get_number("年度")
            total = get_number("本年度特休時數") or 0
            used = get_number("已使用特休時數") or 0
            remaining_field = get_number("剩餘特休時數")
            remaining = remaining_field if remaining_field is not None else max(0.0, float(total) - float(used))

            rows.append({
                "_page_id": page["id"],
                "員工姓名": name,
                "年度": int(y) if y is not None else None,
                "本年度特休時數": float(total),
                "已使用特休時數": float(used),
                "剩餘特休時數": float(remaining),
            })

        return rows

    except Exception as e:
        st.error(f"讀取年度特休表失敗：{e}")
        return []


def ensure_vacation_row(employee_name: str, year: int, default_total: float = 0.0) -> bool:
    employee_name = (employee_name or "").strip()
    if not employee_name:
        return False

    try:
        res = notion.databases.query(
            database_id=VACATION_DB_ID,
            filter={
                "and": [
                    {"property": "員工姓名", "title": {"equals": employee_name}},
                    {"property": "年度", "number": {"equals": int(year)}},
                ]
            },
            page_size=1,
        )
        if res.get("results"):
            return True

        notion.pages.create(
            parent={"database_id": VACATION_DB_ID},
            properties={
                "員工姓名": {"title": [{"text": {"content": employee_name}}]},
                "年度": {"number": int(year)},
                "本年度特休時數": {"number": float(default_total)},
                "已使用特休時數": {"number": 0.0},
                "剩餘特休時數": {"number": float(default_total)},
            },
        )
        return True

    except Exception as e:
        st.error(f"初始化年度特休資料失敗：{e}")
        return False


def get_employee_vacation_snapshot(employee_name: str, year: int) -> dict | None:
    ok = ensure_vacation_row(employee_name, year, default_total=0.0)
    if not ok:
        return None

    rows = list_vacation_summary(is_admin=False, employee_name=employee_name, year=year, limit=5)
    if not rows:
        return None

    row = rows[0]
    used = calc_used_vacation_hours(employee_name, year)
    total = float(row.get("本年度特休時數", 0.0) or 0.0)
    remaining = max(0.0, total - used)

    return {
        "employee": employee_name,
        "year": int(year),
        "total": total,
        "used": float(used),
        "remaining": float(remaining),
        "_page_id": row.get("_page_id"),
    }


def validate_vacation_enough(employee_name: str, year: int, request_hours: int) -> tuple[bool, str]:
    snap = get_employee_vacation_snapshot(employee_name, year)
    if not snap:
        return False, "❌ 讀取年度特休資料失敗"

    remaining = float(snap["remaining"])
    if float(request_hours) > remaining:
        return False, f"❌ 特休不足：剩餘 {remaining:.0f} 小時，但你申請 {int(request_hours)} 小時"
    return True, ""


# =========================
# ✅ 9) 薪資表：讀取/新增/更新（12月才算特休折算）+ 🍱 當月餐食費
# =========================
def salary_calc_payable(
    base_salary: float,
    overtime_pay: float,
    bonus: float,
    leave_deduction: float,
    advance_other: float,
    lunch_amount: float,
    vacation_cashout_amount: float,
    include_cashout: bool,
) -> float:
    total = float(base_salary or 0) + float(overtime_pay or 0) + float(bonus or 0)
    total -= float(leave_deduction or 0)
    total -= float(advance_other or 0)
    total += float(lunch_amount or 0)
    if include_cashout:
        total += float(vacation_cashout_amount or 0)
    return float(total)


@st.cache_data(ttl=60)
def get_month_lunch_amount(employee_name: str, y: int, m: int, is_admin: bool) -> float:
    try:
        if (not LUNCH_DB_ID) or (not ATTEND_DB_ID):
            return 0.0
        s = calc_month_lunch_settlement(employee_name, int(y), int(m), is_admin=is_admin)
        return float(s.get("差額(應得-已訂餐)", 0) or 0.0)
    except Exception:
        return 0.0


def get_salary_record(employee_name: str, y: int, m: int) -> dict | None:
    employee_name = (employee_name or "").strip()
    if not employee_name:
        return None

    try:
        res = notion.databases.query(
            database_id=SALARY_DB_ID,
            filter={
                "and": [
                    {"property": "員工姓名", "title": {"equals": employee_name}},
                    {"property": "薪資年份", "number": {"equals": int(y)}},
                    {"property": "薪資月份", "number": {"equals": int(m)}},
                ]
            },
            page_size=1,
        )
        results = res.get("results", [])
        if not results:
            return None

        page = results[0]
        props = page.get("properties", {}) or {}

        # ---------- 小工具：抓 Notion 值 ----------
        def _find_prop_key(prefix: str) -> str | None:
            """用前綴找欄位（避免欄位被改名或加上括號備註）"""
            for k in props.keys():
                if isinstance(k, str) and k.startswith(prefix):
                    return k
            return None

        def _pick_key(candidates: list[str], prefix: str | None = None) -> str | None:
            """優先精準命中，其次用 prefix 模糊命中"""
            for k in candidates:
                if k in props:
                    return k
            if prefix:
                k2 = _find_prop_key(prefix)
                if k2:
                    return k2
            return candidates[0] if candidates else None

        def get_title(name: str) -> str:
            k = _pick_key([name], prefix=name)
            if not k:
                return ""
            v = (props.get(k, {}) or {}).get("title", []) or []
            return v[0].get("plain_text", "") if v else ""

        def get_number(name: str, *, candidates: list[str] | None = None, prefix: str | None = None) -> float:
            key_list = candidates if candidates else [name]
            k = _pick_key(key_list, prefix=prefix or name)
            if not k:
                return 0.0
            v = (props.get(k, {}) or {}).get("number")
            try:
                return float(v or 0.0)
            except Exception:
                return 0.0

        def get_rich_text(name: str) -> str:
            k = _pick_key([name], prefix=name)
            if not k:
                return ""
            v = (props.get(k, {}) or {}).get("rich_text", []) or []
            return v[0].get("plain_text", "") if v else ""

        # ---------- 可留：發薪月份（若你 DB 還有這欄） ----------
        pay_date = None
        d = (props.get("發薪月份", {}) or {}).get("date")
        if d and d.get("start"):
            try:
                dt = datetime.fromisoformat(d["start"].replace("Z", "+00:00"))
                if dt.tzinfo is not None:
                    dt = dt.astimezone(timezone(timedelta(hours=8)))  # 台灣時區
                pay_date = dt.date()
            except Exception:
                pay_date = None

        # ---------- 新版欄位：加項 / 扣項 / 總計 ----------
        # 加項（照你 Notion 欄位）
        add_keys = [
            "全薪",
            "負責人職務津貼",
            "職務津貼",
            "績效獎金",
            "交通津貼",
            "營業津貼",
            "配合",
            "全勤獎金",
            "證照加給",
            "伙食津貼",
            "平日(中晚)加班費",
            "週六加班費",
            "交際費",
            "年終補助",
        ]

        # 扣項
        deduct_keys = [
            "借支",
            "病假請假",
            "事假請假",
            "借款利息",
            "遲到/早退",
            "勞保費",
            "健保費",
            "其他",
        ]

        # 總計
        total_keys = [
            "薪資總計",
            "應扣總計",
            "實發金額",
        ]

        data = {
            "_page_id": page.get("id"),
            "員工姓名": get_title("員工姓名"),
            "薪資年份": int(get_number("薪資年份") or 0),
            "薪資月份": int(get_number("薪資月份") or 0),
            "備註": get_rich_text("備註"),
            "發薪月份": pay_date,
            "建立時間": (props.get("建立時間", {}) or {}).get("created_time", page.get("created_time", "")),
            "最後更新時間": (props.get("最後更新時間", {}) or {}).get("last_edited_time", page.get("last_edited_time", "")),
        }

        # 寫入加項/扣項/總計數值
        for k in add_keys:
            data[k] = get_number(k, prefix=k)

        for k in deduct_keys:
            data[k] = get_number(k, prefix=k)

        for k in total_keys:
            data[k] = get_number(k, prefix=k)

        return data

    except Exception as e:
        st.error(f"讀取薪資資料失敗：{e}")
        return None



def upsert_salary_record(
    employee_name: str,
    y: int,
    m: int,

    # ✅ 兼容：讓你可以直接丟 data=payload（你 UI 現在就是這樣）
    data: dict | None = None,

    # ✅ 加項（對齊最新 Notion）
    full_salary: float = 0.0,        # 全薪
    leader_allowance: float = 0.0,   # 負責人職務津貼
    job_allowance: float = 0.0,      # 職務津貼
    perf_bonus: float = 0.0,         # 績效獎金
    traffic_allowance: float = 0.0,  # 交通津貼
    sales_allowance: float = 0.0,    # 營業津貼
    coop: float = 0.0,               # 配合
    attend_bonus: float = 0.0,       # 全勤獎金
    cert_allowance: float = 0.0,     # 證照加給
    meal_allowance: float = 0.0,     # 伙食津貼
    ot_weekday: float = 0.0,         # 平日(中晚)加班費
    ot_sat: float = 0.0,             # 週六加班費
    social_fee: float = 0.0,         # 交際費
    year_end: float = 0.0,           # 年終補助
    gross_total: float | None = None,# 薪資總計（可不傳，會自算）

    # ✅ 扣項（對齊最新 Notion）
    advance: float = 0.0,            # 借支
    sick_leave: float = 0.0,         # 病假請假
    personal_leave: float = 0.0,     # 事假請假
    loan_interest: float = 0.0,      # 借款利息
    late_early: float = 0.0,         # 遲到/早退
    labor_fee: float = 0.0,          # 勞保費
    health_fee: float = 0.0,         # 健保費
    other_ded: float = 0.0,          # 其他
    deduct_total: float | None = None,# 應扣總計（可不傳，會自算）
    net_pay: float | None = None,    # 實發金額（可不傳，會自算）

    note: str = "",
    actor: str = "",
) -> bool:

    # =========================
    # ✅ 兼容：若有傳 data，就從 data 映射到本函式欄位（以 Notion 欄位名為準）
    # =========================
    if isinstance(data, dict) and data:
        employee_name = (data.get("員工姓名") or employee_name or "").strip()
        y = int(data.get("薪資年份", y))
        m = int(data.get("薪資月份", m))

        def _as_float(v, default=0.0) -> float:
            try:
                if v is None or v == "":
                    return float(default)
                return float(v)
            except Exception:
                return float(default)

        full_salary = _as_float(data.get("全薪", full_salary))
        leader_allowance = _as_float(data.get("負責人職務津貼", leader_allowance))
        job_allowance = _as_float(data.get("職務津貼", job_allowance))
        perf_bonus = _as_float(data.get("績效獎金", perf_bonus))
        traffic_allowance = _as_float(data.get("交通津貼", traffic_allowance))
        sales_allowance = _as_float(data.get("營業津貼", sales_allowance))
        coop = _as_float(data.get("配合", coop))
        attend_bonus = _as_float(data.get("全勤獎金", attend_bonus))
        cert_allowance = _as_float(data.get("證照加給", cert_allowance))
        meal_allowance = _as_float(data.get("伙食津貼", meal_allowance))
        ot_weekday = _as_float(data.get("平日(中晚)加班費", ot_weekday))
        ot_sat = _as_float(data.get("週六加班費", ot_sat))
        social_fee = _as_float(data.get("交際費", social_fee))
        year_end = _as_float(data.get("年終補助", year_end))

        gross_total = data.get("薪資總計", gross_total)
        gross_total = None if gross_total is None else _as_float(gross_total, 0.0)

        advance = _as_float(data.get("借支", advance))
        sick_leave = _as_float(data.get("病假請假", sick_leave))
        personal_leave = _as_float(data.get("事假請假", personal_leave))
        loan_interest = _as_float(data.get("借款利息", loan_interest))
        late_early = _as_float(data.get("遲到/早退", late_early))
        labor_fee = _as_float(data.get("勞保費", labor_fee))
        health_fee = _as_float(data.get("健保費", health_fee))
        other_ded = _as_float(data.get("其他", other_ded))

        deduct_total = data.get("應扣總計", deduct_total)
        deduct_total = None if deduct_total is None else _as_float(deduct_total, 0.0)

        net_pay = data.get("實發金額", net_pay)
        net_pay = None if net_pay is None else _as_float(net_pay, 0.0)

        note = str(data.get("備註", note) or "")

    employee_name = (employee_name or "").strip()
    if not employee_name:
        st.error("❌ 員工姓名不可為空")
        log_action(actor or "—", "薪資管理", "儲存薪資失敗：員工姓名空白", "失敗")
        return False

    if not SALARY_DB_ID:
        st.error("❌ SALARY_DB_ID 未設定")
        log_action(actor or "—", "薪資管理", "儲存薪資失敗：SALARY_DB_ID 未設定", "失敗")
        return False

    def _f(x) -> float:
        try:
            return float(x or 0.0)
        except Exception:
            return 0.0

    existing = get_salary_record(employee_name, y, m)
    salary_props = get_db_properties(SALARY_DB_ID) or {}

    def has_prop(n: str) -> bool:
        return n in salary_props

    # -------------------------
    # 1) 自動計算總計（若未傳入）
    # -------------------------
    if gross_total is None:
        gross_total = (
            _f(full_salary)
            + _f(leader_allowance)
            + _f(job_allowance)
            + _f(perf_bonus)
            + _f(traffic_allowance)
            + _f(sales_allowance)
            + _f(coop)
            + _f(attend_bonus)
            + _f(cert_allowance)
            + _f(meal_allowance)
            + _f(ot_weekday)
            + _f(ot_sat)
            + _f(social_fee)
            + _f(year_end)
        )

    if deduct_total is None:
        deduct_total = (
            _f(advance)
            + _f(sick_leave)
            + _f(personal_leave)
            + _f(loan_interest)
            + _f(late_early)
            + _f(labor_fee)
            + _f(health_fee)
            + _f(other_ded)
        )

    if net_pay is None:
        net_pay = _f(gross_total) - _f(deduct_total)

    # -------------------------
    # 2) 組 Notion properties（只寫存在的欄位）
    # -------------------------
    props = {}

    if has_prop("員工姓名"):
        props["員工姓名"] = {"title": [{"text": {"content": employee_name}}]}
    else:
        st.error("❌ Notion 薪資表找不到 title 欄位『員工姓名』，請確認該欄位名稱是否正確。")
        log_action(actor or "—", "薪資管理", "儲存薪資失敗：缺少『員工姓名』(title)", "失敗")
        return False

    if has_prop("薪資年份"):
        props["薪資年份"] = {"number": int(y)}
    if has_prop("薪資月份"):
        props["薪資月份"] = {"number": int(m)}

    # 加項
    for k, v in [
        ("全薪", full_salary),
        ("負責人職務津貼", leader_allowance),
        ("職務津貼", job_allowance),
        ("績效獎金", perf_bonus),
        ("交通津貼", traffic_allowance),
        ("營業津貼", sales_allowance),
        ("配合", coop),
        ("全勤獎金", attend_bonus),
        ("證照加給", cert_allowance),
        ("伙食津貼", meal_allowance),
        ("平日(中晚)加班費", ot_weekday),
        ("週六加班費", ot_sat),
        ("交際費", social_fee),
        ("年終補助", year_end),
    ]:
        if has_prop(k):
            props[k] = {"number": _f(v)}

    if has_prop("薪資總計"):
        props["薪資總計"] = {"number": _f(gross_total)}

    # 扣項
    for k, v in [
        ("借支", advance),
        ("病假請假", sick_leave),
        ("事假請假", personal_leave),
        ("借款利息", loan_interest),
        ("遲到/早退", late_early),
        ("勞保費", labor_fee),
        ("健保費", health_fee),
        ("其他", other_ded),
    ]:
        if has_prop(k):
            props[k] = {"number": _f(v)}

    if has_prop("應扣總計"):
        props["應扣總計"] = {"number": _f(deduct_total)}
    if has_prop("實發金額"):
        props["實發金額"] = {"number": _f(net_pay)}

    if has_prop("備註"):
        note = (note or "").strip()
        props["備註"] = {"rich_text": [{"text": {"content": note}}]} if note else {"rich_text": []}

    # -------------------------
    # 3) 寫入 Notion（更新或新增）
    # -------------------------
    try:
        if existing and existing.get("_page_id"):
            notion.pages.update(page_id=existing["_page_id"], properties=props)
        else:
            notion.pages.create(parent={"database_id": SALARY_DB_ID}, properties=props)

        log_action(actor or "—", "薪資管理", f"儲存薪資：{employee_name} {y}/{m}", "成功")
        return True

    except Exception as e:
        st.error(f"寫入薪資資料失敗：{e}")
        log_action(actor or "—", "薪資管理", f"寫入薪資失敗：{employee_name} {y}/{m}｜{e}", "系統錯誤")
        return False


def list_salary_records(is_admin: bool, employee_name: str, y: int | None = None, m: int | None = None, limit: int = 200):
    """
    ✅ 新版薪資表欄位
    - 員工姓名/薪資年份/薪資月份
    - 加項：全薪、負責人職務津貼、職務津貼、績效獎金、交通津貼、營業津貼、配合、全勤獎金、證照加給、伙食津貼、平日(中晚)加班費、週六加班費、交際費、年終補助、薪資總計
    - 扣項：借支、病假請假、事假請假、借款利息、遲到/早退、勞保費、健保費、其他、應扣總計、實發金額
    - 備註、發薪月份（若存在）
    """
    try:
        if not SALARY_DB_ID:
            return []

        salary_props = get_db_properties(SALARY_DB_ID) or {}
        def has_prop(n: str) -> bool:
            return n in salary_props

        query = {
            "database_id": SALARY_DB_ID,
            "page_size": min(limit, 100),
        }

        # ✅ sorts：欄位存在才使用，避免 Notion 噴錯
        sort_candidates = ["建立時間", "最後更新時間", "薪資年份", "薪資月份"]
        sort_prop = next((p for p in sort_candidates if has_prop(p)), None)
        if sort_prop:
            query["sorts"] = [{"property": sort_prop, "direction": "descending"}]

        filters = []
        emp = (employee_name or "").strip()

        if not is_admin:
            # 員工只能看自己
            filters.append({"property": "員工姓名", "title": {"equals": emp}})

        if y is not None and has_prop("薪資年份"):
            filters.append({"property": "薪資年份", "number": {"equals": int(y)}})
        if m is not None and has_prop("薪資月份"):
            filters.append({"property": "薪資月份", "number": {"equals": int(m)}})

        if filters:
            query["filter"] = {"and": filters} if len(filters) > 1 else filters[0]

        res = notion.databases.query(**query)

        rows = []
        for page in res.get("results", []):
            props = page["properties"]

            def get_title(name):
                v = props.get(name, {}).get("title", [])
                return v[0]["plain_text"] if v else ""

            def get_number(name):
                try:
                    return float(props.get(name, {}).get("number") or 0.0)
                except Exception:
                    return 0.0

            def get_rich_text(name):
                v = props.get(name, {}).get("rich_text", [])
                return v[0]["plain_text"] if v else ""

            def get_date(name):
                d = props.get(name, {}).get("date")
                if d and d.get("start"):
                    return d["start"]
                return ""

            row = {"_page_id": page["id"]}

            if has_prop("員工姓名"):
                row["員工姓名"] = get_title("員工姓名")
            if has_prop("薪資年份"):
                row["薪資年份"] = int(get_number("薪資年份") or 0)
            if has_prop("薪資月份"):
                row["薪資月份"] = int(get_number("薪資月份") or 0)

            # 加項
            for f in [
                "全薪","負責人職務津貼","職務津貼","績效獎金","交通津貼","營業津貼","配合",
                "全勤獎金","證照加給","伙食津貼","平日(中晚)加班費","週六加班費","交際費","年終補助",
                "薪資總計"
            ]:
                if has_prop(f):
                    row[f] = get_number(f)

            # 扣項
            for f in ["借支","病假請假","事假請假","借款利息","遲到/早退","勞保費","健保費","其他","應扣總計","實發金額"]:
                if has_prop(f):
                    row[f] = get_number(f)

            if has_prop("備註"):
                row["備註"] = get_rich_text("備註")

            if has_prop("發薪月份"):
                row["發薪月份"] = get_date("發薪月份")

            if has_prop("建立時間"):
                row["建立時間"] = props.get("建立時間", {}).get("created_time", "")
            if has_prop("最後更新時間"):
                row["最後更新時間"] = props.get("最後更新時間", {}).get("last_edited_time", "")

            rows.append(row)

        return rows

    except Exception as e:
        st.error(f"讀取薪資清單失敗：{e}")
        return []


# =========================
# ✅ 匯出 Excel（不用額外套件）
# =========================
def make_duty_excel_bytes(y: int, m: int, df):
    """輸出成『橫向月表』Excel（格式接近你給的參考圖）"""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Alignment, Font, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{m}月值班"

    title = f"{m}月份晚間值班表"
    headers = ["日期", "星期", "檢驗線(中)", "檢驗線(晚)", "收費員(中)", "收費員(晚)", "打掃工作", "手機"]

    # 標題列（合併）
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # 表頭
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=j, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 內容列
    for i, r in enumerate(df.to_dict("records"), start=3):
        day = int(r["日期"])
        weekday = str(r["星期"])
        dt = date(int(y), int(m), day)

        row_values = [
            day,
            weekday,
            "、".join(normalize_multi_people_cell(r.get("檢驗線(中)"))),
            "、".join(normalize_multi_people_cell(r.get("檢驗線(晚)"))),
            "、".join(normalize_multi_people_cell(r.get("收費員(中)"))),
            "、".join(normalize_multi_people_cell(r.get("收費員(晚)"))),
            "、".join(normalize_multi_people_cell(r.get("打掃工作"))),
            "",  # 手機：若你要帶出，可用員工資料表 join（下一版我可以幫你補）
        ]

        for j, v in enumerate(row_values, start=1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 週日/週六顏色（先用字體顏色模擬）
        if dt.weekday() == 6:  # Sunday
            ws.cell(row=i, column=2).font = Font(color="FF0000", bold=True)
        elif dt.weekday() == 5:  # Saturday
            ws.cell(row=i, column=2).font = Font(color="00AA00", bold=True)

    # 表頭框線
    for j in range(1, len(headers)+1):
        ws.cell(row=2, column=j).border = border

    # 欄寬
    widths = [6, 6, 14, 14, 14, 14, 12, 14]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # 輸出 bytes
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def make_excel_bytes(rows: list[dict], filename_hint: str = "salary.xlsx") -> tuple[bytes, str]:
    try:
        import pandas as pd
        from io import BytesIO
        from openpyxl import Workbook  # noqa: F401

        df = pd.DataFrame(rows)
        bio = BytesIO()
        with pd.ExcelWriter(bio, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="薪資清單")
        return bio.getvalue(), filename_hint

    except Exception:
        import csv
        import io

        headers = list(rows[0].keys()) if rows else []
        sio = io.StringIO()
        w = csv.DictWriter(sio, fieldnames=headers)
        w.writeheader()
        for r in rows:
            w.writerow(r)
        data = sio.getvalue().encode("utf-8-sig")
        return data, filename_hint.replace(".xlsx", ".csv")


# ============================================================
# ✅ 出勤記錄表
# ============================================================
def create_attendance_record(employee_name: str, attend_date: date, status: str, actor: str = "") -> bool:
    if not ATTEND_DB_ID:
        st.error("❌ 尚未設定 ATTEND_DB_ID（出勤記錄表 Database ID）")
        return False

    employee_name = (employee_name or "").strip()
    status = (status or "").strip()
    if not employee_name:
        st.error("❌ 員工姓名不可為空")
        return False

    try:
        props_meta = get_db_properties(ATTEND_DB_ID)

        def has_prop(n: str) -> bool:
            return n in (props_meta or {})

        props = {}
        if has_prop("員工姓名"):
            props["員工姓名"] = {"title": [{"text": {"content": employee_name}}]}
        if has_prop("出勤日期"):
            props["出勤日期"] = {"date": {"start": datetime.combine(attend_date, datetime.min.time()).isoformat()}}
        if has_prop("出勤狀態"):
            # ✅ 預設選項：出席/請假/遲到（若你 Notion 已建好，就會用你 Notion 的）
            options = get_select_options(ATTEND_DB_ID, "出勤狀態") or [ATTEND_PRESENT_STATUS, ATTEND_LEAVE_STATUS, ATTEND_LATE_STATUS]
            if status in options:
                props["出勤狀態"] = {"select": {"name": status}}
            else:
                st.error(f"❌ 出勤狀態 Notion 選項不存在：{status}（請先在 Notion 建立選項）")
                return False

        notion.pages.create(parent={"database_id": ATTEND_DB_ID}, properties=props)
        log_action(actor or "—", "出勤新增", f"{employee_name}｜{attend_date.isoformat()}｜{status}", "成功")
        return True

    except Exception as e:
        st.error(f"寫入出勤失敗：{e}")
        log_action(actor or "—", "出勤新增", f"寫入失敗：{e}", "系統錯誤")
        return False

@st.cache_data(ttl=60)
# ============================================================
# ✅ 出勤記錄表（查詢 / 更新）
# ============================================================
def _attend_day_range(attend_date: date) -> tuple[str, str]:
    """回傳 Notion date filter 用的 [start_iso, end_iso) 區間（以該日 00:00:00 起算）。"""
    start_dt = datetime.combine(attend_date, datetime.min.time())
    end_dt = start_dt + timedelta(days=1)
    return start_dt.isoformat(), end_dt.isoformat()


def find_attendance_page(employee_name: str, attend_date: date) -> str | None:
    """用『員工姓名(Title)+出勤日期(Date)』找出當日是否已存在出勤紀錄，回傳 page_id 或 None。"""
    if not ATTEND_DB_ID:
        return None

    employee_name = (employee_name or "").strip()
    if not employee_name:
        return None

    try:
        start_iso, end_iso = _attend_day_range(attend_date)
        res = notion.databases.query(
            database_id=ATTEND_DB_ID,
            page_size=1,
            filter={
                "and": [
                    {"property": "員工姓名", "title": {"equals": employee_name}},
                    {"property": "出勤日期", "date": {"on_or_after": start_iso}},
                    {"property": "出勤日期", "date": {"before": end_iso}},
                ]
            },
        )
        results = (res or {}).get("results") or []
        if results:
            return results[0].get("id")
        return None
    except Exception:
        return None


def upsert_attendance_record(employee_name: str, attend_date: date, status: str, actor: str = "") -> bool:
    """同日同人：有就更新、沒有就新增。"""
    if not ATTEND_DB_ID:
        st.error("❌ 尚未設定 ATTEND_DB_ID（出勤記錄表 Database ID）")
        return False

    employee_name = (employee_name or "").strip()
    status = (status or "").strip()
    if not employee_name:
        st.error("❌ 員工姓名不可為空")
        return False

    try:
        props_meta = get_db_properties(ATTEND_DB_ID)

        def has_prop(n: str) -> bool:
            return n in (props_meta or {})

        # 驗證狀態選項
        options = get_select_options(ATTEND_DB_ID, "出勤狀態") or [ATTEND_PRESENT_STATUS, ATTEND_LEAVE_STATUS, ATTEND_LATE_STATUS]
        if status not in options:
            st.error(f"❌ 出勤狀態 Notion 選項不存在：{status}（請先在 Notion 建立選項）")
            return False

        page_id = find_attendance_page(employee_name, attend_date)

        props = {}
        if has_prop("員工姓名") and (not page_id):
            props["員工姓名"] = {"title": [{"text": {"content": employee_name}}]}
        if has_prop("出勤日期"):
            props["出勤日期"] = {"date": {"start": datetime.combine(attend_date, datetime.min.time()).isoformat()}}
        if has_prop("出勤狀態"):
            props["出勤狀態"] = {"select": {"name": status}}

        if page_id:
            notion.pages.update(page_id=page_id, properties=props)
            log_action(actor or "—", "出勤更新", f"{employee_name}｜{attend_date.isoformat()}｜{status}", "成功")
        else:
            notion.pages.create(parent={"database_id": ATTEND_DB_ID}, properties=props)
            log_action(actor or "—", "出勤新增", f"{employee_name}｜{attend_date.isoformat()}｜{status}", "成功")

        return True

    except Exception as e:
        st.error(f"寫入出勤失敗：{e}")
        log_action(actor or "—", "出勤寫入", f"寫入失敗：{e}", "系統錯誤")
        return False


def get_attendance_status_map_by_date(attend_date: date) -> dict[str, str]:
    """抓出勤記錄表中『某一天』所有員工的狀態，回傳 {員工姓名: 出勤狀態}。"""
    if not ATTEND_DB_ID:
        return {}

    try:
        start_iso, end_iso = _attend_day_range(attend_date)
        out: dict[str, str] = {}
        cursor = None

        def _get_emp_name(props: dict) -> str:
            p = props.get("員工姓名", {}) or {}
            ptype = (p or {}).get("type")
            if ptype == "title":
                return (_title_get_first_plain_text(p) or "").strip()
            if ptype == "rich_text":
                return (_get_prop_plain_text(p) or "").strip()
            if ptype == "select":
                return ((p.get("select") or {}).get("name") or "").strip()
            # fallback：多做一次容錯
            return ((_title_get_first_plain_text(p) or _get_prop_plain_text(p) or "").strip())

        def _get_status(props: dict) -> str:
            p = props.get("出勤狀態", {}) or {}
            ptype = (p or {}).get("type")
            if ptype == "status":
                return (((p.get("status") or {}).get("name")) or "").strip()
            return (((p.get("select") or {}).get("name")) or "").strip()

        while True:
            res = notion.databases.query(
                database_id=ATTEND_DB_ID,
                page_size=100,
                start_cursor=cursor,
                filter={
                    "and": [
                        {"property": "出勤日期", "date": {"on_or_after": start_iso}},
                        {"property": "出勤日期", "date": {"before": end_iso}},
                    ]
                },
            )
            results = (res or {}).get("results") or []
            for p in results:
                props = p.get("properties", {}) or {}
                emp = _get_emp_name(props)
                stt = _get_status(props)
                if emp:
                    out[emp] = stt

            if not res.get("has_more"):
                break
            cursor = res.get("next_cursor")

        return out
    except Exception:
        return {}



def list_attendance_records(start_d: date, end_d: date, employee_name: str | None = None, limit: int = 500) -> list[dict]:
    """
    查詢【出勤記錄表】在區間 [start_d, end_d) 的清單
    - employee_name=None 或 "全部員工"：查全部
    - employee_name=某員工：只查該員工
    回傳欄位：員工姓名 / 出勤日期 / 出勤狀態 / 建立時間 / 最後更新時間 / _page_id
    """
    if not ATTEND_DB_ID:
        return []

    emp = (employee_name or "").strip()

    try:
        filters = [
            {"property": "出勤日期", "date": {"on_or_after": datetime.combine(start_d, datetime.min.time()).isoformat()}},
            {"property": "出勤日期", "date": {"before": datetime.combine(end_d, datetime.min.time()).isoformat()}},
        ]
        if emp and emp != "全部員工":
            filters.insert(0, {"property": "員工姓名", "title": {"equals": emp}})

        query = {
            "database_id": ATTEND_DB_ID,
            "page_size": 100,
            "sorts": [{"property": "出勤日期", "direction": "descending"}],
            "filter": {"and": filters} if len(filters) > 1 else filters[0],
        }

        rows: list[dict] = []
        next_cursor = None

        while True:
            if next_cursor:
                query["start_cursor"] = next_cursor

            res = notion.databases.query(**query)

            for page in res.get("results", []):
                props = page["properties"]

                def get_title(name):
                    v = props.get(name, {}).get("title", [])
                    return v[0]["plain_text"] if v else ""

                def get_select(name):
                    p = props.get(name, {}) or {}
                    t = p.get("type")
                    if t == "select":
                        v = p.get("select")
                        return v.get("name") if v else ""
                    if t == "status":
                        v = p.get("status")
                        return v.get("name") if v else ""
                    if t == "multi_select":
                        vs = p.get("multi_select") or []
                        return "、".join([x.get("name", "") for x in vs if x.get("name")])
                    # fallback（舊資料/未知型態）
                    v = p.get("select")
                    return v.get("name") if v else ""

                def get_date_only(name):
                    d = props.get(name, {}).get("date")
                    if not d or not d.get("start"):
                        return ""
                    try:
                        return datetime.fromisoformat(d["start"].replace("Z", "+00:00")).date().isoformat()
                    except Exception:
                        return d.get("start", "")

                rows.append({
                    "_page_id": page["id"],
                    "員工姓名": get_title("員工姓名"),
                    "出勤日期": get_date_only("出勤日期"),
                    "出勤狀態": get_select("出勤狀態"),
                    "建立時間": props.get("建立時間", {}).get("created_time", ""),
                    "最後更新時間": props.get("最後更新時間", {}).get("last_edited_time", ""),
                })

                if len(rows) >= int(limit):
                    return rows

            if not res.get("has_more"):
                break
            next_cursor = res.get("next_cursor")

        return rows

    except Exception as e:
        st.error(f"讀取出勤紀錄失敗：{e}")
        return []


@st.cache_data(ttl=60)
def _list_lunch_eligible_attendance_days(employee_name: str, start_d: date, end_d: date) -> list[date]:
    """
    回傳：該員工在區間內（start_d ~ end_d）出勤狀態 ∈ {出席, 遲到} 的『紀錄日期列表』（保留重複）。
    - 同一天若有多筆出勤紀錄（例如有重複/異常資料），會視為多次計算（每筆都算 90）。

    ✅ 容錯：
      - 員工姓名欄位可能是 title / rich_text / select
      - 出勤狀態欄位可能是 select / status（Notion 的 Status property）
    """
    if not ATTEND_DB_ID:
        return []

    employee_name = (employee_name or "").strip()
    if not employee_name:
        return []

    try:
        meta = get_db_properties(ATTEND_DB_ID) or {}
        k_emp = resolve_prop_key(meta, "員工姓名") or "員工姓名"
        k_date = resolve_prop_key(meta, "出勤日期") or "出勤日期"
        k_status = resolve_prop_key(meta, "出勤狀態") or "出勤狀態"

        # --- 員工姓名 filter（title / rich_text / select）
        emp_type = (meta.get(k_emp) or {}).get("type")
        if emp_type == "select":
            emp_filter = {"property": k_emp, "select": {"equals": employee_name}}
        elif emp_type == "rich_text":
            emp_filter = {"property": k_emp, "rich_text": {"equals": employee_name}}
        else:
            emp_filter = {"property": k_emp, "title": {"equals": employee_name}}

        # --- 出勤狀態 filter（select / status / multi_select）
        status_type = (meta.get(k_status) or {}).get("type")

        if status_type == "status":
            status_or = [{"property": k_status, "status": {"equals": s}} for s in sorted(ATTEND_LUNCH_ELIGIBLE_STATUSES)]
        elif status_type == "multi_select":
            # multi_select：用 contains
            status_or = [{"property": k_status, "multi_select": {"contains": s}} for s in sorted(ATTEND_LUNCH_ELIGIBLE_STATUSES)]
        else:
            # 預設用 select
            status_or = [{"property": k_status, "select": {"equals": s}} for s in sorted(ATTEND_LUNCH_ELIGIBLE_STATUSES)]

        base_and = [
            emp_filter,
            {"property": k_date, "date": {"on_or_after": datetime.combine(start_d, datetime.min.time()).isoformat()}},
            {"property": k_date, "date": {"before": datetime.combine(end_d, datetime.min.time()).isoformat()}},
        ]
        notion_filter = {"and": base_and + [{"or": status_or}]}

        # 單一員工每月理論上 <= 31 筆，但仍保留分頁以防異常資料
        days: list[date] = []
        next_cursor = None
        while True:
            query = {
                "database_id": ATTEND_DB_ID,
                "filter": notion_filter,
                "page_size": 100,
            }
            if next_cursor:
                query["start_cursor"] = next_cursor

            res = notion.databases.query(**query)

            for page in res.get("results", []):
                props = page.get("properties", {}) or {}
                d = (props.get(k_date, {}) or {}).get("date")
                if not d or not d.get("start"):
                    continue
                try:
                    dd = datetime.fromisoformat(d["start"].replace("Z", "+00:00")).date()
                    days.append(dd)
                except Exception:
                    continue

            if not res.get("has_more"):
                break
            next_cursor = res.get("next_cursor")

        return days

    except Exception:
        return []


def _month_range(y: int, m: int) -> tuple[date, date]:
    start = date(int(y), int(m), 1)
    if m == 12:
        end = date(int(y) + 1, 1, 1)
    else:
        end = date(int(y), int(m) + 1, 1)
    return start, end


def _daterange(d1: date, d2: date):
    # [d1, d2)
    cur = d1
    while cur < d2:
        yield cur
        cur += timedelta(days=1)


def calc_working_days_for_lunch(employee_name: str, y: int, m: int) -> tuple[int, list[date]]:
    """
    ✅ 午餐可領工餐日判定（依你的新規則）：
      - 只看【出勤記錄表】
      - 出勤狀態：出席 / 遲到 → 都算 90
      - 請假 → 不算
      - 週一~週六 才算；週日不算
    回傳 (天數, 日期列表)
    """
    start_d, end_d = _month_range(int(y), int(m))
    eligible_days = _list_lunch_eligible_attendance_days(employee_name, start_d, end_d)

    working_list = []
    for d in sorted(eligible_days):
        if d.weekday() not in WORKDAY_WEEKDAYS:
            continue
        if start_d <= d < end_d:
            working_list.append(d)

    return len(working_list), working_list


def create_lunch_record(employee_name: str, lunch_date: date, amount: float, actor: str = "") -> bool:
    """
    午餐訂餐表欄位（依你截圖）：
      - 員工姓名 (title)
      - 訂餐金額 (number)
      - 訂餐日期 (date)
    """
    if not LUNCH_DB_ID:
        st.error("❌ 尚未設定 LUNCH_DB_ID（午餐訂餐表 Database ID）")
        return False

    employee_name = (employee_name or "").strip()
    if not employee_name:
        st.error("❌ 員工姓名不可為空")
        return False

    try:
        props_meta = get_db_properties(LUNCH_DB_ID)

        def has_prop(n: str) -> bool:
            return n in (props_meta or {})

        props = {}
        if has_prop("員工姓名"):
            props["員工姓名"] = {"title": [{"text": {"content": employee_name}}]}
        if has_prop("訂餐金額"):
            props["訂餐金額"] = {"number": float(amount or 0)}
        if has_prop("訂餐日期"):
            props["訂餐日期"] = {"date": {"start": datetime.combine(lunch_date, datetime.min.time()).isoformat()}}

        notion.pages.create(parent={"database_id": LUNCH_DB_ID}, properties=props)
        log_action(actor or employee_name, "午餐訂餐", f"{employee_name}｜{lunch_date.isoformat()}｜${float(amount or 0):.0f}", "成功")
        return True

    except Exception as e:
        st.error(f"寫入午餐訂餐失敗：{e}")
        log_action(actor or employee_name, "午餐訂餐", f"寫入失敗：{e}", "系統錯誤")
        return False


def list_lunch_records(is_admin: bool, employee_name: str, start_d: date, end_d: date, limit: int = 200) -> list[dict]:
    if not LUNCH_DB_ID:
        return []

    emp = (employee_name or "").strip()
    try:
        meta = get_db_properties(LUNCH_DB_ID) or {}
        k_emp = resolve_prop_key(meta, "員工姓名") or "員工姓名"
        k_date = resolve_prop_key(meta, "訂餐日期") or "訂餐日期"
        k_amt = resolve_prop_key(meta, "訂餐金額") or "訂餐金額"

        emp_type = (meta.get(k_emp) or {}).get("type")
        emp_filter = None
        if emp and emp != "全部員工":
            if emp_type == "select":
                emp_filter = {"property": k_emp, "select": {"equals": emp}}
            elif emp_type == "rich_text":
                emp_filter = {"property": k_emp, "rich_text": {"equals": emp}}
            else:
                emp_filter = {"property": k_emp, "title": {"equals": emp}}

        filters = []
        if emp_filter:
            filters.append(emp_filter)

        filters.append({"property": k_date, "date": {"on_or_after": datetime.combine(start_d, datetime.min.time()).isoformat()}})
        filters.append({"property": k_date, "date": {"before": datetime.combine(end_d, datetime.min.time()).isoformat()}})

        query = {
            "database_id": LUNCH_DB_ID,
            "page_size": min(limit, 100),
            "sorts": [{"property": k_date, "direction": "descending"}],
            "filter": {"and": filters} if len(filters) > 1 else filters[0],
        }

        res = notion.databases.query(**query)

        rows = []
        for page in res.get("results", []):
            props = page.get("properties", {}) or {}

            def get_emp():
                p = props.get(k_emp, {}) or {}
                t = p.get("type")
                if t == "select":
                    return ((p.get("select") or {}).get("name")) or ""
                if t == "rich_text":
                    rt = p.get("rich_text") or []
                    return "".join([x.get("plain_text", "") for x in rt]).strip()
                tt = p.get("title") or []
                return tt[0].get("plain_text", "") if tt else ""

            def get_amt():
                return float(((props.get(k_amt) or {}).get("number")) or 0.0)

            def get_date():
                d = (props.get(k_date, {}) or {}).get("date")
                if not d or not d.get("start"):
                    return ""
                try:
                    return datetime.fromisoformat(d["start"].replace("Z", "+00:00")).date().isoformat()
                except Exception:
                    return d["start"][:10]

            rows.append({
                "員工姓名": get_emp(),
                "訂餐日期": get_date(),
                "訂餐金額": get_amt(),
            })

        return rows

    except Exception:
        return []

def calc_month_lunch_settlement(employee_name: str, y: int, m: int, is_admin: bool) -> dict:
    start_d, end_d = _month_range(int(y), int(m))
    eligible_days, _eligible_list = calc_working_days_for_lunch(employee_name, int(y), int(m))
    entitlement = eligible_days * LUNCH_ALLOWANCE_PER_DAY

    rows = list_lunch_records(is_admin=is_admin, employee_name=employee_name, start_d=start_d, end_d=end_d, limit=500)
    spent = sum(float(r.get("訂餐金額", 0) or 0) for r in rows)
    diff = float(entitlement) - float(spent)

    return {
        "員工姓名": employee_name,
        "年份": int(y),
        "月份": int(m),
        "可領工餐天數(出席/遲到)": int(eligible_days),
        "應得午餐補助(可領×90)": float(entitlement),
        "已訂餐金額": float(spent),
        "差額(應得-已訂餐)": float(diff),
    }




def admin_reset_user_password(target_username: str, temp_password: str, actor: str = "") -> bool:
    """
    管理員重設：寫入「密碼(明碼)」、清空「login_hash」、勾 must_change_password
    讓員工用臨時密碼先登入，登入後會被強制改密碼（改完就會寫回 login_hash）
    """
    target_username = (target_username or "").strip()
    temp_password = (temp_password or "").strip()

    if not target_username or not temp_password:
        st.error("❌ 員工或臨時密碼不可為空")
        log_action(actor or "—", "重設密碼", "員工或臨時密碼為空", "失敗")
        return False

    # 你也可以自己調整臨時密碼規則
    if len(temp_password) < 6:
        st.error("❌ 臨時密碼至少 6 碼")
        log_action(actor or "—", "重設密碼", f"{target_username} 臨時密碼長度不足", "失敗")
        return False

    page = get_account_page_by_username(target_username)
    if not page:
        st.error("❌ 找不到該員工帳號資料")
        log_action(actor or "—", "重設密碼", f"找不到帳號：{target_username}", "失敗")
        return False

    page_id = page["id"]
    props_meta = get_db_properties(ACCOUNT_DB_ID) or {}

    def has_prop(n: str) -> bool:
        return n in props_meta

    props_to_update = {}

    # ✅ 寫入臨時明碼到「密碼」
    if has_prop("密碼"):
        props_to_update["密碼"] = {"rich_text": [{"text": {"content": temp_password}}]}
    else:
        st.error("❌ 帳號管理表缺少『密碼』欄位（rich_text）")
        return False

    # ✅ 清空 login_hash（避免明碼與 hash 同時存在）
    if has_prop("login_hash"):
        props_to_update["login_hash"] = {"rich_text": []}

    # ✅ 強制下次登入改密碼
    if has_prop("must_change_password"):
        props_to_update["must_change_password"] = {"checkbox": True}

    # （可選）記錄重設時間
    if has_prop("last_password_reset"):
        props_to_update["last_password_reset"] = {"date": {"start": datetime.now().isoformat()}}

    try:
        notion.pages.update(page_id=page_id, properties=props_to_update)
        log_action(actor or "—", "重設密碼", f"已重設：{target_username}", "成功")
        return True
    except Exception as e:
        st.error(f"❌ 重設密碼寫回 Notion 失敗：{e}")
        log_action(actor or "—", "重設密碼", f"寫回失敗：{target_username}｜{e}", "系統錯誤")
        return False


# =========================
# 10) Streamlit 設定
# =========================
st.set_page_config(page_title="公司內部系統", layout="wide")



# =========================
# UI Theme (商業風 + 淡淺藍 + 毛玻璃)
# =========================
st.markdown(
    """
    <style>
    /* ----- Base background ----- */
    .stApp {
        background: radial-gradient(1200px 800px at 15% 10%, rgba(208, 235, 255, 0.75) 0%, rgba(230, 246, 255, 0.55) 35%, rgba(245, 250, 255, 0.35) 70%, rgba(255, 255, 255, 1) 100%) !important;
    }

    /* Page padding */
    section.main > div { padding-top: 1.2rem; }

    /* ----- Card / glass containers ----- */
    .glass-card, .stMetric, div[data-testid="stMetric"] {
        background: rgba(255, 255, 255, 0.55) !important;
        border: 1px solid rgba(0, 90, 150, 0.12) !important;
        border-radius: 16px !important;
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
        box-shadow: 0 10px 30px rgba(0, 35, 60, 0.08);
    }

    /* Metric internal spacing */
    div[data-testid="stMetric"] { padding: 14px 16px !important; }

    /* ----- Sidebar look ----- */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, rgba(227, 244, 255, 0.85) 0%, rgba(245, 250, 255, 0.85) 100%) !important;
        border-right: 1px solid rgba(139, 195, 255, 0.12);
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
    }

    /* Sidebar title text tone */
    [data-testid="stSidebar"] * { color: rgba(20, 35, 55, 0.92); }

    /* ----- Buttons (統一樣式：含功能選單/登出/更改密碼/重新同步等) ----- */
    [data-testid="stSidebar"] .stButton > button {
        width: 100%;
        height: 44px;
        border-radius: 12px;
        border: 1px solid rgba(0, 90, 150, 0.18);
        background: rgba(255, 255, 255, 0.55);
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
        box-shadow: 0 8px 18px rgba(0, 35, 60, 0.08);
        font-weight: 600;
        letter-spacing: 0.2px;
        justify-content: flex-start;   /* ✅ 文字靠左 */
        text-align: left;
        padding-left: 14px;
        transition: transform 0.06s ease-in-out, filter 0.12s ease-in-out;
    }
    [data-testid="stSidebar"] .stButton > button:active {
        transform: translateY(1px);
        filter: brightness(0.98);
    }
    [data-testid="stSidebar"] .stButton > button:focus { outline: none; box-shadow: 0 0 0 3px rgba(80, 170, 255, 0.25); }

    /* Remove default hover color shift (盡量維持同色系，避免突兀變色) */
    [data-testid="stSidebar"] .stButton > button:hover {
        /* ✅ 取消 hover 變色：保持與原本一致 */
        background: rgba(255, 255, 255, 0.55);
        border-color: rgba(0, 90, 150, 0.18);
    }

    /* ----- Inputs / selects subtle glass ----- */
    .stTextInput > div > div > input,
    .stTextArea textarea,
    .stSelectbox div[data-baseweb="select"] > div,
    .stNumberInput input {
        background: rgba(255, 255, 255, 0.60) !important;
        border: 1px solid rgba(0, 90, 150, 0.12) !important;
        border-radius: 12px !important;
        backdrop-filter: blur(8px);
        -webkit-backdrop-filter: blur(8px);
    }

    /* ----- Announcement box ----- */
    .announce-box {
        border: 1px solid rgba(0, 90, 150, 0.18);
        border-radius: 16px;
        padding: 10px 12px;
        height: var(--announce-h, 456px);
        overflow-y: auto;
        background: rgba(255,255,255,0.55);
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
        box-shadow: 0 10px 26px rgba(0, 35, 60, 0.08);
    }
    .announce-title {
        position: sticky;
        top: 0;
        z-index: 2;
        text-align: center;
        font-weight: 800;
        padding: 6px 0 10px 0;
        background: linear-gradient(180deg, rgba(255,255,255,0.78) 0%, rgba(255,255,255,0.45) 100%);
        border-bottom: 1px dashed rgba(0, 90, 150, 0.18);
        border-radius: 12px;
        margin-bottom: 8px;
        letter-spacing: 0.5px;
    }
    .announce-empty {
        display:flex;
        align-items:center;
        justify-content:center;
        height: calc(100% - 54px);
        color: rgba(20, 35, 55, 0.55);
        font-weight: 700;
        font-size: 16px;
    }
    

    /* ----- Footer bar ----- */
    .block-container{ padding-bottom: 78px !important; } /* avoid content hidden behind fixed footer */
    .app-footer{
        position: fixed;
        left: 0;
        right: 0;
        bottom: 0;
        height: 28px;
        background: #0b2b5b;
        color: #ffffff;
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 12px;
        line-height: 1;
        padding: 0 10px;
        font-weight: 400;
        letter-spacing: 0.3px;
        z-index: 500;
        box-shadow: 0 -10px 26px rgba(0, 35, 60, 0.18);
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;        
    }

    /* ----- Mobile (RWD) tweaks ----- */
    @media (max-width: 768px){
        section.main > div{ padding-top: 0.8rem !important; padding-left: 0.8rem !important; padding-right: 0.8rem !important; }
        .welcome-card{ padding: 18px 18px !important; }
        .welcome-title{ font-size: 22px !important; }
        .metric-box{ min-width: 100% !important; }
        .announce-box{ height: auto !important; }
        .announce-body{ height: auto !important; max-height: 420px; }
        div[data-testid="stForm"]{ padding: 26px 18px 20px 18px !important; border-radius: 22px !important; }
        div[data-testid="stForm"] .login-title{ font-size: 22px !important; }
        .app-footer{
            height: 26px;
            font-size: 10.5px;
            letter-spacing: 0.2px;
            padding: 0 8px;
        .app-footer p, .app-footer div, .app-footer span{
            margin: 0 !important;
            padding: 0 !important;
            line-height: 1 !important;    
    }
</style>
    """,
    unsafe_allow_html=True,
)

def render_footer():
    # ✅ 全頁面固定底部深藍色版權條（含登入頁）
    st.markdown(
        '<div class="app-footer">元廣順汽車有限公司內部系統｜Copyright © 2026 By LJOU</div>',
        unsafe_allow_html=True,
    )



if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False

if "hours_per_day" not in st.session_state:
    st.session_state["hours_per_day"] = DEFAULT_HOURS_PER_DAY
if "cashout_cap_days" not in st.session_state:
    st.session_state["cashout_cap_days"] = DEFAULT_CASHOUT_CAP_DAYS
if "cashout_amount_per_day" not in st.session_state:
    st.session_state["cashout_amount_per_day"] = DEFAULT_CASHOUT_AMOUNT_PER_DAY
if "cashout_whole_days_only" not in st.session_state:
    st.session_state["cashout_whole_days_only"] = DEFAULT_CASHOUT_WHOLE_DAYS_ONLY


if st.session_state.get("logged_in", False) and st.session_state.get("force_change_pwd", False):
    st.warning("⚠️ 你目前是用明碼登入或被重設密碼，請先更改密碼。")
    change_password_dialog(force=True)
    render_footer()
    st.stop()


# =========================
# 11) 登入介面
# =========================
if not st.session_state["logged_in"]:

    # ✅ 只在登入頁套用的 CSS（其他頁面完全不動）
    st.markdown(
        r'''
        <style>
        /* ===== Login page only (Blue) ===== */
        .stApp {
            background: radial-gradient(1200px 800px at 20% 15%,
                rgba(208, 235, 255, 0.75) 0%,
                rgba(230, 246, 255, 0.55) 35%,
                rgba(245, 250, 255, 0.35) 70%,
                rgba(255, 255, 255, 1) 100%) !important;
        }

        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}
        header {visibility: hidden;}

        section.main > div { padding-top: 0rem !important; }

        /* ✅ 讓登入表單「真的被包在白色卡片內」：直接把 st.form 的容器當作卡片 */
        div[data-testid="stForm"] {
            width: min(520px, 94vw);
            margin: 0 auto;
            background: rgba(255, 255, 255, 0.88) !important;
            border-radius: 26px !important;
            border: 1px solid rgba(0, 90, 150, 0.14) !important;
            box-shadow: 0 18px 50px rgba(0, 35, 60, 0.12) !important;
            padding: 34px 34px 28px 34px !important;
        }

        /* icon + title */
        .login-icon {
            width: 56px;
            height: 56px;
            border-radius: 16px;
            display: grid;
            place-items: center;
            margin: 0 auto 10px auto;
            background: linear-gradient(180deg, rgba(60, 140, 255, 0.95) 0%, rgba(154, 238, 255, 0.8) 100%);
            box-shadow: 0 10px 22px rgba(0, 90, 150, 0.22);
        }
        .login-icon span {
            font-size: 28px;
            line-height: 1;
            filter: drop-shadow(0 6px 12px rgba(0,0,0,0.10));
        }
        .login-title {
            text-align: center;
            font-size: 26px;
            font-weight: 800;
            margin: 6px 0 18px 0;
            color: rgba(55, 62, 120, 0.92);  #文字顏色
            letter-spacing: 0.3px;
        }

        /* inputs */
        div[data-testid="stForm"] label {
            font-weight: 700 !important;
            color: rgba(20, 35, 55, 0.86) !important;
        }
        div[data-testid="stForm"] .stTextInput > div > div > input {
            height: 48px !important;
            border-radius: 12px !important;
            background: rgba(235, 246, 255, 0.55) !important;
            border: 1px solid rgba(0, 90, 150, 0.18) !important;
            box-shadow: inset 0 1px 0 rgba(255,255,255,0.55);
        }
        div[data-testid="stForm"] .stTextInput > div > div > input:focus {
            border-color: rgba(0, 90, 150, 0.48) !important;
            box-shadow: 0 0 0 4px rgba(0, 90, 150, 0.16) !important;
        }

        /* submit button */
        div[data-testid="stForm"] .stButton > button,
        div[data-testid="stForm"] .stFormSubmitButton > button {
            width: 100% !important;
            height: 54px !important;
            border-radius: 14px !important;
            border: none !important;
            color: #fff !important;
            font-weight: 800 !important;
            font-size: 18px !important;
            letter-spacing: 1px !important;
            background: linear-gradient(180deg, rgba(60, 140, 255, 1) 0%, rgba(0, 90, 150, 1) 100%) !important;
            box-shadow: 0 14px 30px rgba(0, 90, 150, 0.20) !important;
        }
        div[data-testid="stForm"] .stButton > button:hover,
        div[data-testid="stForm"] .stFormSubmitButton > button:hover { filter: brightness(1.03); }
        div[data-testid="stForm"] .stButton > button:active,
        div[data-testid="stForm"] .stFormSubmitButton > button:active { transform: translateY(1px); filter: brightness(0.98); }

        div[data-testid="stForm"] .stAlert { border-radius: 12px !important; }
        </style>
        ''',
        unsafe_allow_html=True,
    )

    # ✅ 版面置中（不影響其他頁）
    pad1, center, pad2 = st.columns([1, 1.2, 1])
    with center:
# =========================
# 🛠 部署 Debug（尚未登入也可用）
# =========================
try:
    qp_debug = False
    try:
        qp = st.query_params  # Streamlit 1.54+
        qp_debug = str(qp.get("debug", "0")).strip() in ("1", "true", "True", "yes", "on")
    except Exception:
        qp_debug = False

    env_debug = str(_get_cfg("DEPLOY_DEBUG", "0")).strip() in ("1", "true", "True", "yes", "on")
    if "deploy_debug" not in st.session_state:
        st.session_state["deploy_debug"] = bool(qp_debug or env_debug)

    with st.expander("🛠 部署 Debug（尚未登入也可用）", expanded=bool(st.session_state.get("deploy_debug"))):
        st.session_state["deploy_debug"] = st.checkbox("開啟 Debug", value=bool(st.session_state.get("deploy_debug")), key="deploy_debug_chk")
        st.caption('開啟方式：網址加 ?debug=1 或 Secrets/Env：DEPLOY_DEBUG=1')
        if st.session_state.get("login_debug"):
            st.subheader("🔎 Debug / login()")
            st.json(st.session_state["login_debug"])
except Exception:
    pass

        # ✅ 把登入區改成 st.form：外框就是表單容器，所以「一定會被包在卡片裡」
        with st.form("login_form", clear_on_submit=False):
            st.markdown('<div class="login-icon"><span>🔐</span></div>', unsafe_allow_html=True)
            st.markdown('<div class="login-title">元廣順汽車有限公司<br>員工內部系統</div>', unsafe_allow_html=True)

            # ✅ 保留你原本的登入邏輯（只改外觀）
            user = st.text_input("員工姓名", placeholder="請輸入員工姓名...", key="login_user")
            pwd = st.text_input("密碼", type="password", key="login_pwd")

            submitted = st.form_submit_button("登入系統", use_container_width=True)
            if submitted:
                ok, is_admin, must_change = login(user, pwd)
                if ok:
                    st.session_state["logged_in"] = True
                    st.session_state["user"] = user.strip()
                    st.session_state["is_admin"] = is_admin

                    page = get_account_page_by_username(user.strip())
                    st.session_state["account_page_id"] = page["id"] if page else ""
                    st.session_state["force_change_pwd"] = bool(must_change)

                    st.success("登入成功！")
                    time.sleep(0.4)
                    st.rerun()
                else:
                    st.error("帳號或密碼錯誤")


# =========================
# 12) 主系統介面
# =========================
else:
    is_admin = st.session_state.get("is_admin", False)
    current_user = st.session_state.get("user", "")

    st.sidebar.title(f"🧑‍💼 {current_user}")
    st.sidebar.caption("🔑 管理員" if is_admin else "🧑‍💼 員工")

    

    # ✅ 選單
    if is_admin:
        menu_items = [
            "🏠 個人首頁",
            "📍 每日打卡",
            "🗓️ 年度特休",
            "📝 請假申請",
            "📅 出勤記錄",
            "🍱 午餐管理",
            "💵 薪資計算",
            "🗓️ 值班排班",
            "📋 操作記錄",
            "📢 公告管理",
            "⚙️ 系統設定",

        ]
    else:
        menu_items = [
            "🏠 個人首頁",
            "📍 每日打卡",
            "📝 請假申請",
            "🍱 午餐紀錄",
            "💰 薪資查詢",
        ]

    
    # ✅ 功能選單（按鈕版：與「登出 / 更改密碼」同樣式；文字靠左；尺寸一致）
    if "menu" not in st.session_state:
        st.session_state["menu"] = menu_items[0] if menu_items else ""

    st.sidebar.markdown("### 功能選單")

    for i, _item in enumerate((menu_items or [])):
        _selected = (st.session_state.get("menu") == _item)
        _label = f"▸ {_item}" if _selected else f"  {_item}"

        # ✅ 產生穩定且唯一的 key（避免空字串/符號導致重複）
        _k = re.sub(r"[^0-9a-zA-Z_]+", "_", str(_item)).strip("_")
        if not _k:
            _k = f"item_{i}"

        _role_key = "admin" if st.session_state.get("is_admin") else "staff"
        _btn_key = f"menu_btn_{_role_key}_{i}_{_k}"

        if st.sidebar.button(_label, use_container_width=True, key=_btn_key):
            st.session_state["menu"] = _item
            st.rerun()

    menu = st.session_state.get("menu", menu_items[0] if menu_items else "")


    st.sidebar.divider()

    # ✅ 側邊欄按鈕（直式排列）
    if st.sidebar.button("登出", use_container_width=True):
        log_action(current_user, "登出", "使用者登出", "成功")
        st.session_state["logged_in"] = False
        st.rerun()

    if st.sidebar.button("更改密碼", use_container_width=True):
        change_password_dialog(force=False)

    # ✅ 重新同步：僅管理員可用（清除快取 + 重新載入）
    if is_admin:
        if st.sidebar.button("重新同步", use_container_width=True, help="清除 Streamlit 快取並重新載入 Notion 資料"):
            try:
                st.cache_data.clear()
                st.cache_resource.clear()
            except Exception:
                pass

            try:
                _preserve = {"logged_in", "user", "is_admin", "force_change_pwd", "gps_lat", "gps_lon", "gps_err"}
                _salary_related_keys = {
                    "calc_y", "calc_m", "calc_emp", "list_y", "list_m",
                    "全薪", "負責人職務津貼", "職務津貼", "績效獎金", "交通津貼", "營業津貼", "配合", "全勤獎金",
                    "證照加給", "伙食津貼", "平日(中晚)加班費", "週六加班費", "交際費", "年終補助",
                    "借支", "病假請假", "事假請假", "借款利息", "遲到/早退", "勞保費", "健保費", "其他", "備註",
                }
                for k in list(st.session_state.keys()):
                    if k in _preserve:
                        continue
                    if (k in _salary_related_keys) or str(k).startswith("salary_"):
                        del st.session_state[k]
            except Exception:
                pass

            st.toast("✅ 已清除快取，重新同步中…")
            st.rerun()



    # -------------------------
    # 個人首頁
    # -------------------------
    if menu == "🏠 個人首頁":
        st.header("儀表板")

        # ===== 台灣日期（UTC+8）=====
        tw_now = datetime.now(timezone(timedelta(hours=8)))
        this_year = tw_now.year
        this_month = tw_now.month
        tw_date_str = tw_now.strftime("%Y-%m-%d")

        # ===== 指標：特休 / 午餐差額 =====
        snap = get_employee_vacation_snapshot(current_user, this_year)
        remaining_text = "—"
        if snap and snap.get("remaining") is not None:
            try:
                remaining_text = f"{int(snap['remaining'])} 小時"
            except Exception:
                remaining_text = "—"

        lunch_text = "—"
        if LUNCH_DB_ID and ATTEND_DB_ID:
            try:
                s = calc_month_lunch_settlement(current_user, this_year, this_month, is_admin=False)
                lunch_text = f"${s['差額(應得-已訂餐)']:.0f}"
            except Exception:
                lunch_text = "—"
        elif not ATTEND_DB_ID:
            lunch_text = "（尚未設定 ATTEND_DB_ID）"
        else:
            lunch_text = "（尚未設定 LUNCH_DB_ID）"

        # ===== 歡迎卡片 + 公告：左右並排（左=歡迎卡片/指標；右=公告） =====
        left, right = st.columns([2, 1], gap="large")

        with left:
            # CSS（避免縮排造成 Markdown 當成 code block，所以用 dedent）
            st.markdown(
                textwrap.dedent("""
                <style>
                .welcome-card{
                    background: white;
                    padding: 26px 28px;
                    border-radius: 18px;
                    box-shadow: 0 8px 24px rgba(0,0,0,0.08);
                    margin: 6px 0 18px 0;
                }
                .welcome-title{
                    font-size: 26px;
                    font-weight: 800;
                    color: #1c5fa8;
                    margin: 0 0 10px 0;
                    letter-spacing: 0.2px;
                }
                .welcome-sub{
                    font-size: 15px;
                    color: rgba(20, 35, 55, 0.62);
                    margin: 0 0 16px 0;
                    font-variant-numeric: tabular-nums;
                }
                .metrics-row{
                    display:flex;
                    gap: 16px;
                    flex-wrap: wrap;
                }
                .metric-box{
                    flex: 1;
                    min-width: 220px;
                    background: rgba(245,247,251,0.95);
                    border: 1px solid rgba(0, 90, 150, 0.12);
                    padding: 16px 16px;
                    border-radius: 14px;
                    text-align: center;
                }
                .metric-title{
                    font-size: 13px;
                    color: rgba(20, 35, 55, 0.55);
                    margin-bottom: 6px;
                    font-weight: 700;
                }
                .metric-value{
                    font-size: 22px;
                    font-weight: 800;
                    color: #1c5fa8;
                }
                </style>
                """),
                unsafe_allow_html=True,
            )

            st.markdown(
                textwrap.dedent(f"""
<div class="welcome-card">
<div class="welcome-title">歡迎，{current_user} 👋</div>
<div class="welcome-sub">📅 今日日期：{tw_date_str}</div>

<div class="metrics-row">
<div class="metric-box">
<div class="metric-title">本年度剩餘特休</div>
<div class="metric-value">{remaining_text}</div>
</div>

<div class="metric-box">
<div class="metric-title">本月午餐差額</div>
<div class="metric-value">{lunch_text}</div>
</div>
</div>
</div>
"""),
                unsafe_allow_html=True,
            )


        with right:
            # =========================
            # 📢 公告區塊（固定框 + 捲軸；沒公告也保留；標題置中顯示在框內）
            # =========================
            ROW_HEIGHT = 44
            BOX_HEIGHT_PX = 10 * ROW_HEIGHT + 16

            ann = list_announcements(include_hidden=False, limit=200) if ANNOUNCE_DB_ID else []
            items = []
            for a in ann:
                ds = (a.get("發布日期") or "")[:10]
                content = (a.get("公告內容") or "").strip()
                if content:
                    items.append((ds, content))

            if (not ANNOUNCE_DB_ID):
                inner_html = """
                    <div class="announce-empty">公告</div>
                    <div style="text-align:center; color: rgba(20, 35, 55, 0.45); font-size: 12px; margin-top:-8px;">
                        尚未設定 ANNOUNCE_DB_ID
                    </div>
                """
            elif not items:
                inner_html = '<div class="announce-empty">公告</div>'
            else:
                rows_html = []
                for (ds, content) in items:
                    safe_content = sanitize_announce_text(content)
                    rows_html.append(textwrap.dedent(f"""
                    <div style="
                        display:flex;
                        gap:10px;
                        align-items:flex-start;
                        padding: 8px 6px;
                        border-bottom: 1px dashed rgba(0, 90, 150, 0.16);
                        line-height: 1.4;
                        font-size: 14px;
                    ">
                      <div style="
                        min-width: 92px;
                        color: rgba(20, 35, 55, 0.62);
                        font-variant-numeric: tabular-nums;
                      ">{ds}</div>
                      <div style="flex:1;">{safe_content}</div>
                    </div>
                    """).strip())
                inner_html = "\n".join(rows_html)

            st.markdown(
                textwrap.dedent(f"""
                <div class="announce-box" style="--announce-h:{BOX_HEIGHT_PX}px;">
                  <div class="announce-head">📢 公告</div>
                  <div class="announce-body">{inner_html}</div>
                </div>
                <style>
                  .announce-box{{
                    background: white;
                    border-radius: 18px;
                    box-shadow: 0 8px 24px rgba(0,0,0,0.08);
                    padding: 18px 18px 14px 18px;
                    height: var(--announce-h);
                    overflow: hidden;
                  }}
                  .announce-head{{
                    font-weight: 800;
                    text-align: center;
                    padding-bottom: 10px;
                    border-bottom: 1px dashed rgba(0, 90, 150, 0.18);
                    margin-bottom: 10px;
                  }}
                  .announce-body{{
                    height: calc(var(--announce-h) - 58px);
                    overflow: auto;
                    padding-right: 6px;
                  }}
                  .announce-empty{{
                    text-align:center;
                    font-weight: 800;
                    color: rgba(20, 35, 55, 0.55);
                    padding-top: 8px;
                  }}
                </style>
                """),
                unsafe_allow_html=True,
            )
    elif menu == "🗓️ 值班排班":
        render_duty_schedule_page()


    # -------------------------
    # 管理員：年度特休
    # -------------------------
    elif menu == "🗓️ 年度特休" and is_admin:
            st.header("年度特休（管理員）")

            this_year = datetime.now().year
            year = st.number_input("年度", min_value=2000, max_value=2100, value=this_year, step=1)

            # ✅ 右上角按鈕（像午餐管理那樣）
            employees = list_employee_names()

            @st.dialog("新增當年記錄（全員特休時數設定）")
            def add_year_vacation_dialog(default_year: int):
                # 1) 年度輸入（表單最上方）
                y = st.number_input("要設定的年度", min_value=2000, max_value=2100, value=int(default_year), step=1, key="vac_dialog_year")

                if not employees:
                    st.warning("⚠️ 抓不到員工清單，請確認【帳號管理表】已分享給 Integration")
                    return

                # 2) 先抓「該年度既有資料」，讓表單預設帶入目前值
                existing_rows = list_vacation_summary(is_admin=True, employee_name=current_user, year=int(y), limit=500)
                existing_map = {r.get("員工姓名", ""): r for r in (existing_rows or []) if r.get("員工姓名")}

                st.caption("左邊是員工姓名，右邊輸入該年度『本年度特休時數（小時）』。有資料會更新，沒有資料會新增。")

                inputs = {}
                for emp in employees:
                    default_total = 0.0
                    if emp in existing_map:
                        try:
                            default_total = float(existing_map[emp].get("本年度特休時數", 0.0) or 0.0)
                        except Exception:
                            default_total = 0.0

                    c1, c2 = st.columns([2, 3])
                    with c1:
                        st.write(emp)
                    with c2:
                        inputs[emp] = st.number_input(
                            label=f"vac_total_{emp}",
                            min_value=0.0,
                            step=1.0,
                            value=float(default_total),
                            label_visibility="collapsed",
                        )

                colA, colB = st.columns(2)

                if colA.button("✅ 一鍵寫入（新增/更新 Notion）", use_container_width=True):
                    ok_count = 0
                    for emp, total_hours in inputs.items():
                        emp = (emp or "").strip()
                        if not emp:
                            continue

                        total_hours = float(total_hours or 0.0)
                        used_hours = float(calc_used_vacation_hours(emp, int(y)) or 0.0)
                        remaining_hours = max(0.0, total_hours - used_hours)

                        try:
                            # 若有既有 row → update；否則 create
                            exist = existing_map.get(emp)
                            if exist and exist.get("_page_id"):
                                notion.pages.update(
                                    page_id=exist["_page_id"],
                                    properties={
                                        "員工姓名": {"title": [{"text": {"content": emp}}]},
                                        "年度": {"number": int(y)},
                                        "本年度特休時數": {"number": float(total_hours)},
                                        "已使用特休時數": {"number": float(used_hours)},
                                        "剩餘特休時數": {"number": float(remaining_hours)},
                                    },
                                )
                            else:
                                notion.pages.create(
                                    parent={"database_id": VACATION_DB_ID},
                                    properties={
                                        "員工姓名": {"title": [{"text": {"content": emp}}]},
                                        "年度": {"number": int(y)},
                                        "本年度特休時數": {"number": float(total_hours)},
                                        "已使用特休時數": {"number": float(used_hours)},
                                        "剩餘特休時數": {"number": float(remaining_hours)},
                                    },
                                )

                            ok_count += 1
                        except Exception as e:
                            st.error(f"❌ {emp} 寫入失敗：{e}")

                    st.success(f"✅ 已完成寫入：{ok_count}/{len(employees)} 位員工（年度 {int(y)}）")
                    log_action(current_user, "特休管理", f"新增/更新年度特休：{int(y)}（{ok_count}人）", "成功")
                    st.rerun()

                if colB.button("取消", use_container_width=True):
                    st.rerun()
                    
            @st.dialog("設定特休折算規則（寫入 Notion 特休折算表）")
            def set_cashout_rule_dialog(default_year: int):
                y = st.number_input("年份", min_value=2000, max_value=2100, value=int(default_year), step=1, key="rule_y")
                cap_days = st.number_input("可折算天數", min_value=0.0, max_value=365.0, value=5.0, step=0.5, key="rule_cap")
                hours_per_day = st.number_input("一天時數", min_value=1.0, max_value=24.0, value=8.0, step=0.5, key="rule_hpd")
                amount_per_hour = st.number_input("一小時折算金額", min_value=0.0, value=125.0, step=10.0, key="rule_aph")

                st.caption("送出後會：先檢查特休折算表是否有同年份 → 有則覆蓋、無則新增。")

                c1, c2 = st.columns(2)
                if c1.button("✅ 一鍵新增/覆蓋到 Notion", use_container_width=True):
                    ok = upsert_cashout_rule(
                        year=int(y),
                        cap_days=float(cap_days),
                        hours_per_day=float(hours_per_day),
                        amount_per_hour=float(amount_per_hour),
                        actor=current_user,
                    )
                    if ok:
                        st.success("✅ 已寫入 Notion 特休折算表")
                        st.rerun()

                if c2.button("取消", use_container_width=True):
                    st.rerun()


            # ✅ 右上角按鈕：新增當年記錄 + 設定折算規則
            topL, topR = st.columns([6, 4])
            with topR:
                b1, b2 = st.columns(2)

                with b1:
                    if st.button("➕ 新增當年記錄", use_container_width=True):
                        add_year_vacation_dialog(int(year))

                with b2:
                    if st.button("⚙️ 設定特休折算規則", use_container_width=True):
                        set_cashout_rule_dialog(int(year))


            st.divider()
            
            # ✅ 讀取該年度折算規則（來自 Notion 特休折算表）
            rule = get_cashout_rule_by_year(int(year))
            if rule:
                hours_per_day_rule = float(rule["一天時數"])
                cap_days_rule = float(rule["可折算天數"])
                amount_per_hour_rule = float(rule["一小時折算金額"])
            else:
                hours_per_day_rule = DEFAULT_HOURS_PER_DAY
                cap_days_rule = DEFAULT_CASHOUT_CAP_DAYS
                # 你原本是「每天金額」，但現在表存「每小時金額」
                amount_per_hour_rule = DEFAULT_CASHOUT_AMOUNT_PER_DAY / DEFAULT_HOURS_PER_DAY

            amount_per_day_rule = float(hours_per_day_rule) * float(amount_per_hour_rule)


            # ✅ 下面保留你原本的顯示/計算邏輯（不動）
            data = list_vacation_summary(is_admin=True, employee_name=current_user, year=int(year), limit=200)

            for row in data:
                name = row.get("員工姓名", "")
                if name:
                    used = calc_used_vacation_hours(name, int(year))
                    row["已使用特休時數"] = used
                    row["剩餘特休時數"] = max(0.0, float(row.get("本年度特休時數", 0.0) or 0.0) - used)

                cash = calc_cashout(
                    remaining_hours=row.get("剩餘特休時數", 0.0),
                    hours_per_day=hours_per_day_rule,
                    cap_days=cap_days_rule,
                    amount_per_day=amount_per_day_rule,
                    whole_days_only=True,  # 你目前需求只要整天（向下取整）
                )

                row["可折算天數(規則)"] = cash["cashout_days"]
                row["折算金額(規則)"] = cash["cashout_amount"]

            if data:
                display_data = [{k: v for k, v in row.items() if k != "_page_id"} for row in data]
            if not is_admin:
                display_data = strip_meta_columns(display_data)
            if not is_admin:
                display_data = strip_meta_columns(display_data)
                st.dataframe(display_data, use_container_width=True)
            else:
                st.info("目前查不到該年度的特休資料。")
                
    # -------------------------
    # 📢 公告管理（管理員）
    # -------------------------
    elif menu == "📢 公告管理" and is_admin:
        st.header("📢 公告管理（管理員）")

        if not ANNOUNCE_DB_ID:
            st.warning("⚠️ 尚未設定 ANNOUNCE_DB_ID（公告紀錄表 Database ID）")
            st.stop()

        # ✅ 右上角：新增公告
        @st.dialog("➕ 新增公告")
        def add_announce_dialog():
            pub = st.date_input("發布日期", value=date.today())
            content = st.text_area("公告內容", height=120, placeholder="輸入公告內容…")
            end = st.date_input("結束時間（可選）", value=date.today() + timedelta(days=7))
            use_end = st.checkbox("我要設定結束時間", value=True)

            c1, c2 = st.columns(2)
            if c1.button("✅ 新增", use_container_width=True):
                ok = create_announcement(
                    publish_date=pub,
                    content=content,
                    end_date=(end if use_end else None),
                    actor=current_user,
                )
                if ok:
                    st.success("✅ 已新增公告")
                    # 清快取：讓首頁立刻看到
                    try:
                        list_announcements.clear()
                    except Exception:
                        pass
                    st.rerun()

            if c2.button("取消", use_container_width=True):
                st.rerun()

        topL, topR = st.columns([7, 3])
        with topR:
            if st.button("➕ 新增公告", use_container_width=True):
                add_announce_dialog()

        st.divider()

        show_hidden = st.checkbox("顯示已隱藏（已完成 / 已過期）", value=False)
        rows = list_announcements(include_hidden=show_hidden, limit=300)

        if not rows:
            st.info("目前沒有公告。")
            st.stop()

        # ✅ 管理員表格顯示（含完成/結束時間）
        show = [{
            "發布日期": (r.get("發布日期") or "")[:10],
            "公告內容": r.get("公告內容") or "",
            "完成情況": bool(r.get("完成情況", False)),
            "結束時間": (r.get("結束時間") or "")[:10],
        } for r in rows]
        st.dataframe(show, use_container_width=True)

        st.divider()
        st.subheader("快速操作（勾完成 / 封存）")

        # 用 label 讓你挑選
        def _label(r: dict) -> str:
            ds = (r.get("發布日期") or "")[:10]
            done = "✅" if r.get("完成情況") else "⬜"
            content = (r.get("公告內容") or "").strip().replace("\n", " ")
            content = content[:30] + ("…" if len(content) > 30 else "")
            return f"{done} {ds}｜{content}"

        label_map = {_label(r): r for r in rows}
        sel = st.selectbox("選擇公告", list(label_map.keys()))
        picked = label_map[sel]
        pid = picked["_page_id"]

        c1, c2, c3 = st.columns([1, 1, 1])
        with c1:
            new_done = st.checkbox("標記為已完成（隱藏）", value=bool(picked.get("完成情況", False)))
            if st.button("✅ 更新完成狀態", use_container_width=True):
                ok = mark_announcement_done(pid, bool(new_done), actor=current_user)
                if ok:
                    try:
                        list_announcements.clear()
                    except Exception:
                        pass
                    st.success("✅ 已更新")
                    st.rerun()

        with c2:
            st.caption("封存 = 從 Notion 封存（通常不再使用）")
            confirm = st.checkbox("我確認要封存", key=f"ann_confirm_{pid}")
            if st.button("🗑️ 封存公告", use_container_width=True, disabled=not confirm):
                ok = archive_announcement(pid, actor=current_user)
                if ok:
                    try:
                        list_announcements.clear()
                    except Exception:
                        pass
                    st.success("✅ 已封存")
                    st.rerun()

        with c3:
            st.caption("提示：未勾完成，但到期（結束時間<=今天）也會自動隱藏")


    # -------------------------
    # 📍 每日打卡（管理員/員工都可用）
    # -------------------------
    elif menu == "📍 每日打卡":
        st.header("📍 GPS 打卡")

        # 基本檢查
        if not PUNCH_DB_ID:
            st.warning("⚠️ 尚未設定 PUNCH_DB_ID（Notion 打卡記錄表 Database ID）")
            st.stop()

        if COMPANY_LAT == 0 or COMPANY_LON == 0:
            st.warning("⚠️ 尚未設定公司座標 COMPANY_LAT / COMPANY_LON（請在 .env 設定）")
            st.stop()

        today = date.today()
        TW_TZ = timezone(timedelta(hours=8))
        tw_now = datetime.now(TW_TZ)
        today = tw_now.date()
        # -------------------------
        # ✅ Query Params 相容層（新版 st.query_params / 舊版 experimental）
        # -------------------------
        def _get_qp():
            try:
                return dict(st.query_params)
            except Exception:
                try:
                    return st.experimental_get_query_params()
                except Exception:
                    return {}

        def _clear_qp():
            try:
                st.query_params.clear()
            except Exception:
                try:
                    st.experimental_set_query_params()
                except Exception:
                    pass

        # -------------------------
        # ✅ 把 query params 的 lat/lon 同步進 session_state
        # -------------------------
        qp = _get_qp()

        if "gps_err" in qp:
            st.session_state["gps_err"] = str(qp.get("gps_err"))
            _clear_qp()

        if ("lat" in qp) and ("lon" in qp):
            try:
                st.session_state["gps_lat"] = float(qp.get("lat"))
                st.session_state["gps_lon"] = float(qp.get("lon"))
                st.session_state["gps_err"] = ""
            except Exception:
                st.session_state["gps_err"] = "GPS 座標解析失敗"
            _clear_qp()

        # -------------------------
        # ✅ 顯示定位狀態 + 取得/重新定位按鈕（重點：不再 st.stop 擋按鈕）
        # -------------------------
        gps_err = st.session_state.get("gps_err", "")
        lat = st.session_state.get("gps_lat")
        lon = st.session_state.get("gps_lon")

        # 用欄位排版讓畫面更像你截圖那種「乾淨」樣子
        infoL, infoR = st.columns([6, 4])

        with infoL:
            if gps_err:
                st.error(f"定位失敗：{gps_err}")
            elif (lat is None) or (lon is None):
                st.info("尚未取得定位：請按右側『📍 取得定位』並允許瀏覽器定位。")
            else:
                st.success("已取得定位 ✅")

        with infoR:
            btn_text = "📍 取得定位" if (lat is None or lon is None) else "🔄 重新定位"
            if st.button(btn_text, use_container_width=True):
                loc = get_geolocation()  # 會跳授權視窗

                if loc and "coords" in loc:
                    st.session_state["gps_lat"] = float(loc["coords"]["latitude"])
                    st.session_state["gps_lon"] = float(loc["coords"]["longitude"])
                    st.session_state["gps_err"] = ""
                    st.success("✅ 已取得定位")
                    st.rerun()
                else:
                    st.session_state["gps_err"] = "定位失敗或未授權（請確認瀏覽器定位權限）"
                    st.error(st.session_state["gps_err"])

        # -------------------------
        # ✅ 若還沒有定位：先不要給打卡按鈕（但頁面不會空）
        # -------------------------
        if (lat is None) or (lon is None) or gps_err:
            st.caption("※ 桌機瀏覽器 GPS 可能不準，建議用手機打卡。")
            st.stop()

        st.caption(f"公司座標：lat={COMPANY_LAT} lon={COMPANY_LON} 半徑={COMPANY_RADIUS_M}m")
        st.caption(f"目前座標：lat={lat} lon={lon}")


        # -------------------------
        # ✅ 計算距離 + 顯示「距離公司幾公尺」（你要的功能）
        # -------------------------
        dist = haversine_m(float(lat), float(lon), float(COMPANY_LAT), float(COMPANY_LON))
        gps_ok = dist <= float(COMPANY_RADIUS_M)

        m1, m2, m3 = st.columns(3)
        m1.metric("目前緯度", f"{float(lat):.6f}")
        m2.metric("目前經度", f"{float(lon):.6f}")
        m3.metric("距離公司(公尺)", f"{dist:.1f} m")

        st.caption(f"允許範圍：{int(COMPANY_RADIUS_M)} m｜狀態：{'✅ 可打卡' if gps_ok else '❌ 超出範圍'}")

        if not gps_ok:
            st.warning("你目前不在公司範圍內，因此無法打卡。")
            st.stop()

        st.divider()

        # -------------------------
        # ✅ 兩個按鈕：上班/下班（一天各一次）
        # -------------------------
        already_in = has_punch(current_user, today, "上班")
        already_out = has_punch(current_user, today, "下班")

        c1, c2 = st.columns(2)

        with c1:
            if already_in:
                st.button("✅ 今日已完成上班打卡", disabled=True, use_container_width=True)
            else:
                if st.button("🟢 上班打卡", use_container_width=True):
                    ok = create_punch_record(
                        employee_name=current_user,
                        punch_type="上班",
                        lat=float(lat),
                        lon=float(lon),
                        dist_m=float(dist),
                        passed=True,
                        note="",
                        actor=current_user,
                    )
                    if ok:
                        try:
                            has_punch.clear()
                        except Exception:
                            pass
                        try:
                            list_punch_records.clear()   # ✅ 新增：清掉查詢打卡紀錄的快取
                        except Exception:
                            pass

                        # ✅ 同步出勤：上班打卡 → 出勤記錄表（出席/遲到）
                        try:
                            if ATTEND_DB_ID:
                                tw_now = datetime.now(timezone(timedelta(hours=8)))
                                cutoff = tw_now.replace(hour=8, minute=6, second=0, microsecond=0)
                                att_status = ATTEND_PRESENT_STATUS if tw_now <= cutoff else ATTEND_LATE_STATUS
                                upsert_attendance_record(current_user, tw_now.date(), att_status, actor=current_user)
                        except Exception:
                            pass

                        # ✅ 同步出勤：上班打卡 → 出勤記錄表（出席/遲到）
                        try:
                            if ATTEND_DB_ID:
                                tw_now = datetime.now(timezone(timedelta(hours=8)))
                                cutoff = tw_now.replace(hour=8, minute=6, second=0, microsecond=0)
                                att_status = ATTEND_PRESENT_STATUS if tw_now <= cutoff else ATTEND_LATE_STATUS
                                upsert_attendance_record(current_user, tw_now.date(), att_status, actor=current_user)
                        except Exception:
                            pass

                        st.success("✅ 上班打卡完成")
                        st.rerun()

        with c2:
            if already_out:
                st.button("✅ 今日已完成下班打卡", disabled=True, use_container_width=True)
            else:
                if st.button("🔴 下班打卡", use_container_width=True):
                    ok = create_punch_record(
                        employee_name=current_user,
                        punch_type="下班",
                        lat=float(lat),
                        lon=float(lon),
                        dist_m=float(dist),
                        passed=True,
                        note="",
                        actor=current_user,
                    )
                    if ok:
                        try:
                            has_punch.clear()
                        except Exception:
                            pass
                        try:
                            list_punch_records.clear()   # ✅ 新增：清掉查詢打卡紀錄的快取
                        except Exception:
                            pass

                        st.success("✅ 下班打卡完成")
                        st.rerun()

        st.divider()

        # -------------------------
        # ✅ 查詢我的打卡紀錄
        # -------------------------
        st.subheader("查詢我的打卡紀錄（最近 6 個月）")
        now_dt = datetime.now()
        y = st.number_input("年份", min_value=2000, max_value=2100, value=int(now_dt.year), step=1, key="punch_y")
        m = st.number_input("月份", min_value=1, max_value=12, value=int(now_dt.month), step=1, key="punch_m")

        diff = (now_dt.year - int(y)) * 12 + (now_dt.month - int(m))
        if diff < 0:
            st.warning("⚠️ 不能查未來月份")
        elif diff > 5:
            st.warning("⚠️ 最多只能查最近 6 個月")
        else:
            rows = list_punch_records(current_user, int(y), int(m), limit=500)
            if rows:
                st.dataframe(strip_meta_columns(rows), use_container_width=True)
            else:
                st.info("本月沒有打卡紀錄。")

        st.caption("※ 提醒：桌機瀏覽器 GPS 可能不準，建議用手機打卡。")


    # -------------------------
    # 請假申請
    # -------------------------
    elif menu == "📝 請假申請":
        st.header("請假")

        status_options = get_select_options(LEAVE_DB_ID, "狀態") or ["待審核", "通過", "退回"]
        leave_type_options = get_select_options(LEAVE_DB_ID, "假別") or ["特休", "病假", "事假"]

        employee_options = []
        if is_admin:
            employee_options = list_employee_names()

        this_year = datetime.now().year
        vac_year = st.number_input("年度（特休統計）", min_value=2000, max_value=2100, value=this_year, step=1)

        snap = get_employee_vacation_snapshot(current_user, int(vac_year))
        st.subheader("我的年度特休（先看再請假）")
        if snap:
            m1, m2, m3 = st.columns(3)
            m1.metric("本年度特休", f"{int(snap['total'])} 小時")
            m2.metric("已使用", f"{int(snap['used'])} 小時")
            m3.metric("剩餘", f"{int(snap['remaining'])} 小時")
        else:
            st.warning("⚠️ 讀取年度特休資料失敗，請稍後再試。")

        st.divider()

        @st.dialog("➕ 新增請假（管理員代填）")
        def admin_create_leave_dialog():
            target_employee = st.selectbox("請假人（代填）", employee_options if employee_options else [current_user], index=0)

            col1, col2 = st.columns(2)
            leave_type = col1.selectbox("假別", leave_type_options)
            hours = col2.number_input("請假時數", min_value=1, step=1)

            st.markdown("**請假期間（到小時）**")
            c1, c2, c3, c4 = st.columns(4)
            start_date = c1.date_input("開始日期", value=date.today())
            start_hour = c2.selectbox("開始時間（時）", list(range(0, 24)), index=9)
            end_date = c3.date_input("結束日期", value=date.today())
            end_hour = c4.selectbox("結束時間（時）", list(range(0, 24)), index=18)

            reason = st.text_area("請假事由")

            btn1, btn2 = st.columns(2)
            submit = btn1.button("✅ 送出申請", use_container_width=True)
            cancel = btn2.button("取消", use_container_width=True)
            if cancel:
                st.rerun()

            if submit:
                start_dt = datetime.combine(start_date, datetime.min.time()).replace(hour=int(start_hour))
                end_dt = datetime.combine(end_date, datetime.min.time()).replace(hour=int(end_hour))

                if end_dt <= start_dt:
                    st.error("❌ 結束時間必須晚於開始時間")
                    return

                if leave_type == "特休":
                    if start_dt.year != end_dt.year:
                        st.error("❌ 特休申請目前不支援跨年度，請拆成兩筆（或調整日期）")
                        return
                    ok, msg = validate_vacation_enough(target_employee, start_dt.year, int(hours))
                    if not ok:
                        st.error(msg)
                        return

                ok = create_leave_request(
                    employee_name=current_user,
                    target_employee_name=target_employee,
                    created_by=current_user,
                    leave_type=leave_type,
                    hours=hours,
                    start_dt=start_dt,
                    end_dt=end_dt,
                    reason=reason,
                )

                if ok:
                    st.success("✅ 請假申請已送出（待審核）")
                    st.rerun()

        top_left, top_right = st.columns([7, 3])
        with top_right:
            if is_admin:
                if st.button("➕ 新增請假", use_container_width=True):
                    admin_create_leave_dialog()

        st.subheader("請假紀錄")
        data = list_leave_requests(is_admin=is_admin, employee_name=current_user, limit=50)
        if data:
            display_data = [{k: v for k, v in row.items() if k != "_page_id"} for row in data]
            if not is_admin:
                display_data = strip_meta_columns(display_data)
            st.dataframe(display_data, use_container_width=True)
        else:
            st.info("目前沒有請假紀錄。")

        if is_admin and data:
            st.subheader("管理員審核（改狀態 / 刪除）")
            label_to_row = {make_leave_label(r): r for r in data}
            labels = list(label_to_row.keys())
            selected_label = st.selectbox("選擇要操作的請假紀錄", labels)
            selected_row = label_to_row[selected_label]
            page_id = selected_row["_page_id"]

            colA, colB = st.columns([2, 1])
            with colA:
                new_status = st.selectbox("更新為狀態", status_options)
                if st.button("✅ 更新狀態", use_container_width=True):
                    ok = update_leave_status(page_id, new_status, actor=current_user)
                    if ok:
                        st.success("✅ 狀態已更新")
                        st.rerun()
            with colB:
                st.markdown("**刪除（封存）**")
                confirm = st.checkbox("我確認要刪除此筆", key=f"confirm_del_{page_id}")
                if st.button("🗑️ 刪除這筆", use_container_width=True, disabled=not confirm):
                    ok = delete_leave_request(page_id, actor=current_user)
                    if ok:
                        st.success("✅ 已刪除（封存）")
                        st.rerun()

    # -------------------------
    # ✅ 管理員：出勤記錄
    # -------------------------
    elif menu == "📅 出勤記錄" and is_admin:
        st.header("📅 出勤記錄（管理員）")

        if not ATTEND_DB_ID:
            st.warning("⚠️ 尚未設定 ATTEND_DB_ID（出勤記錄表 Database ID）")
            st.stop()

        # ✅ 先抓員工清單（可同步刷新）
        topL, topR = st.columns([7, 3])
        with topR:
            if st.button("🔄 同步所有員工資料", use_container_width=True):
                # 清掉 cache，重新抓 Notion 員工
                try:
                    list_employee_names.clear()
                except Exception:
                    pass
                st.success("✅ 已同步員工清單")
                st.rerun()

        employees = list_employee_names()
        if not employees:
            st.warning("⚠️ 抓不到員工清單，請確認【帳號管理表】已分享給 Integration")
            st.stop()

        # ✅ 出勤狀態（若 Notion 有選項就用 Notion 的）
        status_options = get_select_options(ATTEND_DB_ID, "出勤狀態") or [ATTEND_PRESENT_STATUS, ATTEND_LEAVE_STATUS, ATTEND_LATE_STATUS]
        default_idx = status_options.index(ATTEND_PRESENT_STATUS) if ATTEND_PRESENT_STATUS in status_options else 0

        @st.dialog("新增當日出勤（全員）")
        def add_today_attendance_dialog():
            attend_date = st.date_input("出勤日期", value=date.today())
            # ✅ 先抓 Notion 既有紀錄：有就預設呈現，沒有就預設『出席』
            existing_map = get_attendance_status_map_by_date(attend_date)
            st.caption("每位員工三選一（橫向）。若 Notion 當天已有紀錄會先帶出；沒有則預設『出席』。")

            selections = {}
            for emp in employees:
                c1, c2 = st.columns([2, 6])
                with c1:
                    st.write(emp)
                with c2:
                    selections[emp] = st.radio(
                        label=f"att_{emp}",
                        options=status_options,
                        index=(status_options.index(existing_map.get(emp)) if existing_map.get(emp) in status_options else default_idx),
                        horizontal=True,
                        label_visibility="collapsed",
                    )

            colA, colB = st.columns(2)
            if colA.button("✅ 一鍵新增（寫入 Notion）", use_container_width=True):
                ok_count = 0
                for emp, stt in selections.items():
                    if upsert_attendance_record(emp, attend_date, stt, actor=current_user):
                        ok_count += 1
                st.success(f"✅ 已新增 {ok_count}/{len(employees)} 筆出勤")
                st.rerun()

            if colB.button("取消", use_container_width=True):
                st.rerun()

        # ✅ 上方按鈕：新增出勤
        btnL, btnR = st.columns([7, 3])
        with btnR:
            if st.button("➕ 新增出勤", use_container_width=True):
                add_today_attendance_dialog()

        st.divider()

        # =========================
        # ✅ 查詢：年/月 + 員工篩選
        # =========================
        now = datetime.now()
        q1, q2, q3 = st.columns([1, 1, 2])

        with q1:
            y = st.number_input("年份", min_value=2000, max_value=2100, value=int(now.year), step=1, key="att_y_admin")
        with q2:
            m = st.number_input("月份", min_value=1, max_value=12, value=int(now.month), step=1, key="att_m_admin")
        with q3:
            emp_filter = st.selectbox("員工（可篩選）", ["全部員工"] + employees, index=0, key="att_emp_admin")

        start_d, end_d = _month_range(int(y), int(m))

        # ✅ 查詢顯示（可篩某員工）
        rows = list_attendance_records(start_d=start_d, end_d=end_d, employee_name=emp_filter, limit=2000)

        st.subheader(f"{int(y)} 年 {int(m)} 月 出勤清單（{emp_filter}）")
        if rows:
            display = [{k: v for k, v in r.items() if k != "_page_id"} for r in rows]
            st.dataframe(display, use_container_width=True)

            # ✅ 匯出目前篩選結果
            file_bytes, file_name = make_excel_bytes(
                display,
                filename_hint=f"attendance_{int(y)}_{int(m)}_{('all' if emp_filter=='全部員工' else emp_filter)}.xlsx"
            )
            st.download_button(
                "📤 匯出 Excel（目前篩選結果）",
                data=file_bytes,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.info("本條件下沒有出勤資料。")

        st.divider()

        # =========================
        # ✅ 匯出：整個月份「全部員工」
        # =========================
        st.subheader("📦 月報匯出（全部員工）")
        all_rows = list_attendance_records(start_d=start_d, end_d=end_d, employee_name="全部員工", limit=5000)

        if all_rows:
            export_all = [{k: v for k, v in r.items() if k != "_page_id"} for r in all_rows]
            file_bytes2, file_name2 = make_excel_bytes(
                export_all,
                filename_hint=f"attendance_month_all_{int(y)}_{int(m)}.xlsx"
            )
            st.download_button(
                "📤 匯出 Excel（本月全部員工出勤）",
                data=file_bytes2,
                file_name=file_name2,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            st.caption("（這個就是你每月結算要匯出的那份）")
        else:
            st.info("本月目前沒有任何員工出勤資料可匯出。")

    # -------------------------
    # 🍱 管理員：午餐管理（全員月結/清單 + 一鍵新增當日記錄）
    # -------------------------
    elif menu == "🍱 午餐管理" and is_admin:
        st.header("🍱 午餐管理（管理員）")

        if not LUNCH_DB_ID:
            st.warning("⚠️ 尚未設定 LUNCH_DB_ID（午餐訂餐表 Database ID），目前無法使用午餐功能。")
            st.stop()

        employees = list_employee_names()
        if not employees:
            st.warning("⚠️ 抓不到員工清單，請確認【帳號管理表】已分享給 Integration")
            st.stop()

        @st.dialog("新增當日午餐（全員）")
        def add_today_lunch_dialog():
            lunch_date = st.date_input("訂餐日期", value=date.today())
            st.caption("每位員工填當天吃的金額，最後一鍵送出。")

            amounts = {}
            for emp in employees:
                c1, c2 = st.columns([2, 3])
                with c1:
                    st.write(emp)
                with c2:
                    amounts[emp] = st.number_input(
                        label=f"l_amt_{emp}",
                        min_value=0.0,
                        step=10.0,
                        value=0.0,
                        label_visibility="collapsed",
                    )

            colA, colB = st.columns(2)
            if colA.button("✅ 一鍵送出（寫入 Notion）", use_container_width=True):
                ok_count = 0
                for emp, amt in amounts.items():
                    if create_lunch_record(emp, lunch_date, float(amt), actor=current_user):
                        ok_count += 1
                st.success(f"✅ 已新增 {ok_count}/{len(employees)} 筆午餐")
                st.rerun()

            if colB.button("取消", use_container_width=True):
                st.rerun()

        topL, topR = st.columns([7, 3])
        with topR:
            if st.button("➕ 新增當日記錄", use_container_width=True):
                add_today_lunch_dialog()

        st.divider()

        now = datetime.now()
        y = st.number_input("年份", min_value=2000, max_value=2100, value=int(now.year), step=1, key="l_y_admin")
        m = st.number_input("月份", min_value=1, max_value=12, value=int(now.month), step=1, key="l_m_admin")

        if not ATTEND_DB_ID:
            st.warning("⚠️ 目前尚未設定 ATTEND_DB_ID（出勤記錄表），所以午餐月結算無法用『出席/遲到×90』計算。")

        st.subheader("全員月結算（可領工餐×90 - 訂餐金額）")
        settlements = []
        for emp in employees:
            settlements.append(calc_month_lunch_settlement(emp, int(y), int(m), is_admin=True))

        if settlements:
            st.dataframe(settlements, use_container_width=True)

            file_bytes, file_name = make_excel_bytes(
                settlements,
                filename_hint=f"lunch_settlement_{int(y)}_{int(m)}.xlsx"
            )
            st.download_button(
                "📤 匯出 Excel（午餐月結算）",
                data=file_bytes,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        st.divider()
        st.subheader("查詢某員工訂餐清單")
        emp = st.selectbox("員工姓名", employees, key="l_admin_emp")
        start_d, end_d = _month_range(int(y), int(m))
        rows = list_lunch_records(is_admin=True, employee_name=emp, start_d=start_d, end_d=end_d, limit=300)
        if rows:
            display = [{k: v for k, v in r.items() if k != "_page_id"} for r in rows]
            st.dataframe(display, use_container_width=True)
        else:
            st.info("該員工本月尚無訂餐紀錄。")

        st.caption("差額(應得-已訂餐) > 0：公司補；< 0：員工補（從薪水加/扣）")

    # -------------------------
    # 🍱 員工：午餐紀錄（年/月選擇 + 清單三欄）
    # -------------------------
    elif menu == "🍱 午餐紀錄" and (not is_admin):
        st.header("🍱 午餐紀錄（員工）")

        if not LUNCH_DB_ID:
            st.warning("⚠️ 尚未設定 LUNCH_DB_ID（午餐訂餐表 Database ID），目前無法查看午餐紀錄。")
            st.stop()

        now = datetime.now()
        y = st.number_input("年份", min_value=2000, max_value=2100, value=int(now.year), step=1, key="l_y_emp_view")
        m = st.number_input("月份", min_value=1, max_value=12, value=int(now.month), step=1, key="l_m_emp_view")

        start_d, end_d = _month_range(int(y), int(m))
        rows = list_lunch_records(is_admin=False, employee_name=current_user, start_d=start_d, end_d=end_d, limit=300)

        st.subheader(f"{int(y)} 年 {int(m)} 月 午餐訂餐清單")
        if rows:
            display = [{
                "員工姓名": r.get("員工姓名", ""),
                "訂餐金額": r.get("訂餐金額", 0),
                "訂餐日期": r.get("訂餐日期", ""),
            } for r in rows]
            st.dataframe(display, use_container_width=True)
        else:
            st.info("本月尚無訂餐紀錄。")

    # -------------------------
    # ✅ 管理員：薪資計算表（對齊最新 Notion 欄位）
    # -------------------------
    elif menu == "💵 薪資計算" and is_admin:
        st.header("薪資計算（管理員）")

        employees = list_employee_names()
        if not employees:
            st.warning("⚠️ 抓不到員工清單，請確認【帳號管理表】已分享給 Integration")
            st.stop()

        # -------------------------
        # 欄位名稱（對齊 Notion）
        # -------------------------
        FIELD_EMP = "員工姓名"
        FIELD_Y = "薪資年份"
        FIELD_M = "薪資月份"

        FIELD_BASE = "全薪"
        FIELD_LEADER = "負責人職務津貼"  
        FIELD_JOB = "職務津貼"
        FIELD_PERF = "績效獎金"
        FIELD_TRAFFIC = "交通津貼"
        FIELD_SALES = "營業津貼"
        FIELD_COOP = "配合"
        FIELD_ATTEND = "全勤獎金"
        FIELD_CERT = "證照加給"
        FIELD_MEAL = "伙食津貼"
        FIELD_OT_WEEKDAY = "平日(中晚)加班費"
        FIELD_OT_SAT = "週六加班費"
        FIELD_SOCIAL = "交際費"
        FIELD_YEAR_END = "年終補助"

        FIELD_GROSS = "薪資總計"

        FIELD_ADVANCE = "借支"
        FIELD_SICK = "病假請假"
        FIELD_PERSONAL = "事假請假"
        FIELD_LOAN_INT = "借款利息"
        FIELD_LATE = "遲到/早退"
        FIELD_LABOR = "勞保費"
        FIELD_HEALTH = "健保費"
        FIELD_OTHER_DED = "其他"

        FIELD_DEDUCT = "應扣總計"
        FIELD_NET = "實發金額"
        FIELD_NOTE = "備註"

        def _to_float(v, default=0.0) -> float:
            try:
                if v is None or v == "":
                    return float(default)
                return float(v)
            except Exception:
                return float(default)

        now = datetime.now()
        y = st.number_input("薪資年份", min_value=2000, max_value=2100, value=int(now.year), step=1, key="calc_y")
        m = st.number_input("薪資月份", min_value=1, max_value=12, value=int(now.month), step=1, key="calc_m")
        target_employee = st.selectbox("員工姓名", employees, key="calc_emp")

        # ✅ 讓「自動帶入」在切換 年/月/員工 時能正確刷新：所有關鍵輸入元件改用「依年月員工」的 key
        _emp_key = (str(target_employee).strip() or "—").replace(" ", "_")
        salary_prefix = f"salary_{int(y)}_{int(m)}_{_emp_key}"

        st.subheader("計算區（輸入/調整）")
        existing = get_salary_record(target_employee, int(y), int(m))

        # -------------------------
        # 先從 Notion 既有紀錄帶入
        # -------------------------
        base_salary = _to_float(existing.get(FIELD_BASE)) if existing else 0.0
        leader_allowance = _to_float(existing.get(FIELD_LEADER)) if existing else 0.0
        job_allowance = _to_float(existing.get(FIELD_JOB)) if existing else 0.0
        perf_bonus = _to_float(existing.get(FIELD_PERF)) if existing else 0.0
        traffic_allowance = _to_float(existing.get(FIELD_TRAFFIC)) if existing else 0.0
        sales_allowance = _to_float(existing.get(FIELD_SALES)) if existing else 0.0

        coop = _to_float(existing.get(FIELD_COOP)) if existing else 0.0
        attend_bonus = _to_float(existing.get(FIELD_ATTEND)) if existing else 0.0
        cert_allowance = _to_float(existing.get(FIELD_CERT)) if existing else 0.0

        # 🍱 伙食津貼：預設用「午餐月結差額」自動帶入（出席/遲到天數×90 - 已訂餐金額）
        # 規則：
        # - 若 Notion 薪資表（既有紀錄）已經有值 → 尊重既有值（不覆蓋）
        # - 若既有值為空 / 不存在 → 依規則自動計算後帶入
        auto_meal = 0.0
        try:
            auto_meal = _to_float(get_month_lunch_amount(target_employee, int(y), int(m), is_admin=bool(is_admin)), 0.0)
        except Exception:
            auto_meal = 0.0

        existing_meal_raw = existing.get(FIELD_MEAL) if existing else None
        has_existing_meal = existing_meal_raw is not None
        meal_allowance = _to_float(existing_meal_raw) if has_existing_meal else float(auto_meal)


        ot_weekday = _to_float(existing.get(FIELD_OT_WEEKDAY)) if existing else 0.0


        # ✅ 自動帶入：平日(中晚)加班費（由 值班排班表 + 加班設定表 推算）
        # 規則：
        # - 若 Notion 薪資表（既有紀錄）已經有值 → 尊重既有值（不覆蓋）
        # - 若既有值為空 / 不存在 → 依規則自動計算後帶入
        calc = calc_weekday_ot_from_duty(target_employee, int(y), int(m))
        suggested_ot_weekday = float(calc.get("amount") or 0.0)

        # existing.get(...) 可能是 None：代表 Notion 這格還沒填
        existing_ot_raw = existing.get(FIELD_OT_WEEKDAY) if existing else None
        has_existing_value = existing_ot_raw is not None

        if not has_existing_value:
            ot_weekday = suggested_ot_weekday






        ot_sat = _to_float(existing.get(FIELD_OT_SAT)) if existing else 0.0
        social_fee = _to_float(existing.get(FIELD_SOCIAL)) if existing else 0.0
        year_end = _to_float(existing.get(FIELD_YEAR_END)) if existing else 0.0

        advance = _to_float(existing.get(FIELD_ADVANCE)) if existing else 0.0
        sick_leave = _to_float(existing.get(FIELD_SICK)) if existing else 0.0
        personal_leave = _to_float(existing.get(FIELD_PERSONAL)) if existing else 0.0
        loan_interest = _to_float(existing.get(FIELD_LOAN_INT)) if existing else 0.0
        late_early = _to_float(existing.get(FIELD_LATE)) if existing else 0.0
        labor_fee = _to_float(existing.get(FIELD_LABOR)) if existing else 0.0
        health_fee = _to_float(existing.get(FIELD_HEALTH)) if existing else 0.0
        other_ded = _to_float(existing.get(FIELD_OTHER_DED)) if existing else 0.0

        note = str(existing.get(FIELD_NOTE, "")) if existing else ""

        # -------------------------
        # UI：加項
        # -------------------------
        st.markdown("### ✅ 加項（薪資/津貼/獎金）")
        r1c1, r1c2, r1c3 = st.columns(3)
        base_salary = r1c1.number_input(FIELD_BASE, min_value=0.0, step=1000.0, value=float(base_salary))
        leader_allowance = r1c2.number_input(FIELD_LEADER, min_value=0.0, step=100.0, value=float(leader_allowance))
        job_allowance = r1c3.number_input(FIELD_JOB, min_value=0.0, step=100.0, value=float(job_allowance))

        r2c1, r2c2, r2c3 = st.columns(3)
        perf_bonus = r2c1.number_input(FIELD_PERF, min_value=0.0, step=100.0, value=float(perf_bonus))
        traffic_allowance = r2c2.number_input(FIELD_TRAFFIC, min_value=0.0, step=100.0, value=float(traffic_allowance))
        sales_allowance = r2c3.number_input(FIELD_SALES, min_value=0.0, step=100.0, value=float(sales_allowance))

        r3c1, r3c2, r3c3 = st.columns(3)
        coop = r3c1.number_input(FIELD_COOP, step=100.0, value=float(coop))
        attend_bonus = r3c2.number_input(FIELD_ATTEND, step=100.0, value=float(attend_bonus))
        cert_allowance = r3c3.number_input(FIELD_CERT, step=100.0, value=float(cert_allowance))

        r4c1, r4c2, r4c3, r4c4 = st.columns(4)
        meal_allowance = r4c1.number_input(FIELD_MEAL, step=50.0, value=float(meal_allowance), key=f"{salary_prefix}_meal")
        ot_weekday = r4c2.number_input(FIELD_OT_WEEKDAY, step=100.0, value=float(ot_weekday), key=f"{salary_prefix}_ot_weekday")
        ot_sat = r4c3.number_input(FIELD_OT_SAT, step=100.0, value=float(ot_sat))
        social_fee = r4c4.number_input(FIELD_SOCIAL, step=100.0, value=float(social_fee))

        year_end = st.number_input(FIELD_YEAR_END, step=1000.0, value=float(year_end))

        # -------------------------
        # UI：扣項
        # -------------------------
        st.markdown("### ✅ 扣項（借支/請假/保費/其他）")
        d1, d2, d3, d4 = st.columns(4)
        advance = d1.number_input(FIELD_ADVANCE, step=100.0, value=float(advance))
        sick_leave = d2.number_input(FIELD_SICK, step=100.0, value=float(sick_leave))
        personal_leave = d3.number_input(FIELD_PERSONAL, step=100.0, value=float(personal_leave))
        loan_interest = d4.number_input(FIELD_LOAN_INT, step=100.0, value=float(loan_interest))

        d5, d6, d7, d8 = st.columns(4)
        late_early = d5.number_input(FIELD_LATE, step=50.0, value=float(late_early))
        labor_fee = d6.number_input(FIELD_LABOR, step=50.0, value=float(labor_fee))
        health_fee = d7.number_input(FIELD_HEALTH, step=50.0, value=float(health_fee))
        other_ded = d8.number_input(FIELD_OTHER_DED, step=100.0, value=float(other_ded))

        note = st.text_input(FIELD_NOTE, value=note)

        # -------------------------
        # 計算：薪資總計 / 應扣總計 / 實發金額
        # -------------------------
        gross_total = (
            float(base_salary)
            + float(leader_allowance)
            + float(job_allowance)
            + float(perf_bonus)
            + float(traffic_allowance)
            + float(sales_allowance)
            + float(coop)
            + float(attend_bonus)
            + float(cert_allowance)
            + float(meal_allowance)
            + float(ot_weekday)
            + float(ot_sat)
            + float(social_fee)
            + float(year_end)
        )

        deduct_total = (
            float(advance)
            + float(sick_leave)
            + float(personal_leave)
            + float(loan_interest)
            + float(late_early)
            + float(labor_fee)
            + float(health_fee)
            + float(other_ded)
        )

        net_pay = gross_total - deduct_total

        st.success(f"✅ {FIELD_GROSS} = ${gross_total:,.0f}")
        st.info(f"✅ {FIELD_DEDUCT} = ${deduct_total:,.0f}")
        st.success(f"✅ {FIELD_NET} = ${net_pay:,.0f}")
        st.caption(f"計算：{FIELD_NET} = {FIELD_GROSS} − {FIELD_DEDUCT}")

        # -------------------------
        # 儲存
        # -------------------------
        colA, colB = st.columns([2, 3])
        with colA:
            if st.button("💾 儲存到 Notion 薪資表並匯出Excel", use_container_width=True):
                # 這裡直接把「最新 Notion 欄位」寫回去
                payload = {
                    FIELD_EMP: target_employee,
                    FIELD_Y: int(y),
                    FIELD_M: int(m),

                    FIELD_BASE: float(base_salary),
                    FIELD_LEADER: float(leader_allowance),
                    FIELD_JOB: float(job_allowance),
                    FIELD_PERF: float(perf_bonus),
                    FIELD_TRAFFIC: float(traffic_allowance),
                    FIELD_SALES: float(sales_allowance),

                    FIELD_COOP: float(coop),
                    FIELD_ATTEND: float(attend_bonus),
                    FIELD_CERT: float(cert_allowance),
                    FIELD_MEAL: float(meal_allowance),
                    FIELD_OT_WEEKDAY: float(ot_weekday),
                    FIELD_OT_SAT: float(ot_sat),
                    FIELD_SOCIAL: float(social_fee),
                    FIELD_YEAR_END: float(year_end),

                    FIELD_GROSS: float(gross_total),

                    FIELD_ADVANCE: float(advance),
                    FIELD_SICK: float(sick_leave),
                    FIELD_PERSONAL: float(personal_leave),
                    FIELD_LOAN_INT: float(loan_interest),
                    FIELD_LATE: float(late_early),
                    FIELD_LABOR: float(labor_fee),
                    FIELD_HEALTH: float(health_fee),
                    FIELD_OTHER_DED: float(other_ded),

                    FIELD_DEDUCT: float(deduct_total),
                    FIELD_NET: float(net_pay),
                    FIELD_NOTE: note,
                }

                ok = upsert_salary_record(
                    employee_name=target_employee,
                    y=int(y),
                    m=int(m),
                    data=payload,        # ✅ 建議你的 upsert_salary_record 支援用 data dict 寫入
                    actor=current_user,
                )

                if ok:
                    st.success("✅ 已儲存/更新 Notion 薪資資料（已產生 Excel 備份）")
                    export_rows = [dict(payload)]
                    file_bytes, file_name = make_excel_bytes(
                        export_rows,
                        filename_hint=f"salary_{target_employee}_{int(y)}_{int(m)}.xlsx"
                    )
                    st.download_button(
                        "⬇️ 下載 Excel 備份",
                        data=file_bytes,
                        file_name=file_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )


        with colB:
            st.caption("提示：若 Notion 薪資表欄位名稱有差一個字，會導致寫入被略過。請以 Notion 欄位名稱為準。")

        # -------------------------
        # 查詢 / 匯出
        # -------------------------
        st.divider()
        st.subheader("薪資清單（查詢/匯出）")

        qcol1, qcol2 = st.columns([1, 2])
        with qcol1:
            query_year = st.number_input(
                "查詢年份",
                min_value=2000,
                max_value=2100,
                value=int(now.year),
                step=1,
                key="list_y",
            )
        with qcol2:
            q_month = st.selectbox(
                "查詢月份（可選）",
                ["全部"] + [str(i) for i in range(1, 13)],
                index=0,
                key="list_m",
            )

        m_filter = None if q_month == "全部" else int(q_month)

        rows = list_salary_records(is_admin=True, employee_name=current_user, y=int(query_year), m=m_filter, limit=200)

        if rows:
            display = [{k: v for k, v in r.items() if k != "_page_id"} for r in rows]
            st.dataframe(display, use_container_width=True)

            file_bytes, file_name = make_excel_bytes(
                display,
                filename_hint=f"salary_{int(query_year)}_{('all' if m_filter is None else int(m_filter))}.xlsx"
            )
            st.download_button(
                "📤 匯出 Excel",
                data=file_bytes,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.info("目前查不到薪資資料。")


    # -------------------------
    # ✅ 員工：薪資查詢（新版 Notion 欄位）
    # -------------------------
    elif menu == "💰 薪資查詢" and (not is_admin):
        st.header("薪資查詢（員工）")

        now = datetime.now()
        y = st.number_input("薪資年份", min_value=2000, max_value=2100, value=int(now.year), step=1)
        m = st.number_input("薪資月份", min_value=1, max_value=12, value=int(now.month), step=1)

        existing = get_salary_record(current_user, int(y), int(m))
        if not existing:
            st.info("目前查不到該月薪資資料。")
        else:
            # ===== 小工具：安全取值 / 找欄位 =====
            def _num(v, default=0.0):
                try:
                    if v is None or v == "":
                        return float(default)
                    return float(v)
                except Exception:
                    return float(default)

            def _money(v):
                return f"${_num(v):.0f}"

            def _find_key_by_prefix(d: dict, prefix: str):
                for k in (d or {}).keys():
                    if isinstance(k, str) and k.startswith(prefix):
                        return k
                return None

            def _get_first_existing_key(d: dict, candidates: list[str], prefix: str | None = None):
                # 1) 先用候選名精準匹配
                for k in candidates:
                    if k in d:
                        return k
                # 2) 再用 prefix 模糊找（避免 Notion 欄位顯示被截斷）
                if prefix:
                    k2 = _find_key_by_prefix(d, prefix)
                    if k2:
                        return k2
                # 3) 找不到就回傳第一個候選（用來顯示時 get() 會吃到 0）
                return candidates[0] if candidates else None

        

            st.subheader("我的薪資明細")

            # ===== 加項（新版 Notion）=====
            add_items = [
                ("全薪", "全薪"),
                ("負責人職務津貼", "負責人職務津貼"),
                ("職務津貼", "職務津貼"),
                ("績效獎金", "績效獎金"),
                ("交通津貼", "交通津貼"),
                ("營業津貼", "營業津貼"),
                ("配合", "配合"),
                ("全勤獎金", "全勤獎金"),
                ("證照加給", "證照加給"),
                ("伙食津貼", "伙食津貼"),
                ("平日(中晚)加班費", "平日(中晚)加班費"),
                ("週六加班費", "週六加班費"),
                ("交際費", "交際費"),
                ("年終補助", "年終補助"),
            ]

            st.markdown("### ➕ 加項")
            cols = st.columns(4)
            for i, (label, key) in enumerate(add_items):
                v = existing.get(key, 0)
                cols[i % 4].metric(label, _money(v))

            # ===== 扣項（新版 Notion）=====
            deduct_items = [
                ("借支", "借支"),
                ("病假請假", "病假請假"),
                ("事假請假", "事假請假"),
                ("借款利息", "借款利息"),
                ("遲到/早退", "遲到/早退"),
                ("勞保費", "勞保費"),
                ("健保費", "健保費"),
                ("其他", "其他"),
            ]

            st.markdown("### ➖ 扣項")
            cols2 = st.columns(4)
            for i, (label, key) in enumerate(deduct_items):
                v = existing.get(key, 0)
                cols2[i % 4].metric(label, _money(v))

            # ===== 總計（新版 Notion）=====
            st.markdown("### 🧾 總計")
            c1, c2, c3 = st.columns(3)
            c1.metric("薪資總計", _money(existing.get("薪資總計", 0)))
            c2.metric("應扣總計", _money(existing.get("應扣總計", 0)))
            c3.metric("實發金額", _money(existing.get("實發金額", 0)))

            st.write("")
            st.write(f"備註：{existing.get('備註', '') or '—'}")

            # （保留你原本的 12 月特休折算提示：這不是 Notion 薪資欄位，但很多人會需要）
            is_december = (int(m) == 12)
            if is_december:
                snap = get_employee_vacation_snapshot(current_user, int(y))
                cashout_amount = 0.0
                cashout_days = 0.0
                if snap:
                    cash = calc_cashout(
                        remaining_hours=snap["remaining"],
                        hours_per_day=st.session_state["hours_per_day"],
                        cap_days=st.session_state["cashout_cap_days"],
                        amount_per_day=st.session_state["cashout_amount_per_day"],
                        whole_days_only=st.session_state["cashout_whole_days_only"],
                    )
                    cashout_amount = cash["cashout_amount"]
                    cashout_days = cash["cashout_days"]

                st.info(f"💱 年底特休折算：可折算 {cashout_days:.0f} 天，折算金額 ${cashout_amount:.0f}")
                st.caption("※ 特休折算僅於 12 月顯示，避免其他月份造成誤會。")

    # -------------------------
    # ✅ 管理員：操作記錄表
    # -------------------------
    elif menu == "📋 操作記錄" and is_admin:
        st.header("操作記錄（管理員）")

        if not OPLOG_DB_ID:
            st.warning("⚠️ 你尚未在 .env 設定 OPLOG_DB_ID（操作記錄表 Database ID），所以目前無法讀取。")
            st.info("✅ 你已在 Notion 建好『操作結果』(成功/失敗/系統錯誤) 就對了，接下來只要把 DB ID 填進 .env 即可。")
        else:
            # ✅ 員工清單：跟「帳號管理表」同步（含管理員）
            employees = list_employee_names()

            # ✅ 四個篩選：顯示筆數 / 操作結果 / 員工 / 關鍵字（關鍵字放最右）
            colA, colB, colC, colD = st.columns([1, 1, 1, 2])
            with colA:
                limit = st.number_input("顯示筆數", min_value=10, max_value=200, value=100, step=10)
            with colB:
                result_filter = st.selectbox("操作結果", ["全部", "成功", "失敗", "系統錯誤"], index=0)
            with colC:
                emp_filter = st.selectbox("員工", ["全部"] + employees, index=0)
            with colD:
                keyword = st.text_input("關鍵字", value="")

            logs = list_operation_logs(limit=int(limit))

            # 1) 操作結果
            if result_filter != "全部":
                logs = [r for r in logs if (r.get("操作結果") == result_filter)]

            # 2) 員工（精準篩選：只看該員工的紀錄）
            if emp_filter != "全部":
                logs = [r for r in logs if (r.get("員工姓名", "") == emp_filter)]

            # 3) 關鍵字（可單用；若已選員工，則是在該員工範圍內再做關鍵字搜尋）
            if keyword.strip():
                k = keyword.strip()
                logs = [
                    r for r in logs
                    if (k in (r.get("員工姓名", "") or ""))
                    or (k in (r.get("操作內容", "") or ""))
                    or (k in (r.get("操作類型", "") or ""))
                    or (k in (r.get("操作結果", "") or ""))
                ]

            if logs:
                st.dataframe(logs, use_container_width=True)
            else:
                st.info("目前沒有符合條件的操作記錄。")



    # -------------------------
    # 管理員：系統設定
    # -------------------------
    elif menu == "⚙️ 系統設定":
        st.header("系統設定")

        if is_admin:
            st.subheader("🔑 管理員：重設員工密碼（忘記密碼用）")
            employees = list_employee_names()
            target = st.selectbox("選擇員工", employees)

            temp_pwd = st.text_input("設定臨時密碼（先給員工登入用）", type="password")
            if st.button("✅ 重設密碼", use_container_width=True):
                ok = admin_reset_user_password(target, temp_pwd, actor=current_user)
                if ok:
                    st.success("✅ 已重設！員工下次登入會被強制更改密碼。")
                    st.info("⚠️ 規則：Notion 會只保留『密碼』，並清空『login_hash』，避免同時存在。")
                    st.rerun()

        else:
            st.write("（建置中...）")

# =========================
# ✅ Global footer
# =========================
render_footer()
